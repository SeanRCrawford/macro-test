"""Search the Pokemon pool for answers to specific threats.

    python counter_table.py --vs "Kingambit,Basculegion"
    python counter_table.py --vs "Kingambit,Basculegion" --threshold 100
    python counter_table.py --vs "Kingambit" --chip-from "Ninetales-Alola" --chip-move "Blizzard"
    python counter_table.py --vs "Kingambit,Basculegion,Garchomp" --pairs
    python counter_table.py --vs "Kingambit,Basculegion" --pairs --chip-from "Ninetales-Alola" --chip-move "Blizzard"
    python counter_table.py --vs "Kingambit,Basculegion" --top 15
    python counter_table.py --vs "Kingambit,Basculegion,Urshifu-Rapid-Strike,Chien-Pao,Landorus-Therian,Rillaboom" --team "Big 6" --item "Gallade=Choice Scarf"
    python counter_table.py --vs "Kingambit,Basculegion,Mega Floette" --speed
    python counter_table.py --vs "Kingambit,Basculegion" --threshold 100 --max-taken 50 --outspeed natural
    python counter_table.py --vs "Kingambit,Basculegion" --threshold 50 --max-taken 33 --outspeed scarf
    python counter_table.py --vs "Kingambit,Basculegion,Garchomp" --joint --partner "Whimsicott"
    python counter_table.py --vs "Kingambit,Basculegion,Garchomp" --joint --partner "Whimsicott" --turns 3
    python counter_table.py --vs "Kingambit,Basculegion" --joint --pool-size 30
    python counter_table.py --our "Mega Scizor,Ninetales-Alola" --deep --vs "Basculegion,Mega Charizard Y,Mega Floette,Garchomp,Kingambit,Whimsicott"
    python counter_table.py --our "Mega Scizor,Ninetales-Alola" --deep --switches --vs "Basculegion,Mega Charizard Y,Mega Floette,Garchomp,Kingambit,Whimsicott" --pool-size 60
    python counter_table.py --our "Mega Scizor,Ninetales-Alola" --deep --switches --bench "Gyarados,Arcanine-Hisui" --vs "Basculegion,Mega Charizard Y"
    python counter_table.py --our "Garchomp,Incineroar,Gallade,Hydreigon,Whimsicott,Kingambit" --bring4 --vs "Kingambit,Basculegion,Whimsicott,Sinistcha,Mega Charizard Y,Sylveon"
    python counter_table.py --our "Garchomp,Incineroar,Gallade,Hydreigon,Whimsicott,Kingambit" --bring4 --vs-team "Rain" --xlsx bring4.xlsx
    python counter_table.py --multi-bring4 --vs-team "Kingambit,Basculegion,Whimsicott,Sinistcha,Mega Charizard Y,Sylveon" --vs-team "Garchomp,Landorus-Therian,Rillaboom,Chien-Pao,Urshifu-Rapid-Strike,Farigiraf" --pool-size 60
    python counter_table.py --multi-bring4 --vs-team "..." --vs-team "..." --max-weak 2 --type-limit "Fire:max_weak=1,max_net=-2" --allow-scarf
    python counter_table.py --multi-bring4 --vs-team "Rain" --vs-team "Big 6" --vs-team "NAIC" --pool-size 60
    python counter_table.py --multi-bring4 --vs-team "Rain" --vs-team "Big 6" --xlsx multi.xlsx --deep-dive-core "1,3,5"

Eight modes, pick one (or combine --chip-from/--chip-move with --pairs):

  (default)   OHKO / threshold search: best legal item, WORST-roll % on EACH
              named target (a guarantee), ranked on the worst of them.
  --chip-from / --chip-move
              Who finishes the named targets off (worst roll) after a
              partner's named move has already landed? Spread moves (e.g.
              Blizzard) take the doubles 0.75x multi-target penalty.
  --pairs     Who KOes one of every PAIR drawn from the named targets BEFORE
              the pair KOes it -- one full turn, priority then speed order,
              AVERAGE rolls (a realistic line, not a guarantee). Add
              --chip-from/--chip-move to give it a partner's help; every
              combination of who the candidate and the partner go for is
              tried and the best is kept.
  --speed     Instead of a damage search, a SPEED-TIER chart: `--vs`'s
              targets plus the pool, in real turn order (priority bracket
              first, then effective speed under each one's own best legal
              item -- Choice Scarf and a weather-boosting ability both
              apply). "to have an option to make sure my guys (accounting
              for priority like bullet punch) outspeed their enemies" --
              read down the list; anyone above a target in the SAME
              priority bracket outspeeds it, and a higher bracket outspeeds
              it regardless of the speed numbers.
  --joint     A JOINT pair table: does OUR PAIR (both attacking with their
              own real set, not one fixed move) beat every PAIR drawn from
              the named targets, over --turns turns (default 2), in real
              priority-then-speed order? With --partner, fixes one half and
              searches the pool for the other; WITHOUT --partner, GENERATES
              the whole pair -- every legal pair from the pool. Classified as
              a clean SWEEP (both enemies dead before either of them ever
              acted), an OUT-TRADE win (both enemies dead within the window,
              ours took hits but didn't faint), a LOSS, or NO-KO (window
              elapsed, nobody finished). Also replays the same race with the
              enemy side's speed doubled (a Tailwind hypothesis, same spirit
              as --outspeed scarf) and reports whether the pair is still
              safe. Every printed line shows the actual damage roll each hit
              did, both directions -- not just the win/loss classification.
  --deep      A DEEP DIVE on --our's specific, already-chosen pair -- no
              search. Every enemy pair drawn from --vs shown in FULL (pass a
              whole 6-Pokemon roster for "all 15 possible enemy leads"):
              which of theirs is an outright OHKO RISK on ours (checked on
              their best roll, a structural fact about the matchup -- "Scizor
              is always OHKO'd by Mega Charizard Y" -- not contingent on
              which line the race happens to play out), the full 2x2 damage
              grid both directions (not just whichever move the race chose),
              and the turn-by-turn line each matchup collapses into. Losses
              and OHKO risks sort first, since those are what's actionable.
              Add --switches to also check, for every LOSING enemy pair,
              whether swapping in a --bench candidate (or the whole pool)
              for one of ours turns it around -- one turn spent on the
              switch (the incoming Pokemon can't act that turn, same as a
              real doubles switch), ranked by least damage taken switching
              in among candidates that actually fix the loss.
  --bring4    For an ALREADY-DECIDED team (--our, 3, 4, 5, or 6 names) vs one
              enemy roster (--vs): every one of its C(len,2) pairs (the
              same table --joint's own pool search prints), then every one
              of its C(len,min(4,len)) possible BRING subsets, ranked by how
              bad its WORST internal pair does (maximin -- "I always have
              options no matter what position I am in", not just one great
              pair propping up a hole behind it). Exactly 3 or 4 names is
              not a special case -- it degenerates to the single bring that
              IS your team, still summarised by its own internal pairs.
              --good-threshold sets the bar (default 100%% -- must beat
              every enemy pair drawn from --vs) for what counts as a
              "good" pair when counting how many of a bring's pairs clear it.
  --multi-bring4
              Instead of one already-decided 6, SEARCH the pool for the
              best CORE against one enemy roster, or several at once
              (--vs-team, repeated 1+ times -- either a comma-separated
              roster or the name of a saved team from data/teams.csv/
              data/teams/data/my_teams, the same library --team already
              searches):
              for each enemy, the best bring-4 this
              core can field against it (a bring-4 may differ per
              opponent, real Team Preview); ranked on the WORST of those
              per-enemy best cases (maximin, same idea as --bring4's own
              Stage 2, generalised across enemies). A core is 4, 5, OR 6
              Pokemon, not always 6 -- "a full team of only 4-5 members is
              not a problem, in some ways it is actually better and more
              efficient" -- and any core with a member that never actually
              gets brought against any named enemy is dropped outright (the
              smaller core without that dead weight is already shown on its
              own). Each printed core also shows its type-weakness synergy
              (member_weakness_summary: how many members are weak to 2+
              types / exactly 1 / none) and a TEAMSHEET -- each bring-4
              member's own best item/moveset for that specific enemy, with
              every move's roster-wide usage%%. --max-weak/--type-limit add
              HARD synergy limits (a core that breaks one is never shown,
              however well it wins its bring-4s) -- e.g. --max-weak 2 for
              "only 2 members may be weak to any one type", or
              --type-limit "Fire:max_weak=1,max_net=-2" for "only 1 member
              weak to Fire, Fire must have net 2 resistances".
              --max-weak-types adds a BREADTH cap on top -- e.g.
              --max-weak-types 3 for "no more than 3 different types may
              have 2+ weak members", distinct from --max-weak's own
              per-type ceiling (a core can pass --max-weak while still
              being broadly fragile across many types at once). A core is
              also always capped at --max-megas (default 2) Mega-capable
              members -- "a full team can only have two mega stone users"
              -- distinct from the per-BATTLE mega-vs-stay-base choice each
              pair's own minimax already searches. Exhaustive
              by default over a candidate pool narrowed by --min-enemies/
              --good-threshold; --beam searches the whole pool instead when
              that's too large to sweep exhaustively.

By default the WHOLE ~270-Pokemon dataset is searched, not a pre-narrowed
"generically good" subset -- see `_pool` below for why (short version: "why
does Mega Scizor not show up" used to have an answer, and it was a bug in
which Pokemon ever got asked, not in the damage/item/move search itself).

DEFAULT MODE ONLY: --max-taken and --outspeed add hard requirements on top
of --threshold's KO bar, all three composed by AND -- a row failing any of
them is dropped, not just flagged.

    --max-taken 50           every named target's best roll against this
                              Pokemon must do under 50%
    --outspeed natural       must out-speed every named target under its
                              own chosen item
    --outspeed scarf         ... or would, if it held Choice Scarf instead
                              (a hypothesis; pin --item yourself if you want
                              that reflected in the actual damage numbers)

"my attackers in counter_table must either be faster than the enemy, able
to be faster with choice scarf, and/or take max X damage from the enemy's
best attack (e.g., OHKO all, take less than 50%, outspeed. or 2HKO all,
take less than 33%, outspeed.)" -- those two examples are exactly the last
two commands above (2HKO all == --threshold 50, the same worst-roll->=50%
guarantee --threshold already means).

By default every pool member's item AND moveset are searched for the best
legal answer to `--vs` -- "or to just select optimal item" is the default,
not something you have to ask for. To instead PIN a specific item on a named
Pokemon (e.g. "Choice Scarf on Gallade beats Big 6 easily" -- go verify
that claim directly), use --item:

    --item "Gallade=Choice Scarf"
    --item "Gallade=Choice Scarf;Ninetales-Alola=Life Orb"

A pinned item does not freeze the moveset to some hardcoded default -- "For
Choice Scarf, a pokemon can use 4 moves, which could be highly useful": the
4 moves are still genuinely re-optimised FOR that item, same search as
always, just with the item fixed instead of searched. Pin the moves too
(skipping move search entirely) with --moves, semicolon-separated by
Pokemon and comma-separated within:

    --moves "Gallade=Psycho Cut,Sacred Sword,Close Combat,Ice Punch"

--partner-item pins the --chip-from/--pairs partner's item the same way (it
defaults to the partner's own best legal item otherwise). Pins are checked
against Regulation MB's banned list (Assault Vest, Choice Band, Choice
Specs) the same as a searched item would be -- a pin is a decision, not a
loophole around the ban.

Choice Scarf is legal in Regulation MB, but by default every item SEARCH
(everywhere above -- not a pin) excludes it anyway: "it is too easy to
punish" as a recommendation, even though it's a real option. Pass
--allow-scarf to let the search consider it again. An explicit --item pin
(e.g. "Gallade=Choice Scarf") already bypasses this exclusion regardless --
same "a pin is a decision, not a loophole" rule as the Regulation MB ban.

Every damage number printed is the FULL roll (worst-average-best%). A
"Mega X" name evaluated ALONE (the default search, --chip-from/--chip-move)
always means the mega form, stats and all; inside a real PAIR (--pairs'
partner, --joint, --deep, --switches), only one member of a side may
actually Mega Evolve -- every legal assignment (including "neither does",
which keeps a Mega pick's own base ability and typing, e.g. Gyarados's
Intimidate and Water/Flying instead of Mold Breaker and Water/Dark) is
searched, ours for the best result and the named enemy pair's for the worst
-- see src/counter_finder.py for exactly what that means and what this
deliberately does not model (no Tailwind/screens/redirection outside the
--joint family, no whole-game item effects, no Intimidate stat-drop
simulation -- a hypothesis for the detailed lead-race tools, not a verdict).
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "..", "src"))
import blas_limits  # noqa: E402,F401

import argparse  # noqa: E402

import _harness  # noqa: E402,F401

from counter_finder import (DEFAULT_EXCLUDED_ITEMS, _answer_for,  # noqa: E402
                            _fixed_sets_from_pair_rows, _pair_sort_key,
                            bring4_pair_depth, bring4_search,
                            chip_then_ko, core_deep_dive, deep_dive,
                            enemy_has_real_tailwind, joint_pair_search,
                            joint_pool_search, member_weakness_summary,
                            multi_bring4_beam, multi_bring4_coverage,
                            multi_bring4_exhaustive, own_pair_has_real_tailwind,
                            pair_search, recommended_lead, speed_tiers,
                            switch_in_search, threshold_search)


def _parse_item_overrides(spec):
    """"Gallade=Choice Scarf;Ninetales-Alola=Life Orb" -> {name: item}."""
    if not spec:
        return None
    out = {}
    for entry in spec.split(";"):
        entry = entry.strip()
        if not entry:
            continue
        if "=" not in entry:
            raise SystemExit(f"--item entry {entry!r} must be 'Pokemon=Item'")
        name, item = entry.split("=", 1)
        out[name.strip()] = item.strip()
    return out


def _parse_type_limits(entries):
    """["Fire:max_weak=1,max_net=-2", "Ice:max_weak=0", ...] -> {"Fire":
    {"max_weak": 1, "max_net": -2}, "Ice": {"max_weak": 0}} -- --type-limit's
    format, one flag per type, repeatable. Feeds `_effective_type_limits`/
    `team_search.hard_violations` (via `multi_bring4_exhaustive`/
    `multi_bring4_beam`) the same {type: {"max_weak", "max_net"}} shape
    those already expect.
    """
    if not entries:
        return None
    from species_data import TYPES
    out = {}
    for entry in entries:
        if ":" not in entry:
            raise SystemExit(f"--type-limit entry {entry!r} must be "
                             f"'Type:max_weak=N,max_net=M'")
        type_name, rest = entry.split(":", 1)
        type_name = type_name.strip()
        if type_name not in TYPES:
            raise SystemExit(f"--type-limit: unknown type {type_name!r}")
        limits = {}
        for kv in rest.split(","):
            kv = kv.strip()
            if not kv:
                continue
            if "=" not in kv:
                raise SystemExit(f"--type-limit entry {entry!r}: {kv!r} "
                                 f"must be 'key=value'")
            k, v = kv.split("=", 1)
            k = k.strip()
            if k not in ("max_weak", "max_net"):
                raise SystemExit(f"--type-limit: unknown key {k!r} "
                                 f"(max_weak/max_net)")
            try:
                limits[k] = int(v.strip())
            except ValueError:
                raise SystemExit(f"--type-limit: {k}={v!r} must be an integer")
        out[type_name] = limits
    return out


# "if I don't use the argument then none are applied" -- the default list
# only kicks in when --strict-weak-types is passed BARE (no value); omitted
# entirely, nothing here is used at all.
_STRICT_WEAK_TYPES_DEFAULT = ("Fire", "Water", "Fairy", "Steel", "Dark",
                              "Ghost", "Ground")


def _resolve_strict_weak_types(raw):
    """--strict-weak-types's own three-way argument: `None` (flag omitted)
    -> no types at all; `"__DEFAULT__"` (flag given bare, via `nargs="?",
    const="__DEFAULT__"`) -> `_STRICT_WEAK_TYPES_DEFAULT`; otherwise -> the
    given comma-separated list, validated against `species_data.TYPES`.
    Returns a list of type names (possibly empty).
    """
    if raw is None:
        return []
    if raw == "__DEFAULT__":
        return list(_STRICT_WEAK_TYPES_DEFAULT)
    from species_data import TYPES
    names = [t.strip() for t in raw.split(",") if t.strip()]
    unknown = [t for t in names if t not in TYPES]
    if unknown:
        raise SystemExit(f"--strict-weak-types: unknown type(s) "
                         f"{', '.join(unknown)}")
    return names


def _merge_strict_weak_types(type_limits, strict_types):
    """Fold `strict_types` into `type_limits` as `max_weak=1` each -- an
    explicit `--type-limit` entry for the SAME type wins outright (its own
    full {max_weak, max_net} dict replaces the shorthand's bare max_weak=1,
    "more specific/intentional always beats the shorthand"). `type_limits`
    may be `None` (no --type-limit given); returns `None` only when both
    inputs are empty, matching `_parse_type_limits`'s own "no limits at
    all" convention downstream code already expects.
    """
    if not strict_types:
        return type_limits
    merged = {t: {"max_weak": 1} for t in strict_types}
    if type_limits:
        merged.update(type_limits)
    return merged


def _parse_move_overrides(spec):
    """"Gallade=Psycho Cut,Sacred Sword;X=..." -> {name: [move, ...]}."""
    if not spec:
        return None
    out = {}
    for entry in spec.split(";"):
        entry = entry.strip()
        if not entry:
            continue
        if "=" not in entry:
            raise SystemExit(f"--moves entry {entry!r} must be 'Pokemon=Move,Move,...'")
        name, moves = entry.split("=", 1)
        out[name.strip()] = [m.strip() for m in moves.split(",") if m.strip()]
    return out


def _resolve_vs_team(raw, teams, merged):
    """One `--vs-team` value -> its resolved roster (list of names),
    validated. `raw` is EITHER the name of a saved team (a `teams.csv` row,
    or a pokepaste in `data/teams/`/`data/my_teams/` -- whatever
    `species_data.load_teams` already indexes under `teams`) OR a raw
    comma-separated Pokemon list -- shared by `--multi-bring4` (repeated)
    and `--bring4` (a single one, as an alternative to `--vs`)."""
    name = raw.strip()
    team = list(teams[name]) if name in teams else [
        n.strip() for n in raw.split(",") if n.strip()]
    if len(team) < 2:
        raise SystemExit(f"--vs-team {raw!r} needs at least 2 Pokemon to "
                         f"form any pairs")
    unknown_team = [n for n in team if n not in merged]
    if unknown_team:
        raise SystemExit(f"unknown Pokemon: {', '.join(unknown_team)}")
    return team


def _parse_deep_dive_core(spec):
    """"5,85,16" (also space- or mixed-separated) -> [5, 85, 16], the ranks
    to run `core_deep_dive` on -- "I can deep mode run for promising teams
    5, 85, and 16". Order is preserved as given; a repeated rank is
    dropped after its first occurrence (each rank becomes an xlsx sheet
    name -- 'Dive 5 Sets' twice would crash the workbook write, and diving
    the same core again gives nothing new anyway)."""
    if not spec:
        return []
    parts = [p for p in spec.replace(",", " ").split() if p]
    out = []
    for p in parts:
        try:
            n = int(p)
        except ValueError:
            raise SystemExit(f"--deep-dive-core: {p!r} is not an integer")
        if n < 1:
            raise SystemExit(f"--deep-dive-core: {n} must be >= 1 "
                             f"(1 = top result)")
        if n not in out:
            out.append(n)
    return out


def _prompt_deep_dive_ranks(n_rows, no_prompt):
    """After a --multi-bring4/--bring4 search, ask (once) which core/bring-4
    RANKS (1-based, over the FULL ranked list -- the same numbering the
    xlsx's '#' column and --deep-dive-core itself use) to deep dive --
    "there was a prompt after a multi-bring4 that lets me run deep-dive on
    named cores just generated ... deep mode run for promising teams 5, 85,
    and 16". Returns [] (no prompt shown at all) whenever there's nothing to
    pick from, `--no-prompt` was passed, or stdin isn't a real terminal --
    a scripted or piped run must never block waiting on input it can't get.
    """
    if no_prompt or n_rows < 1 or not sys.stdin.isatty():
        return []
    while True:
        try:
            raw = input(f"\nDeep dive which cores? (1-{n_rows}, e.g. "
                       f"5,85,16 -- blank to skip): ")
        except EOFError:
            return []
        if not raw.strip():
            return []
        try:
            ranks = _parse_deep_dive_core(raw)
        except SystemExit as e:
            print(e)
            continue
        bad = [r for r in ranks if r > n_rows]
        if bad:
            print(f"Only {n_rows} found -- out of range: "
                 f"{', '.join(str(b) for b in bad)}")
            continue
        return ranks


def _pool(args, merged):
    """Who gets searched. Default: EVERY Pokemon in the dataset.

    "Why does Mega Scizor not show up" -- the earlier default reused
    `generate_team.build_candidate_pool`, which ranks by roster.csv's generic
    Score for TEAM GENERATION (a balanced-team-of-6 concern) and truncates to
    the top N of that ranking. Mega Scizor's generic Score put it outside
    that top 40 even though it is a genuinely strong, correctly-calculated
    answer to a SPECIFIC trio (Close Combat/Knock Off/Bullet Punch clears
    all three) -- the search itself was never the problem, the pool it never
    got to run on was. A single-Pokemon threshold/chip/pairs search over the
    full ~270-entry dataset takes low single-digit seconds, so there is no
    real cost to searching everything by default; `--pool-size` still exists
    to explicitly narrow it (e.g. to that same generic-Score top N) if
    something ever needs to be faster.

    Whatever the source, `data/preferences.csv`'s Include/Exclude are then
    applied (`_apply_preferences`) -- "make sure preferences.csv is taken
    into account (includes, excludes) so that it reduces the pool of
    eligible mons."
    """
    if args.team:
        from _harness import load_world
        W = load_world()
        base = list(W["teams"][args.team])
    elif args.pool_size:
        from generate_team import build_candidate_pool
        base = list(build_candidate_pool(merged, top_n=args.pool_size))
    else:
        base = sorted(merged)
    return _apply_preferences(base, merged, verbose=True)


def _apply_preferences(pool, merged, verbose=False):
    """preferences.csv Include/Exclude, honoured against whatever pool was
    built (full dataset, --pool-size top-N, or a --team roster) -- the same
    "always honour it, regardless of pool source" rule
    `team_search.build_candidate_pool` documents for team generation.

    Exclude also covers Mega forms of an excluded base species (and vice
    versa) -- banning one form but not the other is almost never what's
    meant, same reasoning `build_candidate_pool` applies. Include guarantees
    a name is actually searched, adding it back in if a --pool-size cut or
    --team roster left it out -- the same "a pin only means something if
    it's actually searched" rule this file's --item/--moves overrides
    already follow (see `main`).
    """
    from species_data import load_preferences
    prefs = load_preferences()
    excluded = set(prefs["exclude"])
    for e in list(excluded):
        if e.startswith("Mega "):
            excluded.add(e[5:])
        else:
            excluded.update({f"Mega {e}", f"Mega {e} X", f"Mega {e} Y"})

    before = len(pool)
    kept = [n for n in pool if n not in excluded]
    dropped = before - len(kept)
    added = [n for n in prefs["include"]
            if n in merged and n not in excluded and n not in kept]
    out = kept + added

    if verbose:
        if dropped:
            print(f"preferences.csv Exclude drops {dropped} Pokemon: "
                 f"pool {before} -> {len(kept)}.")
        if added:
            print(f"preferences.csv Include adds {sorted(added)}: "
                 f"pool -> {len(out)}.")
    return out


def _roll(h):
    """"what the damage roll is": worst-avg-best, as %, one Hit."""
    return f"{h.lo * 100:.0f}-{h.avg * 100:.0f}-{h.hi * 100:.0f}%"


def _print_threshold(rows, targets, threshold, top, max_taken=None, outspeed=None):
    screened = max_taken is not None or outspeed is not None
    header = f"{'#':>3} {'Pokemon':20s} {'item':16s}"
    for t in targets:
        header += f" {t[:22]:>22s}"
    header += f" {'worst':>6s} {'all >= thr':>10s}"
    if screened:
        header += f" {'max taken':>10s} {'outspeeds all':>14s}"
    print(header)
    print("-" * len(header))
    for i, r in enumerate(rows[:top], start=1):
        line = f"{i:>3} {r['name'][:20]:20s} {(r['item'] or '-')[:16]:16s}"
        for t in targets:
            h = r["per_target"][t]
            line += f" {_roll(h):>22s}"
        line += f" {r['worst_pct'] * 100:>5.0f}% {('YES' if r['meets_all'] else ''):>10s}"
        if screened:
            worst_taken = max((h.hi for h in r["incoming"].values()), default=0.0)
            outsped = (r["outspeeds_scarf"] if outspeed == "scarf"
                      else r["outspeeds"])
            line += f" {worst_taken * 100:>9.0f}% {('YES' if all(outsped.values()) else ''):>14s}"
        print(line)
    print()
    print("Each cell is the damage ROLL (worst-average-best %) that Pokemon's")
    print("best move (from its own optimised set, under its best LEGAL item for")
    print("this matchup) does to that target at full HP -- the WORST end is what")
    print("counts for the verdict (a guarantee, not a hope). 'worst' is the")
    print("worst-roll minimum across targets -- ranked on it, since answering")
    print("one target and whiffing the other is not the answer that was asked")
    print("for. A 'Mega X' name is always its mega stats.")
    if screened:
        print()
        print("Rows that failed --max-taken/--outspeed are not shown at all --")
        print("'max taken' is the worst of EACH named target's best roll against")
        print("this Pokemon (the most they could do, guaranteed-survival reading);")
        print("'outspeeds all' is against every named target " +
             ("naturally OR with Choice Scarf (a hypothesis)."
              if outspeed == "scarf" else "under its own chosen item."))
    for i, r in enumerate(rows[:top], start=1):
        bits = [f"{t}: {h.move_name or '-'} {_roll(h)}"
                for t, h in r["per_target"].items()]
        print(f"  {i:>3} {r['name']}: " + "; ".join(bits))
        if screened:
            in_bits = [f"{t}: {h.move_name or '-'} {_roll(h)}"
                      for t, h in r["incoming"].items()]
            print(f"      incoming: " + "; ".join(in_bits))


def _print_speed(rows, targets, top):
    print("Speed tiers -- real turn order: PRIORITY bracket first, then")
    print("effective speed (Choice Scarf / a weather-boosting ability already")
    print("applied) within the same bracket. No field: no Tailwind/Trick Room.")
    print("'role' marks the named --vs targets so you can read straight down")
    print("for anyone in your pool sitting above one in the same bracket.\n")
    header = f"{'#':>3} {'Pokemon':20s} {'item':16s} {'pri':>4s} {'move':16s} {'speed':>6s}  role"
    print(header)
    print("-" * len(header))
    shown = 0
    for r in rows:
        is_target = r["name"] in targets
        if not is_target:
            if shown >= top:
                continue
            shown += 1
        pri = f"+{r['priority']}" if r["priority"] > 0 else "-"
        num = "" if is_target else str(shown)
        print(f"{num:>3} {r['name'][:20]:20s} "
              f"{(r['item'] or '-')[:16]:16s} {pri:>4s} "
              f"{(r['priority_move'] or '-')[:16]:16s} {r['speed']:>6.1f}  "
              f"{'TARGET' if is_target else ''}")


def _print_chip(rows, targets, partner, move, top):
    print(f"Chip from {partner}'s {move}:")
    for t in targets:
        h = rows[0]["chip"][t] if rows else None
        if h:
            spread = " (spread, 0.75x)" if h.num_targets_hit > 1 else ""
            print(f"  {t}: {_roll(h)}{spread}")
    print()
    header = f"{'#':>3} {'Pokemon':20s} {'item':16s} {'KOs':>4s}"
    print(header)
    print("-" * len(header))
    for i, r in enumerate(rows[:top], start=1):
        print(f"{i:>3} {r['name'][:20]:20s} {(r['item'] or '-')[:16]:16s} "
              f"{r['n_ko']:>4d}/{len(targets)}")
    print()
    print("The chip is the PARTNER's damage roll, worst end used as the")
    print("guaranteed amount taken off before the finisher swings. Each row")
    print("below is the finisher's roll against what's left. A spread chip")
    print("move (Blizzard, ...) is shown with its doubles 0.75x noted -- it")
    print("assumes a real 2v2 field even if only one target was named.")
    for i, r in enumerate(rows[:top], start=1):
        bits = []
        for t, (ko, h) in r["finishes"].items():
            tag = "KO" if ko else _roll(h)
            bits.append(f"{t}: {tag} ({h.move_name or '-'})")
        print(f"  {i:>3} {r['name']}: " + "; ".join(bits))


_RANK = {"clean": 0, "trade": 1, "no_ko": 2, "pinned": 3}


def _print_pairs(rows, targets, top, partner=None, move=None):
    total = rows[0]["pairs_total"] if rows else 0
    if partner:
        print(f"With help from {partner}'s {move} (average roll):\n")
    header = (f"{'#':>3} {'Pokemon':20s} {'item':16s} {'beaten':>7s} "
              f"{'clean':>6s} {'trade':>6s} {'no KO':>6s} {'pinned':>7s}")
    print(header)
    print("-" * len(header))
    for i, r in enumerate(rows[:top], start=1):
        beaten = r["pairs_clean"] + r["pairs_trade"]
        print(f"{i:>3} {r['name'][:20]:20s} {(r['item'] or '-')[:16]:16s} "
              f"{beaten:>4d}/{total:<2d} {r['pairs_clean']:>3d}/{total:<2d} "
              f"{r['pairs_trade']:>3d}/{total:<2d} {r['pairs_no_ko']:>3d}/{total:<2d} "
              f"{r['pairs_pinned']:>4d}/{total:<2d}")
    print()
    print("One full turn per pair, priority THEN speed order, average rolls --")
    print("does this Pokemon KO one of the pair before the pair KOs it? Tried")
    print("against every (candidate target, partner target) combination; the")
    print("better outcome is kept. A spread move (candidate's, or the")
    print("partner's) hits BOTH pair members at once, 0.75x each. Only one")
    print("Mega Evolution per side is legal -- every legal assignment (ours")
    print("for the best result, theirs for the worst) is searched, including")
    print("'neither does' (keeps a Mega pick's own base ability/typing).")
    print("clean   it KOs one of them and survives the rest of the turn.")
    print("trade   it KOs one of them but is also KO'd, later the same turn,")
    print("        by the pair member it wasn't aimed at -- still 'KO'd them")
    print("        before being KO'd', just not for free.")
    print("no KO   it acts (isn't removed before its own move fires) but")
    print("        nothing in the pair goes down.")
    print("pinned  removed before it ever gets to act.")
    print("beaten  clean + trade -- pairs where the stated condition holds.")
    for i, r in enumerate(rows[:top], start=1):
        # WORST first -- _RANK ranks clean(0) best and pinned(3) worst, so a
        # plain ascending sort shows the 3 BEST matchups under a variable
        # named "worst" (the opposite of what a risk-scrutiny preview should
        # show, and the opposite of what `_print_deep`'s own detail listing
        # already promises: "LOSSES AND OHKO RISKS FIRST, since those are
        # what's actionable"). Sort DESCENDING so the riskiest 3 print here.
        worst = sorted(r["detail"].items(), key=lambda kv: _RANK[kv[1]["outcome"]],
                       reverse=True)
        print(f"  {i:>3} {r['name']}:")
        for (e1, e2), d in worst[:3]:
            print(f"      {e1} + {e2}: {d['outcome']}"
                 + (f" -> {d['target']}" if d["outcome"] in ("clean", "trade") else ""))
            for role, hits in d["hits"].items():
                for tgt, h in hits.items():
                    spread = " (spread)" if h.num_targets_hit > 1 else ""
                    print(f"          {role} -> {tgt}: {h.move_name or '-'} "
                         f"{_roll(h)}{spread}")


_JOINT_RANK = {"sweep": 0, "out_trade": 1, "no_ko": 2, "loss": 3}


def _row_pair(r, partner):
    """(c_name, p_name, items_str) for either row shape: `joint_pair_search`
    (`{"name", "item"}`, `partner` fixed by the caller) or `joint_pool_search`
    (`{"pair", "item1", "item2"}`, both slots searched)."""
    if "pair" in r:
        n1, n2 = r["pair"]
        return n1, n2, f"{r['item1'] or '-'} / {r['item2'] or '-'}"
    return r["name"], partner, r["item"] or "-"


def _print_joint(rows, targets, top, partner, turns):
    total = rows[0]["pairs_total"] if rows else 0
    if partner:
        print(f"Paired with {partner} (both attacking with their own real "
              f"set, {turns} turns, average rolls):\n")
    else:
        print(f"Generating the pair -- both slots searched from the pool, "
              f"{turns} turns, average rolls:\n")
    header = (f"{'#':>3} {'Pair':34s} {'items':30s} {'beaten':>7s} "
              f"{'swept':>6s} {'traded':>7s} {'lost':>5s} {'no KO':>6s} "
              f"{'clean':>10s} "
              f"{'tw-safe':>8s} {'pr-safe':>8s}")
    print(header)
    print("-" * len(header))
    for i, r in enumerate(rows[:top], start=1):
        c_name, p_name, items = _row_pair(r, partner)
        pair_str = f"{c_name} + {p_name}"
        beaten = r["pairs_swept"] + r["pairs_traded"]
        clean_max = 2 * total
        print(f"{i:>3} {pair_str[:34]:34s} {items[:30]:30s} "
              f"{beaten:>4d}/{total:<2d} {r['pairs_swept']:>3d}/{total:<2d} "
              f"{r['pairs_traded']:>4d}/{total:<2d} {r['pairs_lost']:>2d}/{total:<2d} "
              f"{r['pairs_no_ko']:>3d}/{total:<2d} "
              f"{r['pairs_clean_win_total']:>5.1f}/{clean_max:<3d} "
              f"{r['pairs_tailwind_safe']:>5d}/{total:<2d} "
              f"{r['pairs_protect_safe']:>5d}/{total:<2d}")
    print()
    print("Up to `turns` turns, priority THEN speed order, average rolls -- both")
    print("of ours attack with their own real optimised set (not one fixed")
    print("move). A spread move (Eruption, Heat Wave, ...) hits both of the")
    print("enemy pair at once, 0.75x each. Only one Mega Evolution per side")
    print("is legal -- every legal assignment (ours for the best result,")
    print("theirs for the worst) is searched, including 'neither does'")
    print("(keeps a Mega pick's own base ability/typing).")
    print("swept      both of the pair are dead before EITHER of them ever")
    print("           got to act -- outsped and KO'd before they could move.")
    print("out-trade  both of the pair die within the turn window, but ours")
    print("           took a hit along the way (still won the race).")
    print("lost       one of ours fainted before both of the pair did.")
    print("no KO      the turn window elapsed with neither side finished.")
    print("beaten     swept + traded -- pairs where the joint fight is won.")
    print("clean      OUR OWN retained HP across every enemy pair (0-2.0")
    print("           each, C+P), summed -- \"losing 1 pokemon and taking a")
    print("           lot of damage and KOing 2 enemies [is] far inferior to")
    print("           KOing the enemy without taking damage\": a swept pair")
    print("           always scores the max here (zero damage taken, by")
    print("           definition); a messy out-trade scores less, the more")
    print("           it cost. Two equally-beaten pairs are ranked by THIS")
    print("           when protect-safe and raw beaten count both tie.")
    print("tw-safe    the SAME race, replayed with the enemy pair's speed")
    print("           doubled (a Tailwind hypothesis) -- still swept or")
    print("           traded, not lost or no-KO'd, once they move first.")
    print("pr-safe    the SAME race, replayed twice more -- once with EACH")
    print("           enemy Protecting turn 1 instead of attacking -- still")
    print("           swept or traded both times. A turn-1 scouting Protect")
    print("           is the classic doubles 50/50 (e.g. the second enemy")
    print("           protects, the first still KOs one of ours, then the")
    print("           protector cleans up next turn) -- this flags it rather")
    print("           than hiding it behind the no-Protect line of play.")
    print()
    print("DAMAGE, both directions: each line is the roll (worst-average-best")
    print("%) that hit did, read against whatever HP the target had left AT")
    print("THAT POINT in the turn -- a second hit on an already-damaged target")
    print("is a % of what remained, not of full HP.")
    for i, r in enumerate(rows[:top], start=1):
        c_name, p_name, _items = _row_pair(r, partner)
        role_name = {"C": c_name, "P": p_name}
        # WORST first -- see the matching note in `_print_pairs`: _JOINT_RANK
        # ranks sweep(0) best and loss(3) worst, so this must sort DESCENDING
        # to actually show the riskiest 3 matchups here, not the safest 3.
        worst = sorted(r["detail"].items(),
                       key=lambda kv: _JOINT_RANK[kv[1]["outcome"]], reverse=True)
        print(f"  {i:>3} {c_name} + {p_name}:")
        for (e1, e2), d in worst[:3]:
            role_name["E1"], role_name["E2"] = e1, e2
            if d.get("tailwind_forced"):
                tw = (f"  [REAL TAILWIND THREAT: {e1}/{e2} can actually set "
                     f"it -- assumed, result is {d['tailwind_outcome']} "
                     f"(would be {d['outcome_without_tailwind']} without it)]")
            elif not d["tailwind_safe"]:
                tw = f"  [tailwind: {d['tailwind_outcome']}]"
            else:
                tw = ""
            pr = "" if d["protect_safe"] else (
                f"  [protect: {e1} protects->{d['protect_outcomes']['E1']}, "
                f"{e2} protects->{d['protect_outcomes']['E2']}]")
            print(f"      {e1} + {e2}: {d['outcome']} "
                 f"(turn {d['turns_used']}){tw}{pr}")
            for turn_i, turn_hits in enumerate(d["log"], 1):
                for role, tgt_role, h in turn_hits:
                    spread = " (spread)" if h.num_targets_hit > 1 else ""
                    print(f"          T{turn_i} {role_name[role]} -> "
                         f"{role_name[tgt_role]}: {h.move_name or '-'} "
                         f"{_roll(h)}{spread}")


def _print_deep(name1, name2, item1, item2, targets, detail, summary, turns):
    """One named pair, every enemy pair drawn from `targets` in full --
    OHKO risk, the 2x2 damage grid both directions, and the turn-by-turn
    line. Deliberately shows ALL of them (not top-N): "a deep dive" means
    seeing every one, not a ranked shortlist -- but LOSSES AND OHKO RISKS
    FIRST, since those are what's actionable.
    """
    total = summary["pairs_total"]
    beaten = summary["pairs_swept"] + summary["pairs_traded"]
    print(f"Deep dive: {name1} ({item1 or '-'}) + {name2} ({item2 or '-'})")
    print(f"vs {total} enemy pair(s) drawn from: {', '.join(targets)}\n")
    print(f"{beaten}/{total} beaten ({summary['pairs_swept']} swept, "
         f"{summary['pairs_traded']} traded), {summary['pairs_lost']} lost, "
         f"{summary['pairs_no_ko']} no-KO, "
         f"{summary['pairs_tailwind_safe']}/{total} tailwind-safe, "
         f"{summary['pairs_protect_safe']}/{total} protect-safe, "
         f"{summary['pairs_clean_win_total']:.1f}/{total * 2} clean win\n")

    def sort_key(kv):
        (_e1, _e2), d = kv
        # _JOINT_RANK ranks sweep(0) best and loss(3) worst -- negate it so
        # the ascending sort below actually puts losses (and, tied on
        # outcome, the most OHKO risks) FIRST, matching this function's own
        # stated "LOSSES AND OHKO RISKS FIRST" promise instead of the
        # opposite of it.
        return (-_JOINT_RANK[d["outcome"]], -len(d["ohko_risk"]))

    ordered = sorted(detail.items(), key=sort_key)
    role_name = {"C": name1, "P": name2}
    for (e1, e2), d in ordered:
        role_name["E1"], role_name["E2"] = e1, e2
        if d.get("tailwind_forced"):
            tw = (f"  [REAL TAILWIND THREAT: {e1}/{e2} can actually set it "
                 f"-- assumed, result is {d['tailwind_outcome']} (would be "
                 f"{d['outcome_without_tailwind']} without it)]")
        elif not d["tailwind_safe"]:
            tw = f"  [UNSAFE under tailwind: {d['tailwind_outcome']}]"
        else:
            tw = ""
        pr = "" if d["protect_safe"] else (
            f"  [UNSAFE if {e1} protects: {d['protect_outcomes']['E1']}; "
            f"if {e2} protects: {d['protect_outcomes']['E2']}]")
        print(f"  {e1} + {e2}: {d['outcome'].upper()} (turn {d['turns_used']}){tw}{pr}")
        for r in d["ohko_risk"]:
            print(f"      OHKO RISK: {role_name[r['attacker']]}'s {r['move']} "
                 f"could one-shot {role_name[r['target']]} "
                 f"(worst roll {r['hi'] * 100:.0f}%)")
        grid = d["grid"]
        print("      Damage we deal (average roll):")
        for (atk, tgt), h in grid["ours"].items():
            print(f"          {role_name[atk]} -> {role_name[tgt]}: "
                 f"{h.move_name or '-'} {_roll(h)}"
                 f"{' (spread)' if h.num_targets_hit > 1 else ''}")
        print("      Damage we take (average roll):")
        for (atk, tgt), h in grid["theirs"].items():
            print(f"          {role_name[atk]} -> {role_name[tgt]}: "
                 f"{h.move_name or '-'} {_roll(h)}"
                 f"{' (spread)' if h.num_targets_hit > 1 else ''}")
        print("      How it plays out:")
        for turn_i, turn_hits in enumerate(d["log"], 1):
            if not turn_hits:
                print(f"          T{turn_i}: nobody left to act")
                continue
            for role, tgt_role, h in turn_hits:
                spread = " (spread)" if h.num_targets_hit > 1 else ""
                print(f"          T{turn_i} {role_name[role]} -> "
                     f"{role_name[tgt_role]}: {h.move_name or '-'} "
                     f"{_roll(h)}{spread}")
        print()


def _print_bring4(pair_rows, bring4_rows, our6, targets, top, turns, good_threshold):
    """`bring4_search`'s own two-stage output: every pair drawn from
    `our6`, then every possible bring-4 subset ranked by how bad its WORST
    pair is (maximin -- "always have options no matter what position I am
    in")."""
    total = pair_rows[0]["pairs_total"] if pair_rows else 0
    print(f"Bring-4 robustness: {' / '.join(our6)}")
    print(f"vs {', '.join(targets)} ({turns} turns, average rolls, "
         f"good-pair bar {good_threshold * 100:.0f}% beaten)\n")

    print(f"Stage 1 -- all {len(pair_rows)} pairs drawn from your "
         f"{len(our6)}:")
    header = (f"  {'#':>3} {'Pair':34s} {'beaten':>7s} {'swept':>6s} "
             f"{'traded':>7s} {'lost':>5s} {'no KO':>6s} {'tw-safe':>8s} "
             f"{'pr-safe':>8s} {'own-tw':>7s}")
    print(header)
    print("  " + "-" * (len(header) - 2))
    for i, r in enumerate(pair_rows, start=1):
        pair_str = " + ".join(r["pair"])
        beaten = r["pairs_swept"] + r["pairs_traded"]
        print(f"  {i:>3} {pair_str[:34]:34s} {beaten:>4d}/{total:<2d} "
             f"{r['pairs_swept']:>3d}/{total:<2d} {r['pairs_traded']:>4d}/{total:<2d} "
             f"{r['pairs_lost']:>2d}/{total:<2d} {r['pairs_no_ko']:>3d}/{total:<2d} "
             f"{r['pairs_tailwind_safe']:>5d}/{total:<2d} "
             f"{r['pairs_protect_safe']:>5d}/{total:<2d} "
             f"{r['pairs_own_tailwind_used']:>4d}/{total:<2d}")

    n_pairs = bring4_rows[0]["pairs_total"] if bring4_rows else 6
    print(f"\nStage 2 -- all {len(bring4_rows)} possible bring-4s, ranked by "
         f"how many enemy pairs\nNONE of its {n_pairs} pairs can beat (fewest "
         f"first), then by how their WORST pair does\n(best worst-case "
         f"first), then by how many of its {n_pairs} pairs clear the "
         f"good-pair bar:")
    header2 = (f"  {'#':>3} {'Bring-4':46s} {'uncov':>6s} {'good':>6s} "
              f"{'worst pair':34s} {'worst beaten':>13s}")
    print(header2)
    print("  " + "-" * (len(header2) - 2))
    for i, b in enumerate(bring4_rows[:top], start=1):
        bring4_str = " / ".join(b["bring4"])
        worst_str = " + ".join(b["worst_pair"])
        wr = b["worst_pair_row"]
        uncov = len(b["uncovered_enemy_pairs"])
        print(f"  {i:>3} {bring4_str[:46]:46s} {uncov:>3d}/{total:<2d} "
             f"{b['pairs_good']:>3d}/{n_pairs:<2d}"
             f"{worst_str[:34]:34s} "
             f"{wr['pairs_swept'] + wr['pairs_traded']:>4d}/{total:<2d}")
        if b["uncovered_enemy_pairs"]:
            print("           no answer to: " + ", ".join(
                f"{e1}+{e2}" for e1, e2 in b["uncovered_enemy_pairs"]))
    print()
    print("beaten     swept + traded -- how many of the enemy pairs drawn")
    print("           from --vs this OUR pair actually beats.")
    print("own-tw     (Stage 1) how many of the enemy pairs this OUR pair")
    print("           only beats by choosing to open with ITS OWN real")
    print("           Tailwind (a setter + attacker hyper-offense answer) --")
    print("           0 for a pair with no real Tailwind setter of its own.")
    bring_size = len(bring4_rows[0]["bring4"]) if bring4_rows else 4
    print(f"uncov      (Stage 2) enemy pairs that NONE of this bring-4's "
         f"{n_pairs}")
    print(f"           pairs beat -- a real, unconditional loss whichever of")
    print(f"           the {bring_size} you're forced to send out. Ranked "
         f"on THIS")
    print("           FIRST: a bring-4 with one such pair loses to one that")
    print("           merely has a lower average, even at an equal beaten")
    print("           fraction -- \"having a pair that every pair of yours")
    print("           loses against is terrible.\"")
    print(f"good       (Stage 2) how many of this bring-4's {n_pairs} "
         f"internal pairs")
    print("           clear the good-pair bar above -- \"several perform")
    print("           very well\" rather than relying on just one.")
    print(f"worst pair (Stage 2) the WEAKEST of this bring-4's {n_pairs} "
         f"pairs, ranked")
    print("           protect-safe wins first, then beaten count -- the one")
    print("           you're stuck with if the game forces exactly that")
    print("           board state. Ranked on THIS, not the average: a")
    print("           bring-4 with one great pair and one awful one loses to")
    print("           a bring-4 that's merely good everywhere.")


def _print_pair_summary(coverage, top=10):
    """"I want to at least see the results (i.e., high performing pairs,
    results /15)" -- `multi_bring4_coverage`'s own Stage A already ran the
    full pool-wide pair search against EVERY named enemy (`per_enemy`), a
    cost that's linear-ish in pool size (C(pool,2) pairs, same as --joint),
    entirely independent of whether the LATER team-of-6 sweep
    (`multi_bring4_exhaustive`/`multi_bring4_beam`) can run at all -- so
    this always has something to show, even when the candidate pool is far
    too large for an exhaustive C(N,6) sweep. For each enemy, the top pairs
    by beaten fraction (swept+traded)/total -- "/15" for a full 6-Pokemon
    enemy roster (C(6,2)=15 pairs), whatever the real total is otherwise.
    """
    for i, (target_names, rows) in enumerate(
            zip(coverage["target_name_lists"], coverage["per_enemy"]), start=1):
        total = rows[0]["pairs_total"] if rows else 0
        print(f"  Enemy {i} ({', '.join(target_names)}) -- top pairs:")
        ranked = sorted(rows, key=_pair_sort_key)
        for r in ranked[:top]:
            beaten = r["pairs_swept"] + r["pairs_traded"]
            n1, n2 = r["pair"]
            print(f"      {n1} + {n2}: {beaten}/{total} beaten "
                 f"({r['pairs_swept']} swept, {r['pairs_traded']} traded)")
    print()


def _print_teamsheet_member(name, target_names, merged, moves_db, natures,
                            typechart, item_overrides, move_overrides,
                            excluded_items, indent="            ",
                            fixed_items=None, fixed_moves=None):
    """One member's chosen item/moves + each move's roster-wide USAGE%,
    against the specific enemy roster it's actually brought -- "Display
    teamsheets, i.e., moves for the teams and usage."

    `fixed_items`/`fixed_moves` ({name: ...}), when given (the
    `--multi-bring4` case, from `multi_bring4_coverage`'s own return): READ
    from there instead of re-deriving -- "for a team, the moves must stay
    the same, i.e., they can't be adjusted battle to battle", so the set
    SHOWN here must be the exact one the printed numbers next to it were
    actually computed from, not a fresh per-enemy `_answer_for` call that
    could legally come back different. Falls back to a fresh `_answer_for`
    call (the same search `threshold_search`/etc already run) only when
    neither is given (every other mode, which has no multi-enemy "must stay
    fixed" concern to begin with -- a single `target_names` list has
    nothing to be inconsistent WITH). Usage%% is read straight off
    `merged[name]["moves_usage"]`, 0.0 for a move that usage data never
    recorded (the same "any move in the game" escape hatch
    `solver.build_moveset` already allows).
    """
    if fixed_items is not None and name in fixed_items:
        item, move_names = fixed_items[name], fixed_moves.get(name)
    else:
        item, move_names, _weather = _answer_for(
            name, merged, moves_db, natures, typechart, target_names,
            item_overrides=item_overrides, move_overrides=move_overrides,
            excluded_items=excluded_items)
    usage_by_move = dict(merged[name].get("moves_usage") or [])
    moves_str = ", ".join(f"{m} ({usage_by_move.get(m, 0.0):.0f}%)"
                          for m in (move_names or []))
    print(f"{indent}{name} @ {item or '-'}: {moves_str or '-'}")


def _print_multi_bring4(rows, target_name_lists, top, mode_label, good_threshold,
                        candidate_pool_size, pool_size, merged, moves_db,
                        natures, typechart, item_overrides, move_overrides,
                        excluded_items, fixed_items=None, fixed_moves=None,
                        core_sizes=(4, 5, 6)):
    """`multi_bring4_exhaustive`/`multi_bring4_beam`'s own row shape --
    every printed CORE (`core_sizes`, default 4/5/6 Pokemon -- "a full team
    of only 4-5 members is not a problem, in some ways it is actually
    better and more efficient" -- `unused` is always empty by the time a
    row gets here, both
    search functions already drop any core with dead weight) shows its best
    bring-4 against EACH enemy (they can differ), with the BOTTLENECK enemy
    (the worst of those best cases -- what the core is actually ranked on)
    marked, a per-member weakness-count synergy line
    (`member_weakness_summary`), and a teamsheet (item + moves + usage%%)
    for every member of that enemy's bring-4."""
    print(f"Multi-bring4 search ({mode_label}): {candidate_pool_size} "
         f"candidate(s) from a {pool_size}-Pokemon search pool, "
         f"good-pair bar {good_threshold * 100:.0f}%\n")
    for i, targets in enumerate(target_name_lists, start=1):
        print(f"  Enemy {i}: {', '.join(targets)}")
    print()
    if not rows:
        print(f"No core ({', '.join(str(s) for s in core_sizes)} Pokemon) "
             f"found -- widen the pool, lower --good-threshold/--min-enemies, "
             f"relax --max-weak/--type-limit, or pass --beam for a broader "
             f"(non-exhaustive) search.\n")
        return
    for i, r in enumerate(rows[:top], start=1):
        core = r["core"]
        print(f"  {i:>3} ({r['core_size']}) {' / '.join(core)}")
        weak = member_weakness_summary(core, merged)
        types_2plus = sum(1 for c in weak["per_type"].values() if c >= 2)
        print(f"        synergy: weak to 2+ types: {weak['weak_to_2plus']}, "
             f"weak to 1: {weak['weak_to_1']}, weak to 0: {weak['weak_to_0']} "
             f"(members)")
        print(f"        types with 2+ weak members: {types_2plus}")
        print(f"        per-member weak-type counts: " + ", ".join(
            f"{n}={c}" for n, c in weak["per_member"].items()))
        by_type = sorted(((t, c) for t, c in weak["per_type"].items() if c > 0),
                         key=lambda tc: -tc[1])
        print("        weaknesses by type: " + (
            ", ".join(f"{t} {c}" for t, c in by_type) if by_type else "none"))
        core_fixed_items = fixed_items
        if r.get("item_clause_resolved_items"):
            core_fixed_items = {**(fixed_items or {}), **r["item_clause_resolved_items"]}
        for e_idx, pe in enumerate(r["per_enemy"], start=1):
            wr = pe["best_bring4_row"]["worst_pair_row"]
            total = wr["pairs_total"]
            beaten = wr["pairs_swept"] + wr["pairs_traded"]
            uncovered = pe["best_bring4_row"]["uncovered_enemy_pairs"]
            bottleneck = "  <-- bottleneck" if e_idx - 1 == r["worst_enemy_idx"] else ""
            print(f"        vs Enemy {e_idx}: bring "
                 f"{' / '.join(pe['best_bring4'])} "
                 f"(worst pair beats {beaten}/{total} of Enemy {e_idx}'s "
                 f"pairs){bottleneck}")
            if uncovered:
                print(f"          no answer to: " + ", ".join(
                    f"{e1}+{e2}" for e1, e2 in uncovered))
            for name in pe["best_bring4"]:
                _print_teamsheet_member(
                    name, pe["target_names"], merged, moves_db, natures,
                    typechart, item_overrides, move_overrides, excluded_items,
                    fixed_items=core_fixed_items, fixed_moves=fixed_moves)
        print()
    print("bottleneck the enemy this core is WEAKEST against, even using")
    print("           its own best available bring-4 there -- what the")
    print("           ranking is actually maximin'd on: a core that's")
    print("           spectacular vs 2 enemies and shaky vs the 3rd loses")
    print("           to one that's merely solid against all three. An")
    print("           enemy pair NONE of the core's pairs can beat")
    print("           ('no answer to') outranks everything else here --")
    print("           an equal beaten fraction with no unconditional loss")
    print("           is always preferred.")
    print("worst pair the WEAKEST of that bring-4's own internal pairs,")
    print("           ranked protect-safe wins first, then beaten count")
    print("           (same reading as --bring4's own 'worst pair') -- the")
    print("           bring-4 shown is the one with fewest enemy pairs it")
    print("           has no answer to, then whose worst pair is least")
    print("           bad, for that specific enemy.")
    print("synergy    per-CORE type-weakness counts (`--max-weak`/")
    print("           `--type-limit` hard-filter on this same data): how")
    print("           many of the core's members are weak to 2+ types, to")
    print("           exactly 1, or to none at all.")
    print("types with 2+ weak members")
    print("           how many DIFFERENT types have 2+ of the core's own")
    print("           members weak to them (`--max-weak-types` hard-filters")
    print("           on this) -- a BREADTH measure, distinct from the")
    print("           synergy line above's per-member counts.")
    print("teamsheet  each bring-4 member's own best legal item/moveset,")
    print("           searched ONCE against every named enemy and held")
    print("           fixed for all of them (a real team's set can't be")
    print("           adjusted battle to battle), with each move's")
    print("           roster-wide usage%% alongside it.")


def _row_line(row, label):
    """One `_ROW_TOTAL_FIELDS`-shaped dict as a single readable line --
    shared by every level `_print_core_deep_dive` reports at (per pair per
    enemy, per pair overall, whole-core overall)."""
    beaten = row["pairs_swept"] + row["pairs_traded"]
    total = row["pairs_total"]
    return (f"{label}: {beaten}/{total} beaten ({row['pairs_swept']} swept, "
           f"{row['pairs_traded']} traded, {row['pairs_lost']} lost, "
           f"{row['pairs_no_ko']} no-KO), {row['pairs_tailwind_safe']}/{total} "
           f"tw-safe, {row['pairs_protect_safe']}/{total} pr-safe, "
           f"{row['pairs_clean_win_total']:.1f}/{total * 2} clean win")


def _print_core_deep_dive(dive, top_gameplans=None):
    """`core_deep_dive`'s own return shape, printed in full -- "I want to
    see the full beaten/swept/traded/lost/no-KO/tw-safe/pr-safe for each
    pair, and then vs each enemy, how does each team perform on each one
    across each pair ... I also want to see the gameplans for each pair
    included in a team vs enemies." Deliberately verbose (this is the
    opt-in --deep-dive-core path, not the default per-core summary)."""
    core = dive["core"]
    print(f"\nDeep dive: {' / '.join(core)}\n")
    print("  set: " + ", ".join(f"{n} @ {s['item'] or '-'}"
                                for n, s in dive["sets"].items()))
    print("  " + _row_line(dive["overall"], "OVERALL, every pair vs every "
                           "enemy"))
    print()
    for (n1, n2), pair in dive["per_pair"].items():
        print(f"  {n1} + {n2}")
        print("    " + _row_line(pair["total"], "total, all enemies"))
        for pe in pair["per_enemy"]:
            print("    " + _row_line(pe["summary"],
                                     f"vs {', '.join(pe['target_names'])}"))
            role_name = {"C": n1, "P": n2}
            for (e1, e2), d in pe["detail"].items():
                role_name["E1"], role_name["E2"] = e1, e2
                tw = "" if d["tailwind_safe"] else f"  [tailwind: {d['tailwind_outcome']}]"
                pr = "" if d["protect_safe"] else (
                    f"  [protect: {e1}->{d['protect_outcomes']['E1']}, "
                    f"{e2}->{d['protect_outcomes']['E2']}]")
                print(f"        {e1} + {e2}: {d['outcome']} "
                     f"(turn {d['turns_used']}){tw}{pr}")
                for turn_i, turn_hits in enumerate(d["log"], 1):
                    for role, tgt_role, h in turn_hits:
                        spread = " (spread)" if h.num_targets_hit > 1 else ""
                        print(f"            T{turn_i} {role_name[role]} -> "
                             f"{role_name[tgt_role]}: {h.move_name or '-'} "
                             f"{_roll(h)}{spread}")
        print()


def _write_teamsheet_json(path, core_dive):
    """`core_dive["sets"]` (item + moveset, already fixed for the whole
    core -- see `core_deep_dive`) as the exact JSON shape the Streamlit
    app's Team Builder tab already reads/writes: `{"pool": [...], "sets":
    {name: {"item", "moves"}}}`.

        "I am primarily using the CLI for counter_table.py and then
         checking results in the streamlit app, so the export in the CLI
         needs to be able to export to the streamlit app, whether through
         a paste or otherwise"

    `path` of `"-"` prints the JSON to stdout instead, framed with clear
    markers so it can be copy-pasted straight into the app's paste box
    (Team Builder -> "...or paste a team .json") without also having to
    save/upload a file -- the CLI and the app are typically two different
    windows/machines for this workflow, so a plain stdout dump is the
    lowest-friction bridge between them.
    """
    import json
    payload = {"pool": list(core_dive["core"]),
              "sets": {n: {"item": s["item"], "moves": s["moves"]}
                       for n, s in core_dive["sets"].items()}}
    text = json.dumps(payload, indent=2)
    if path == "-":
        print("\n=== Teamsheet JSON -- paste into the Streamlit app's "
             "Team Builder tab ('...or paste a team .json') ===")
        print(text)
        print("=== end teamsheet JSON ===")
    else:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(text)
        print(f"\nTeamsheet JSON: {os.path.abspath(path)} -- upload it via "
             f"the Streamlit app's Team Builder tab (\"...or upload a team "
             f".json\"), or open the file and paste its contents into the "
             f"paste box there.")


def _fixed_sets_for(names, fixed_items, fixed_moves, merged, moves_db, natures,
                    typechart, target_names, item_overrides, move_overrides,
                    excluded_items):
    """{name: {"item", "moves"}} for `names`, read from the already-computed
    FIXED sets when present -- NEVER re-derived, "a real team's set can't be
    adjusted battle to battle" -- falling back to a fresh `_answer_for` call
    against `target_names` only for a name the fixed dicts don't cover
    (mirrors `_print_teamsheet_member`'s own fallback, same reason: not
    every xlsx export path has a precomputed fixed set to read, e.g.
    --bring4's, which searches only one team against one roster)."""
    out = {}
    for name in names:
        if name in fixed_items:
            out[name] = {"item": fixed_items[name],
                         "moves": fixed_moves.get(name) or []}
        else:
            item, move_names, _weather = _answer_for(
                name, merged, moves_db, natures, typechart, target_names,
                item_overrides=item_overrides, move_overrides=move_overrides,
                excluded_items=excluded_items)
            out[name] = {"item": item, "moves": move_names or []}
    return out


def _teamsheet_export_cells(pool, sets, merged):
    """(pool, sets) -> (pokepaste text, base64 teamsheet token) -- the two
    exportable forms an xlsx 'Teamsheets' sheet cell offers: "export any of
    the teamsheets generated ... maybe with a base64 encoded pokepaste in
    an excel cell". The token round-trips exactly (paste it into the
    Streamlit app's Team Builder); the pokepaste is the readable, portable
    (pokepast.es/Showdown) form of the same team."""
    from species_data import team_to_showdown_export
    from team_sheet import encode_teamsheet
    return team_to_showdown_export(pool, sets, merged), encode_teamsheet(pool, sets)


def _write_sets_sheet(wb, rows_by_rank, merged):
    """A 'Sets' sheet: one row per (rank, Pokemon) -- item, ability, nature,
    EVs, moves and each move's roster-wide usage%%, "I also need to see
    what the moves etc are" made visible for EVERY row of the main table,
    not just an opt-in deep dive. `rows_by_rank`: [(rank, core_or_bring4_label,
    {name: {"item", "moves"}}), ...] -- name order within each entry is the
    display order."""
    from export_excel import _autosize, _style_header
    from species_data import EV_LABEL, EV_ORDER, resolve_export_fields
    ws = wb.create_sheet("Sets")
    ws.append(["#", "Team", "Pokemon", "Item", "Ability", "Nature", "EVs",
              "Moves", "Move usage %"])
    _style_header(ws)
    for rank, label, sets in rows_by_rank:
        for name, s in sets.items():
            f = resolve_export_fields(name, s, merged)
            ev_str = " / ".join(f"{f['evs'][k]} {EV_LABEL[k]}" for k in EV_ORDER
                                if f["evs"].get(k)) or "-"
            usage_by_move = dict(merged.get(name, {}).get("moves_usage") or [])
            moves = s["moves"]
            ws.append([rank, label, name, f["item"] or "-", f["ability"],
                      f["nature"], ev_str, ", ".join(moves),
                      ", ".join(f"{usage_by_move.get(m, 0.0):.0f}%" for m in moves)])
    _autosize(ws)


def _write_teamsheets_sheet(wb, entries, merged):
    """A 'Teamsheets' sheet: one row per exportable unit -- "export any of
    the teamsheets generated in the CLI's .xlsx to the streamlit app".
    `entries`: [(rank, label, scope, members, sets), ...] where `sets` is
    {name: {"item", "moves"}} for exactly `members`."""
    from openpyxl.styles import Alignment
    from export_excel import _autosize, _style_header
    ws = wb.create_sheet("Teamsheets")
    ws.append(["#", "Team", "Scope", "Members", "Pokepaste", "Teamsheet (base64)"])
    _style_header(ws)
    for rank, label, scope, members, sets in entries:
        paste, token = _teamsheet_export_cells(members, sets, merged)
        ws.append([rank, label, scope, " / ".join(members), paste, token])
        ws.cell(row=ws.max_row, column=5).alignment = Alignment(wrap_text=True)
    _autosize(ws)
    # Pokepaste text is long and multi-line -- _autosize's plain "longest
    # cell text" measure would otherwise blow this column out to hundreds
    # of characters wide; a fixed, generous width (wrapped, so nothing is
    # lost) reads far better.
    ws.column_dimensions["E"].width = 45


def _loses_to(detail_dicts):
    """Enemy pairs this pair actually LOSES to, across one or more
    `pe["detail"]` mappings (`{(e1,e2): d, ...}`) -- the NAMES behind the
    "Lost" count column, same `outcome == "loss"` predicate, so the two
    columns always agree on count."""
    names = [f"{e1} + {e2}" for detail in detail_dicts
            for (e1, e2), d in detail.items() if d["outcome"] == "loss"]
    return "; ".join(names)


def _write_dive_sheets(wb, core_dives, _safe_sheet_name, _style_header, _autosize):
    """One 'Dive {rank} Sets'/'Summary'/'Gameplans' sheet trio per
    `(rank, dive)` in `core_dives` -- `core_deep_dive`'s own return shape,
    in full, "I also want to see the gameplans for each pair included in a
    team vs enemies". Shared by --bring4/--multi-bring4's xlsx writers so
    several dived cores/bring-4s from one run ("deep mode run for promising
    teams 5, 85, and 16") land in one workbook instead of colliding on a
    single fixed sheet name.

    The Summary and Gameplans sheets both carry a "Mega Used" column --
    `dive["mega_used"]`, constant for every row of one dive -- "note which
    one is used" whenever `core` carried 2 Mega-stone holders and the
    bring-4-consistent-mega choice had to pick one."""
    for rank, dive in core_dives:
        mega_used = dive.get("mega_used") or ""

        ws = wb.create_sheet(_safe_sheet_name(f"Dive {rank} Sets"))
        ws.append(["Pokemon", "Item", "Moves"])
        _style_header(ws)
        for name, s in dive["sets"].items():
            ws.append([name, s["item"] or "-", ", ".join(s["moves"])])
        _autosize(ws)

        ws = wb.create_sheet(_safe_sheet_name(f"Dive {rank} Summary"))
        ws.append(["Pair", "Enemy Team", "Beaten", "Total", "Swept", "Traded",
                   "Lost", "No KO", "Tailwind Safe", "Protect Safe",
                   "Clean Win Total", "Loses To", "Mega Used"])
        _style_header(ws)
        ov = dive["overall"]
        all_details = [pe["detail"] for pair in dive["per_pair"].values()
                       for pe in pair["per_enemy"]]
        ws.append(["OVERALL", "every pair, every enemy",
                   ov["pairs_swept"] + ov["pairs_traded"], ov["pairs_total"],
                   ov["pairs_swept"], ov["pairs_traded"], ov["pairs_lost"],
                   ov["pairs_no_ko"], ov["pairs_tailwind_safe"],
                   ov["pairs_protect_safe"], round(ov["pairs_clean_win_total"], 1),
                   _loses_to(all_details), mega_used])
        for (n1, n2), pair in dive["per_pair"].items():
            label = f"{n1} + {n2}"
            t = pair["total"]
            pair_details = [pe["detail"] for pe in pair["per_enemy"]]
            ws.append([label, "all enemies", t["pairs_swept"] + t["pairs_traded"],
                      t["pairs_total"], t["pairs_swept"], t["pairs_traded"],
                      t["pairs_lost"], t["pairs_no_ko"], t["pairs_tailwind_safe"],
                      t["pairs_protect_safe"], round(t["pairs_clean_win_total"], 1),
                      _loses_to(pair_details), mega_used])
            for pe in pair["per_enemy"]:
                s = pe["summary"]
                ws.append([label, ", ".join(pe["target_names"]),
                          s["pairs_swept"] + s["pairs_traded"], s["pairs_total"],
                          s["pairs_swept"], s["pairs_traded"], s["pairs_lost"],
                          s["pairs_no_ko"], s["pairs_tailwind_safe"],
                          s["pairs_protect_safe"], round(s["pairs_clean_win_total"], 1),
                          _loses_to([pe["detail"]]), mega_used])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)

        ws = wb.create_sheet(_safe_sheet_name(f"Dive {rank} Gameplans"))
        ws.append(["Pair", "Enemy Team", "Enemy Pair", "Outcome", "Turn",
                   "Actor", "Target", "Move", "Damage %", "Type Eff", "Spread",
                   "Mega Used"])
        _style_header(ws)
        for (n1, n2), pair in dive["per_pair"].items():
            role_name = {"C": n1, "P": n2}
            for pe in pair["per_enemy"]:
                for (e1, e2), d in pe["detail"].items():
                    role_name["E1"], role_name["E2"] = e1, e2
                    for turn_i, turn_hits in enumerate(d["log"], 1):
                        for role, tgt_role, h in turn_hits:
                            ws.append([f"{n1} + {n2}", ", ".join(pe["target_names"]),
                                      f"{e1} + {e2}", d["outcome"], turn_i,
                                      role_name[role], role_name[tgt_role],
                                      h.move_name or "-", round(h.frac * 100, 1),
                                      h.eff, h.num_targets_hit > 1, mega_used])
                    # A blank row between each enemy-pair "match" -- purely
                    # for legibility once several matches sit one after
                    # another in the same sheet.
                    ws.append([])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)


def _avg_score(names, merged):
    """Mean roster.csv Score across `names` -- "the average score (score
    from roster.csv) of the team". `species_data.build_merged_dataset`
    already merges roster.csv's Score onto every record as `merged[name]
    ["score"]`, so this is a plain lookup, not a new data source. A name
    roster.csv doesn't carry a Score for (`None`) is skipped rather than
    treated as 0, so one missing entry doesn't drag the average down for
    no real reason. `None` if nothing in `names` has a Score at all.
    """
    scores = [merged[n]["score"] for n in names
             if merged.get(n, {}).get("score") is not None]
    return sum(scores) / len(scores) if scores else None


def _pairs_note(pair_rows):
    """A compact "name+name beaten/total" note for EVERY one of a bring-4's
    own internal pairs (6 for a 4-Pokemon bring, fewer for a 3-Pokemon
    core), not just the single worst one the main columns already track --
    "keep a note of the 6 pairs for each enemy team and their overall
    performance along with the full summary"."""
    return "; ".join(
        f"{' + '.join(r['pair'])} {r['pairs_swept'] + r['pairs_traded']}/{r['pairs_total']}"
        for r in pair_rows)


def _per_90(count, n_pairs, pairs_total, scale=1.0):
    """`count` (out of `n_pairs * pairs_total * scale`) normalised to a
    fixed /90 unit -- "Avg Wins/90", "Clean wins/90" -- so the column stays
    comparable across cores of different sizes (a 3-Pokemon core has only 3
    pairs, not 6) and enemy rosters of different sizes (fewer than 6 named
    enemies means fewer than 15 enemy pairs, not the full 90), instead of
    silently mislabeling a smaller raw scale as "/90". `scale` is 2.0 for
    `pairs_clean_win_total`'s own 0.0-2.0-per-pair units, 1.0 (the default)
    for a plain beaten/tailwind-safe/protect-safe/no-faint COUNT.
    """
    denom = n_pairs * pairs_total * scale
    return (count / denom) * 90 if denom else 0.0


def _write_multi_bring4_xlsx(path, rows, target_name_lists, merged, moves_db,
                             natures, typechart, item_overrides, move_overrides,
                             excluded_items, fixed_items, fixed_moves, core_dives):
    """--multi-bring4's table as an Excel workbook -- "it may make more
    sense to make counter_table.py export an xlsx rather than a csv, so
    that I can see the performance vs team ... for each pair, and then vs
    each enemy". A flat CSV row squeezes the per-enemy breakdown into one
    joined string per cell; real columns (and, with `core_dives`, a real
    per-pair-per-enemy table plus the full gameplan) are the whole point of
    switching format. Reuses `export_excel.py`'s own header/autosize/sheet-
    name styling rather than re-inventing it.

    `core_dives`: [(rank, core_deep_dive's return), ...] for every core the
    caller opted to deep dive (--deep-dive-core or the interactive prompt)
    -- zero, one, or several, each getting its own 'Dive N ...' sheet trio.

    The 'Cores' sheet's leading '#' column is the FULL 1-based rank over
    `rows` (not just the console's --top slice) -- "this will also require
    the team number to be indexed in the xlsx export" -- and is exactly the
    number --deep-dive-core/the interactive prompt reads back.

    Each enemy's own column block also carries "Enemy N mega used" (the
    single Mega-stone holder that enemy's own best bring-4 treats as
    transformed, `pe["best_bring4_row"]["mega_used"]` -- "note which one is
    used" once a core's 2 stone holders needed the bring-4-consistent
    choice), "Enemy N lead"/"Enemy N backup" (`recommended_lead`: that
    enemy's own best bring's best-ranked pair, and the 2 members left over
    -- "is there a way to assess the best lead vs a given enemy team? Just
    the best pair + their backup"), and "Enemy N 6 pairs" (`_pairs_note`:
    every one of that bring's own internal pairs, by name, with its own
    beaten fraction -- "keep a note of the 6 pairs for each enemy team and
    their overall performance along with the full summary").

    The core-level "Sum 3rd Best ..." columns are a plain SUM (not the
    "Avg .../90" columns' per-90 normalisation) of each enemy's own
    `bring4_pair_depth` 3rd-best-pair reading, across every named enemy --
    "sum-of 3rd best pair for each of the relevant metrics across every
    enemy team in the summary".
    """
    from openpyxl import Workbook
    from export_excel import _autosize, _safe_sheet_name, _style_header
    wb = Workbook()
    all_enemies = sorted({n for t in target_name_lists for n in t})

    ws = wb.active
    ws.title = "Cores"
    header = ["#", "Core", "Size", "Bottleneck Enemy",
              "Weak to 2+ types (members)", "Weak to 1 type (members)",
              "Weak to 0 types (members)", "Types with 2+ weak members",
              "Weaknesses by type", "Average Score",
              "Avg Wins/90", "Avg Wins under Tailwind/90",
              "Avg Wins under Protect/90", "Clean wins/90",
              "Sum 3rd Best Pairs Beaten", "Sum 3rd Best Tailwind-Safe",
              "Sum 3rd Best Protect-Safe", "Sum 3rd Best No-Faint"]
    for i in range(len(target_name_lists)):
        header += [f"Enemy {i + 1}", f"Enemy {i + 1} best bring-4",
                   f"Enemy {i + 1} mega used",
                   f"Enemy {i + 1} lead", f"Enemy {i + 1} backup",
                   f"Enemy {i + 1} worst pair beaten",
                   f"Enemy {i + 1} uncovered enemy pairs",
                   f"Enemy {i + 1} pairs beaten total",
                   f"Enemy {i + 1} pairs beaten 3rd best",
                   f"Enemy {i + 1} pairs beaten 4th best",
                   f"Enemy {i + 1} pairs beaten worst",
                   f"Enemy {i + 1} has real Tailwind",
                   f"Enemy {i + 1} pairs Tailwind-safe total",
                   f"Enemy {i + 1} pairs Tailwind-safe best",
                   f"Enemy {i + 1} pairs Tailwind-safe 3rd best",
                   f"Enemy {i + 1} pairs protect-safe total",
                   f"Enemy {i + 1} pairs protect-safe best",
                   f"Enemy {i + 1} pairs protect-safe 3rd best",
                   f"Enemy {i + 1} pairs clean win total",
                   f"Enemy {i + 1} pairs beaten without fainting best",
                   f"Enemy {i + 1} pairs beaten without fainting 3rd best",
                   f"Enemy {i + 1} 6 pairs"]
    ws.append(header)
    _style_header(ws)
    sets_rows, teamsheet_entries = [], []
    for rank, r in enumerate(rows, start=1):
        core = r["core"]
        weak = member_weakness_summary(core, merged)
        by_type = sorted(((t, c) for t, c in weak["per_type"].items() if c > 0),
                         key=lambda tc: -tc[1])
        types_2plus = sum(1 for c in weak["per_type"].values() if c >= 2)
        # One `bring4_pair_depth` call per enemy, reused for BOTH the
        # core-level "Avg .../90" columns (averaged across every named
        # enemy below) and that enemy's own per-enemy columns further
        # right -- never re-derived.
        per_enemy_depths = [(pe, bring4_pair_depth(pe["best_bring4_row"]))
                            for pe in r["per_enemy"]]
        avg_wins = avg_wins_tw = avg_wins_pr = avg_clean = 0.0
        if per_enemy_depths:
            rates = [
                (_per_90(d["beaten_total"], len(pe["best_bring4_row"]["pair_rows"]),
                        d["pairs_total"]),
                 _per_90(d["tailwind_safe_total"], len(pe["best_bring4_row"]["pair_rows"]),
                        d["pairs_total"]),
                 _per_90(d["protect_safe_total"], len(pe["best_bring4_row"]["pair_rows"]),
                        d["pairs_total"]),
                 _per_90(d["no_faint_total"], len(pe["best_bring4_row"]["pair_rows"]),
                        d["pairs_total"]))
                for pe, d in per_enemy_depths]
            avg_wins = sum(x[0] for x in rates) / len(rates)
            avg_wins_tw = sum(x[1] for x in rates) / len(rates)
            avg_wins_pr = sum(x[2] for x in rates) / len(rates)
            avg_clean = sum(x[3] for x in rates) / len(rates)
        avg_score = _avg_score(core, merged)
        # "sum-of 3rd best pair for each of the relevant metrics across
        # every enemy team" -- unlike the "Avg .../90" columns above (which
        # normalise so cores/enemies of different sizes stay comparable),
        # this is a plain sum of each enemy's own `bring4_pair_depth`
        # 3rd-best-pair reading (`None` only for a core too small to HAVE a
        # 3rd pair, e.g. a 3-Pokemon core's 3-pair bring -- treated as 0,
        # not skipped, so a core that can't even field a 3rd pair reads as
        # weaker here rather than silently shrinking the sum's own count).
        sum_3rd_beaten = sum((d["beaten_3rd"] or 0) for _pe, d in per_enemy_depths)
        sum_3rd_tw = sum((d["tailwind_safe_3rd"] or 0) for _pe, d in per_enemy_depths)
        sum_3rd_pr = sum((d["protect_safe_3rd"] or 0) for _pe, d in per_enemy_depths)
        sum_3rd_nf = sum((d["no_faint_3rd"] or 0) for _pe, d in per_enemy_depths)
        row = [rank, " / ".join(core), r["core_size"], r["worst_enemy_idx"] + 1,
              weak["weak_to_2plus"], weak["weak_to_1"], weak["weak_to_0"],
              types_2plus, ", ".join(f"{t} {c}" for t, c in by_type),
              round(avg_score, 1) if avg_score is not None else "",
              round(avg_wins, 1), round(avg_wins_tw, 1),
              round(avg_wins_pr, 1), round(avg_clean, 1),
              sum_3rd_beaten, sum_3rd_tw, sum_3rd_pr, sum_3rd_nf]
        for pe, depth in per_enemy_depths:
            wr = pe["best_bring4_row"]["worst_pair_row"]
            uncovered = pe["best_bring4_row"]["uncovered_enemy_pairs"]
            pt = depth["pairs_total"]
            n_pairs = len(pe["best_bring4_row"]["pair_rows"])
            lead_backup = recommended_lead(pe["best_bring4_row"])
            row += [", ".join(pe["target_names"]), " / ".join(pe["best_bring4"]),
                   pe["best_bring4_row"].get("mega_used") or "",
                   " + ".join(lead_backup["lead"]), " + ".join(lead_backup["backup"]),
                   f"{wr['pairs_swept'] + wr['pairs_traded']}/{wr['pairs_total']}",
                   ", ".join(f"{e1}+{e2}" for e1, e2 in uncovered),
                   f"{depth['beaten_total']}/{n_pairs * pt}",
                   f"{depth['beaten_3rd']}/{pt}", f"{depth['beaten_4th']}/{pt}",
                   f"{depth['beaten_worst']}/{pt}",
                   enemy_has_real_tailwind(pe["target_names"], merged),
                   f"{depth['tailwind_safe_total']}/{n_pairs * pt}",
                   f"{depth['tailwind_safe_best']}/{pt}",
                   f"{depth['tailwind_safe_3rd']}/{pt}",
                   f"{depth['protect_safe_total']}/{n_pairs * pt}",
                   f"{depth['protect_safe_best']}/{pt}",
                   f"{depth['protect_safe_3rd']}/{pt}",
                   f"{depth['clean_win_total']:.1f}/{n_pairs * pt * 2:.0f}",
                   f"{depth['no_faint_best']}/{pt}",
                   f"{depth['no_faint_3rd']}/{pt}",
                   _pairs_note(pe["best_bring4_row"]["pair_rows"])]
        ws.append(row)
        core_fixed_items = fixed_items
        if r.get("item_clause_resolved_items"):
            core_fixed_items = {**fixed_items, **r["item_clause_resolved_items"]}
        core_sets = _fixed_sets_for(core, core_fixed_items, fixed_moves, merged,
                                    moves_db, natures, typechart, all_enemies,
                                    item_overrides, move_overrides, excluded_items)
        sets_rows.append((rank, " / ".join(core), core_sets))
        teamsheet_entries.append(
            (rank, " / ".join(core), "full core", core, core_sets))
        for e_idx, pe in enumerate(r["per_enemy"], start=1):
            b4 = pe["best_bring4"]
            b4_sets = {n: core_sets[n] for n in b4}
            teamsheet_entries.append(
                (rank, " / ".join(core), f"vs Enemy {e_idx}", b4, b4_sets))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)

    _write_sets_sheet(wb, sets_rows, merged)
    _write_teamsheets_sheet(wb, teamsheet_entries, merged)
    if core_dives:
        _write_dive_sheets(wb, core_dives, _safe_sheet_name, _style_header, _autosize)

    wb.save(path)
    return path


def _write_bring4_xlsx(path, bring4_rows, our6, targets, merged,
                       moves_db, natures, typechart, item_overrides,
                       move_overrides, excluded_items, core_dives,
                       fixed_items=None, fixed_moves=None):
    """--bring4's table as an Excel workbook -- the single-team/single-enemy-
    roster counterpart to `_write_multi_bring4_xlsx`, for "run a bring4 vs
    only ONE named team" with the same xlsx bridge: a ranked '#' column, a
    'Sets' sheet (our6's own item/moveset, searched once and shared by
    every bring-4 -- a real team can't re-optimise battle to battle), a
    'Teamsheets' sheet (pokepaste + base64 token for the full six and every
    bring-4 subset), and a Dive sheet trio per `core_dives` entry.

    `fixed_items`/`fixed_moves`: the sets `bring4_rows` was ACTUALLY raced
    with (`counter_finder._fixed_sets_from_pair_rows`, read back from
    `pair_rows`) -- when given, the 'Sets'/'Teamsheets' sheets read these
    instead of independently re-deriving each member's item/moveset, so
    Item Clause (`--unique-items`) reassignments the race already applied
    are never silently missing from what's displayed ("when unique-items
    are enforced, it is crucial these are reflected in the sets, summary,
    and gameplan for each team"). `None` (the default) falls back to the
    old independent-re-derivation behaviour, for a caller that never
    raced anything and has no `pair_rows` to read back from.

    Also carries "Mega used" (`b["mega_used"]` -- "note which one is used"
    once a bring's own 2 stone holders needed the bring-4-consistent
    choice), "6 pairs" (`_pairs_note`: every one of this bring's own
    internal pairs, by name, with its own beaten fraction -- "keep a note
    of the 6 pairs ... and their overall performance along with the full
    summary"), and "Lead"/"Backup" (`recommended_lead`: the single best-
    ranked of this bring's own pairs -- "is there a way to assess the best
    lead vs a given enemy team? Just the best pair + their backup").
    """
    from openpyxl import Workbook
    from export_excel import _autosize, _safe_sheet_name, _style_header
    wb = Workbook()

    enemy_tw = enemy_has_real_tailwind(targets, merged)
    ws = wb.active
    ws.title = "Bring-4s"
    ws.append(["#", "Bring-4", "Mega used", "Lead", "Backup",
              "Uncovered enemy pairs", "Pairs good",
              "Pairs total", "Worst pair", "Worst pair beaten",
              "Worst pair total", "Average Score",
              "Avg Wins/90", "Avg Wins under Tailwind/90",
              "Avg Wins under Protect/90", "Clean wins/90",
              "pairs beaten total",
              "pairs beaten 3rd best", "pairs beaten 4th best",
              "pairs beaten worst", "Enemy has real Tailwind",
              "pairs Tailwind-safe total", "pairs Tailwind-safe best",
              "pairs Tailwind-safe 3rd best", "pairs protect-safe total",
              "pairs protect-safe best", "pairs protect-safe 3rd best",
              "pairs clean win total",
              "pairs beaten without fainting best",
              "pairs beaten without fainting 3rd best", "6 pairs"])
    _style_header(ws)
    for rank, b in enumerate(bring4_rows, start=1):
        wr = b["worst_pair_row"]
        depth = bring4_pair_depth(b)
        pt = depth["pairs_total"]
        n_pairs = len(b["pair_rows"])
        avg_score = _avg_score(b["bring4"], merged)
        lead_backup = recommended_lead(b)
        ws.append([rank, " / ".join(b["bring4"]), b.get("mega_used") or "",
                  " + ".join(lead_backup["lead"]), " + ".join(lead_backup["backup"]),
                  ", ".join(f"{e1}+{e2}" for e1, e2 in b["uncovered_enemy_pairs"]),
                  b["pairs_good"], b["pairs_total"], " + ".join(b["worst_pair"]),
                  f"{wr['pairs_swept'] + wr['pairs_traded']}/{wr['pairs_total']}",
                  wr["pairs_total"],
                  round(avg_score, 1) if avg_score is not None else "",
                  round(_per_90(depth["beaten_total"], n_pairs, pt), 1),
                  round(_per_90(depth["tailwind_safe_total"], n_pairs, pt), 1),
                  round(_per_90(depth["protect_safe_total"], n_pairs, pt), 1),
                  round(_per_90(depth["no_faint_total"], n_pairs, pt), 1),
                  f"{depth['beaten_total']}/{n_pairs * pt}",
                  f"{depth['beaten_3rd']}/{pt}", f"{depth['beaten_4th']}/{pt}",
                  f"{depth['beaten_worst']}/{pt}", enemy_tw,
                  f"{depth['tailwind_safe_total']}/{n_pairs * pt}",
                  f"{depth['tailwind_safe_best']}/{pt}",
                  f"{depth['tailwind_safe_3rd']}/{pt}",
                  f"{depth['protect_safe_total']}/{n_pairs * pt}",
                  f"{depth['protect_safe_best']}/{pt}",
                  f"{depth['protect_safe_3rd']}/{pt}",
                  f"{depth['clean_win_total']:.1f}/{n_pairs * pt * 2:.0f}",
                  f"{depth['no_faint_best']}/{pt}",
                  f"{depth['no_faint_3rd']}/{pt}",
                  _pairs_note(b["pair_rows"])])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)

    our6_sets = _fixed_sets_for(our6, fixed_items or {}, fixed_moves or {},
                                merged, moves_db, natures,
                                typechart, targets, item_overrides,
                                move_overrides, excluded_items)
    _write_sets_sheet(wb, [(1, " / ".join(our6), our6_sets)], merged)

    teamsheet_entries = [(1, " / ".join(our6), "all 6", our6, our6_sets)]
    for rank, b in enumerate(bring4_rows, start=1):
        b4 = b["bring4"]
        teamsheet_entries.append(
            (rank, " / ".join(our6), f"bring-4 #{rank}", b4,
            {n: our6_sets[n] for n in b4}))
    _write_teamsheets_sheet(wb, teamsheet_entries, merged)

    if core_dives:
        _write_dive_sheets(wb, core_dives, _safe_sheet_name, _style_header, _autosize)

    wb.save(path)
    return path


def _print_switches(switch_results, bench_size):
    """`switch_results`: {(e1, e2): (rows, tried)} -- `switch_in_search`'s
    own return shape, one entry per LOSING enemy pair from the --deep report
    above. Ranked (within each enemy pair) the same way `switch_in_search`
    already ranks: least damage taken switching in first, among candidates
    that actually fix the loss."""
    if not switch_results:
        print("No losses to check switch-ins for -- every enemy pair is "
             "already held.\n")
        return
    print(f"\nSwitch-ins for {len(switch_results)} losing enemy pair(s), "
         f"from a bench of {bench_size}:\n")
    for (e1, e2), (rows, tried) in switch_results.items():
        print(f"  {e1} + {e2}:")
        if not rows:
            print(f"      No switch-in fixes it, among {tried} candidates tried.")
            continue
        for r in rows[:5]:
            tw = "" if r["tailwind_safe"] else f"  [tailwind: {r['tailwind_outcome']}]"
            print(f"      {r['leaving']} -> {r['arriving']}: takes "
                 f"{r['switch_in_taken'] * 100:.0f}% switching in, then "
                 f"{r['outcome']} (turn {r['turns_used']}){tw}")
        print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--vs", default="",
                    help="comma-separated target Pokemon, by library name. "
                         "Required by every mode except --multi-bring4 "
                         "(which uses --vs-team, repeated, instead)")
    ap.add_argument("--threshold", type=float, default=90.0,
                    help="%% threshold for the default mode's 'meets all' "
                         "column (default 90)")
    ap.add_argument("--chip-from", default="", metavar="POKEMON",
                    help="partner Pokemon whose move chips the targets first")
    ap.add_argument("--chip-move", default="", metavar="MOVE",
                    help="the partner's move (requires --chip-from)")
    ap.add_argument("--partner-item", default="", metavar="ITEM",
                    help="pin the --chip-from/--pairs partner's item instead "
                         "of searching for its best legal one")
    ap.add_argument("--item", default="", metavar="POKEMON=ITEM[;...]",
                    help="pin a specific legal item on named pool members "
                         "instead of searching for the best one, e.g. "
                         "'Gallade=Choice Scarf'. The moveset is still "
                         "re-optimised under the pinned item")
    ap.add_argument("--moves", default="", metavar="POKEMON=MOVE,MOVE,...[;...]",
                    help="pin the moveset too (skips move search) for named "
                         "pool members, e.g. 'Gallade=Psycho Cut,Sacred "
                         "Sword,Close Combat,Ice Punch'")
    ap.add_argument("--pairs", action="store_true",
                    help="speed-order pair search instead of the single-hit "
                         "threshold search")
    ap.add_argument("--speed", action="store_true",
                    help="print a speed-tier chart (priority bracket, then "
                         "effective speed) for --vs's targets plus the pool, "
                         "instead of a damage search")
    ap.add_argument("--joint", action="store_true",
                    help="a JOINT pair table: does OUR pair beat every pair "
                         "drawn from --vs, over --turns turns, real "
                         "priority-then-speed order, both of ours attacking "
                         "with their own real optimised set, plus a "
                         "Tailwind-robustness replay? With --partner, "
                         "searches the pool for the second half; without "
                         "it, GENERATES the whole pair -- every legal pair "
                         "from the pool (slower: see --partner's help)")
    ap.add_argument("--partner", default="", metavar="POKEMON",
                    help="--joint only: fix the second half of the pair and "
                         "search the pool for the first. Omit it to instead "
                         "search EVERY pair from the pool (C(pool,2) of "
                         "them) -- much slower, so narrow with --pool-size "
                         "unless the default pool is small enough already "
                         "(measured: ~6ms per candidate pair per 3 named "
                         "targets, so a --pool-size 80 run is under 20s "
                         "but the full ~270-Pokemon default is minutes)")
    ap.add_argument("--deep", action="store_true",
                    help="a DEEP DIVE on --our's specific pair (no search): "
                         "every enemy pair drawn from --vs shown in full -- "
                         "which of theirs could OHKO one of ours outright "
                         "(on their best roll, regardless of who moves "
                         "first), the 2x2 damage grid both directions, and "
                         "the turn-by-turn line it collapses into. Pass a "
                         "whole 6-Pokemon roster to --vs for 'all 15 "
                         "possible enemy leads'")
    ap.add_argument("--our", default="", metavar="POKEMON,POKEMON",
                    help="--deep: the exact pair to check (required). "
                         "--bring4: your already-decided team (required, 3, "
                         "4, 5, or 6 names -- 3 or 4 skips straight to "
                         "summarising its own internal pairs, since "
                         "there's only one possible bring)")
    ap.add_argument("--bring4", action="store_true",
                    help="for an ALREADY-DECIDED team (--our, 3-6 names) "
                         "against one enemy roster (--vs): every one of its "
                         "C(len,2) pairs (same table --joint prints), then "
                         "every one of its C(len,min(4,len)) possible bring "
                         "subsets, ranked by how bad its WORST internal "
                         "pair does (maximin) -- 'I always have options no "
                         "matter what position I am in', not just a strong "
                         "best pair with a hole behind it")
    ap.add_argument("--good-threshold", type=float, default=100.0, metavar="PCT",
                    help="--bring4/--multi-bring4 only: a pair counts as "
                         "'good' once it beats at least PCT%% of the named "
                         "enemy pairs (default 100 -- must beat ALL of them)")
    ap.add_argument("--multi-bring4", action="store_true",
                    help="find the best team-of-6 (drawn from the pool) "
                         "across SEVERAL enemy rosters at once (--vs-team, "
                         "repeated, instead of --vs): runs the pool-wide "
                         "pair search once per enemy, then finds the "
                         "team-of-6 whose WORST-case enemy (even bringing "
                         "that enemy its own best available bring-4 from "
                         "this team-of-6 -- a bring-4 may differ per "
                         "opponent, same as real Team Preview) is as good "
                         "as possible. Exhaustive by default (over a "
                         "candidate pool narrowed to names that appear in "
                         "a good pair for --min-enemies of the named "
                         "enemies); pass --beam for a broader, non-"
                         "exhaustive search over the whole pool instead")
    ap.add_argument("--vs-team", action="append", default=[],
                    metavar="POKEMON,...|TEAM NAME",
                    help="--multi-bring4/--bring4 only: an enemy roster, "
                         "EITHER comma-separated Pokemon, or the name of a "
                         "saved team (a data/teams.csv row, or a pokepaste "
                         "in data/teams/ or data/my_teams/ -- the same "
                         "library --team already searches). --multi-bring4: "
                         "repeat for each enemy team (1+ required -- a "
                         "single --vs-team searches the best core against "
                         "just that one roster), e.g. --vs-team \"Rain\" "
                         "--vs-team \"Big 6\". --bring4: exactly one, as an "
                         "alternative to spelling the same roster out with "
                         "--vs, e.g. --bring4 --vs-team \"Big 6\" instead of "
                         "--vs \"<Big 6's six names>\"")
    ap.add_argument("--vs-all-teams", action="store_true",
                    help="--multi-bring4 only: run against EVERY saved team "
                         "(data/teams.csv plus any pokepaste in data/teams/ "
                         "or data/my_teams/) instead of naming each one with "
                         "--vs-team -- mutually exclusive with --vs-team")
    ap.add_argument("--jobs", type=int, default=1, metavar="N",
                    help="--multi-bring4 only: search N enemy rosters in "
                         "parallel worker processes instead of one after "
                         "another (0 = one per CPU core). Each worker loads "
                         "the dataset once (~14s) and reuses it, so this is "
                         "close to a linear speedup on a run with several "
                         "--vs-team entries or --vs-all-teams -- no effect "
                         "with only one enemy roster (nothing to split "
                         "across workers). Memory is the limit, not CPU: "
                         "budget roughly 1GB per worker, same as "
                         "search_teams.py's --jobs.")
    ap.add_argument("--max-weak", type=int, default=2, metavar="N",
                    help="--multi-bring4 only: hard-drop any candidate CORE "
                         "where more than N of its members are weak to the "
                         "SAME type, for every type (default 2: 'only 2 "
                         "members may be weak to any one type'). "
                         "Overridden per-type by --type-limit. Unlike "
                         "--good-threshold this is a hard exclusion, not a "
                         "ranking factor -- a core that breaks it is never "
                         "shown, however well it wins its bring-4s. Pass a "
                         "high value (e.g. --max-weak 6) to effectively "
                         "disable it")
    ap.add_argument("--type-limit", action="append", default=[],
                    metavar="TYPE:max_weak=N,max_net=M",
                    help="--multi-bring4 only: a hard per-type override on "
                         "top of --max-weak, repeatable. max_weak=N caps "
                         "how many members may be weak to TYPE; max_net=M "
                         "caps (weak count - resist/immune count) for TYPE "
                         "-- e.g. --type-limit \"Fire:max_weak=1,max_net=-2\" "
                         "for 'only 1 member may be weak to Fire, Fire must "
                         "have net 2 resistances'. Either key may be omitted")
    ap.add_argument("--strict-weak-types", nargs="?", const="__DEFAULT__",
                    default=None, metavar="TYPE,TYPE,...",
                    help="--multi-bring4 only: shorthand that hard-caps "
                         "max_weak=1 for the named types (comma-separated), "
                         "on top of --max-weak/--type-limit -- an explicit "
                         "--type-limit for the same type wins over this. "
                         "Bare (no value) uses the default list: Fire, "
                         "Water, Fairy, Steel, Dark, Ghost, Ground. Omit "
                         "the flag entirely for no constraint from this")
    ap.add_argument("--max-weak-types", type=int, default=None, metavar="N",
                    help="--multi-bring4 only: hard-drop any candidate CORE "
                         "where more than N DIFFERENT types have 2+ members "
                         "weak to them -- e.g. --max-weak-types 3 for 'no "
                         "more than 3 types may have 2 weak members'. A "
                         "breadth cap, distinct from --max-weak's own "
                         "per-type ceiling: a core could satisfy --max-weak "
                         "2 (no type ever exceeds 2 weak members) while "
                         "still being broadly fragile across many types at "
                         "once, which this catches instead. Off by default "
                         "(no breadth cap)")
    ap.add_argument("--core-sizes", default="4,5,6", metavar="N,N,...",
                    help="--multi-bring4 only: which candidate CORE sizes "
                         "to search, comma-separated (default \"4,5,6\"). "
                         "Each must be 3-6 -- \"I would like to output the "
                         "best 3-pokemon cores against each team\" widens "
                         "this down to 3; a 3-member core has exactly one "
                         "possible bring (itself, 3 pairs), the same "
                         "degenerate case a 4-member core already is")
    ap.add_argument("--max-megas", type=int, default=2, metavar="N",
                    help="--multi-bring4 only: hard cap on how many "
                         "Mega-stone-capable members a candidate CORE may "
                         "contain (default 2, VGC's real team-composition "
                         "limit -- 'a full team can only have two mega "
                         "stone users'). In an actual pair, EITHER may "
                         "still choose to transform depending on the "
                         "specific matchup -- that per-battle choice is "
                         "the existing mega-vs-stay-base minimax, unaffected "
                         "by this; this only caps how many are BROUGHT at "
                         "all")
    ap.add_argument("--allow-scarf", action="store_true",
                    help="by default Choice Scarf is excluded from every "
                         "item SEARCH (it's legal in Regulation MB, but too "
                         "easy to punish to want as a default recommendation "
                         "-- see counter_finder.DEFAULT_EXCLUDED_ITEMS). "
                         "Pass this to let the search consider it again; an "
                         "explicit --item pin already bypasses the "
                         "exclusion regardless")
    ap.add_argument("--unique-items", action="store_true",
                    help="--bring4/--multi-bring4 only: enforce the VGC "
                         "Item Clause (no two of the team's own Pokemon may "
                         "hold the same item) when resolving items. Off by "
                         "default -- items are searched per-member in "
                         "isolation, same as always, to keep search fast; "
                         "this is meant for verifying a specific team, not "
                         "for bulk ranking. Under --multi-bring4 this only "
                         "affects --deep-dive-core's own per-core re-race "
                         "-- the exhaustive/beam sweep itself never enforces "
                         "this, regardless of this flag")
    ap.add_argument("--min-enemies", type=int, default=2, metavar="N",
                    help="--multi-bring4 only: a pool member only enters "
                         "the exhaustive search's candidate pool once it "
                         "appears in a good pair (--good-threshold) for at "
                         "least N of the named --vs-team enemies (default "
                         "2, auto-clamped to however many --vs-team entries "
                         "were actually given if that's fewer -- a single "
                         "--vs-team never needs this flag). Ignored under "
                         "--beam, which searches the "
                         "whole pool regardless")
    ap.add_argument("--beam", action="store_true",
                    help="--multi-bring4 only: search the WHOLE pool with "
                         "an incremental beam search (same growth pattern "
                         "the Generate Team tab's own search uses) instead "
                         "of an exhaustive sweep of the narrowed candidate "
                         "pool -- not guaranteed optimal, but works when "
                         "that candidate pool is too large to sweep "
                         "exhaustively (or --min-enemies/--good-threshold "
                         "narrowed it down to nothing)")
    ap.add_argument("--beam-width", type=int, default=40, metavar="N",
                    help="--beam only: how many partial teams to keep at "
                         "each growth step (default 40)")
    ap.add_argument("--max-candidates", type=int, default=30, metavar="N",
                    help="--multi-bring4 (exhaustive, not --beam) only: "
                         "refuse to sweep a candidate pool bigger than this "
                         "(C(N,6) grows fast; default 30 -- C(30,6) is "
                         "~593k, still a couple of seconds)")
    ap.add_argument("--switches", action="store_true",
                    help="--deep only: for every enemy pair --our LOSES, "
                         "try swapping in each --bench candidate (or the "
                         "whole pool, if --bench isn't given) for each of "
                         "--our -- one turn is spent on the switch (the "
                         "incoming Pokemon doesn't get to act that turn, "
                         "same as a real doubles switch), then the new pair "
                         "is raced the same way --deep already does. Reports "
                         "candidates that actually fix the loss, ranked by "
                         "LEAST damage taken switching in ('easy and "
                         "optimal')")
    ap.add_argument("--bench", default="", metavar="POKEMON,POKEMON,...",
                    help="--switches only: the specific switch-in "
                         "candidates to try, instead of searching the whole "
                         "pool (--pool-size narrows that pool the same way "
                         "it does everywhere else)")
    ap.add_argument("--turns", type=int, default=2, metavar="N",
                    help="--joint/--deep/--bring4/--multi-bring4 only: how "
                         "many turns to race (default 2)")
    ap.add_argument("--max-taken", type=float, default=None, metavar="PCT",
                    help="default mode only: drop any row where SOME named "
                         "target's best attack could do PCT%% or more to it "
                         "(their best roll, the guaranteed-survival "
                         "direction). E.g. --max-taken 50 for 'take less "
                         "than 50%%'")
    ap.add_argument("--outspeed", choices=("natural", "scarf"), default=None,
                    help="default mode only: drop any row that doesn't "
                         "out-speed EVERY named target. 'natural': under "
                         "its own chosen item. 'scarf': naturally, OR if it "
                         "held Choice Scarf instead (a hypothesis -- pin "
                         "--item yourself for that to show in the damage "
                         "numbers too)")
    ap.add_argument("--pool-size", type=int, default=0, metavar="N",
                    help="search only the top N of the generic team-"
                         "generation Score ranking, instead of the whole "
                         "~270-Pokemon dataset (the default, 0). Narrowing "
                         "is rarely useful -- a full search is a couple of "
                         "seconds -- and can HIDE a real answer whose "
                         "generic Score is unremarkable (see 'Why does Mega "
                         "Scizor not show up' at the top of this file)")
    ap.add_argument("--team", default="",
                    help="search this library team's members instead of the "
                         "full dataset")
    ap.add_argument("--top", type=int, default=20, help="rows to print")
    ap.add_argument("--csv", default="", help="also write the whole table here")
    ap.add_argument("--xlsx", default="", metavar="PATH",
                    help="--multi-bring4/--bring4 only: write the same "
                         "table to an Excel workbook instead of a flat CSV "
                         "row -- lets you actually see the full beaten/"
                         "swept/traded/lost/no-KO/tw-safe/pr-safe breakdown "
                         "per pair per enemy, not just one squeezed-"
                         "together cell, plus a ranked '#' column, every "
                         "core/bring-4's own item+moveset ('Sets' sheet), "
                         "and a pastable pokepaste + base64 teamsheet token "
                         "per core/bring-4 ('Teamsheets' sheet -- paste "
                         "either into the Streamlit app's Team Builder). "
                         "With --deep-dive-core also given, the workbook "
                         "gets a dedicated sheet trio per dived core with "
                         "that full detail plus the turn-by-turn gameplan "
                         "for every enemy")
    ap.add_argument("--teamsheet-json", default="", metavar="PATH",
                    help="--multi-bring4/--bring4 only: write the Nth-ranked "
                         "core/bring-4's teamsheet (--deep-dive-core's N, "
                         "or the top result if --deep-dive-core wasn't "
                         "given) as JSON in exactly the format the "
                         "Streamlit app's Team Builder tab already reads/"
                         "writes -- {\"pool\": [...], \"sets\": {name: "
                         "{\"item\", \"moves\"}}}. Pass '-' to print it to "
                         "stdout (framed for copy-paste) instead of "
                         "writing a file -- \"the export in the CLI needs "
                         "to be able to export to the streamlit app, "
                         "whether through a paste or otherwise\"")
    ap.add_argument("--deep-dive-core", default="", metavar="N[,N,...]",
                    help="--multi-bring4/--bring4 only: after the main "
                         "search, run a full deep dive on the Nth-ranked "
                         "core/bring-4 shown (1 = top result) -- every one "
                         "of its pairs raced against every enemy pair from "
                         "every named enemy team, with the turn-by-turn "
                         "gameplan log and an aggregate beaten/swept/traded/"
                         "lost/no-KO/tw-safe/pr-safe total (both per-pair "
                         "and for the whole core). Deliberately NOT "
                         "computed for every core in the main search -- opt "
                         "in for the ones you actually want to inspect. "
                         "Comma-separated for several at once, e.g. "
                         "--deep-dive-core \"5,85,16\" -- the xlsx's '#' "
                         "column (with --xlsx) is the same ranking this "
                         "reads")
    ap.add_argument("--no-prompt", action="store_true",
                    help="--multi-bring4/--bring4 only: skip the interactive "
                         "'deep dive which cores?' prompt this tool asks "
                         "after the main search when running in a real "
                         "terminal and --deep-dive-core wasn't already "
                         "given. Scripted/piped runs never see the prompt "
                         "regardless -- this is for an interactive run "
                         "that wants the search results only")
    args = ap.parse_args()

    if bool(args.chip_from) != bool(args.chip_move):
        raise SystemExit("--chip-from and --chip-move must be given together")
    if args.partner_item and not (args.chip_from or args.partner):
        raise SystemExit("--partner-item requires --chip-from/--chip-move or "
                         "--joint --partner (there's no fixed partner to pin "
                         "an item on when --joint generates the whole pair)")
    if (args.max_taken is not None or args.outspeed) and (
            args.pairs or args.chip_from or args.speed or args.joint):
        raise SystemExit("--max-taken/--outspeed only apply to the default "
                         "(threshold) mode")
    if args.joint and (args.pairs or args.chip_from or args.speed):
        raise SystemExit("--joint cannot be combined with --pairs/--chip-from/--speed")
    if args.deep and (args.joint or args.pairs or args.chip_from or args.speed):
        raise SystemExit("--deep cannot be combined with --joint/--pairs/"
                         "--chip-from/--speed")
    if args.bring4 and (args.joint or args.deep or args.pairs or args.chip_from
                        or args.speed):
        raise SystemExit("--bring4 cannot be combined with --joint/--deep/"
                         "--pairs/--chip-from/--speed")
    if args.multi_bring4 and (args.joint or args.deep or args.bring4 or args.pairs
                              or args.chip_from or args.speed):
        raise SystemExit("--multi-bring4 cannot be combined with --joint/"
                         "--deep/--bring4/--pairs/--chip-from/--speed")
    if not args.multi_bring4 and not args.bring4 and not args.vs:
        raise SystemExit("--vs is required (except with --multi-bring4/"
                         "--bring4, which can use --vs-team instead)")
    if args.multi_bring4 and args.vs:
        raise SystemExit("--multi-bring4 uses --vs-team (repeated), not --vs")
    if args.vs_all_teams and args.vs_team:
        raise SystemExit("--vs-all-teams can't be combined with --vs-team -- "
                         "it already runs against every saved team")
    if args.vs_all_teams and not args.multi_bring4:
        raise SystemExit("--vs-all-teams requires --multi-bring4")
    if args.multi_bring4 and len(args.vs_team) < 1 and not args.vs_all_teams:
        raise SystemExit("--multi-bring4 needs at least one --vs-team "
                         "(or --vs-all-teams)")
    if args.bring4:
        if bool(args.vs) == bool(args.vs_team):
            raise SystemExit("--bring4 needs exactly one of --vs "
                             "\"Pokemon,...\" or --vs-team "
                             "\"Pokemon,...|TEAM NAME\"")
        if len(args.vs_team) > 1:
            raise SystemExit("--bring4 takes at most one --vs-team (a "
                             "single enemy roster) -- --multi-bring4 is "
                             "for several at once")
    if args.vs_team and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--vs-team requires --multi-bring4 or --bring4")
    if args.deep_dive_core and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--deep-dive-core requires --multi-bring4 or --bring4")
    if args.no_prompt and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--no-prompt requires --multi-bring4 or --bring4")
    if args.xlsx and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--xlsx requires --multi-bring4 or --bring4")
    if (args.beam or args.beam_width != 40 or args.max_candidates != 30
            or args.min_enemies != 2) and not args.multi_bring4:
        raise SystemExit("--beam/--beam-width/--max-candidates/--min-enemies "
                         "only apply to --multi-bring4")
    if (args.max_weak != 2 or args.type_limit) and not args.multi_bring4:
        raise SystemExit("--max-weak/--type-limit only apply to --multi-bring4")
    if args.max_weak_types is not None and not args.multi_bring4:
        raise SystemExit("--max-weak-types only applies to --multi-bring4")
    if args.strict_weak_types is not None and not args.multi_bring4:
        raise SystemExit("--strict-weak-types only applies to --multi-bring4")
    if args.core_sizes != "4,5,6" and not args.multi_bring4:
        raise SystemExit("--core-sizes only applies to --multi-bring4")
    if args.jobs != 1 and not args.multi_bring4:
        raise SystemExit("--jobs only applies to --multi-bring4")
    if args.unique_items and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--unique-items only applies to --bring4/--multi-bring4")
    if args.teamsheet_json and not (args.multi_bring4 or args.bring4):
        raise SystemExit("--teamsheet-json requires --multi-bring4 or --bring4")
    if args.partner and not args.joint:
        raise SystemExit("--partner requires --joint")
    if args.turns != 2 and not (args.joint or args.deep or args.bring4
                                or args.multi_bring4):
        raise SystemExit("--turns only applies to --joint/--deep/--bring4/--multi-bring4")
    if args.deep and not args.our:
        raise SystemExit("--deep requires --our \"Pokemon,Pokemon\"")
    if args.bring4 and not args.our:
        raise SystemExit("--bring4 requires --our \"Pokemon,Pokemon,...\" "
                         "(exactly 6 names)")
    if args.our and not (args.deep or args.bring4):
        raise SystemExit("--our requires --deep or --bring4")
    if args.deep and (args.partner or args.partner_item):
        raise SystemExit("--partner/--partner-item don't apply to --deep -- "
                         "--our already names both of the pair")
    if args.bring4 and (args.partner or args.partner_item):
        raise SystemExit("--partner/--partner-item don't apply to --bring4 -- "
                         "--our already names your whole 6")
    if args.switches and not args.deep:
        raise SystemExit("--switches requires --deep")
    if args.bench and not args.switches:
        raise SystemExit("--bench requires --switches")
    if not (0.0 <= args.good_threshold <= 100.0):
        raise SystemExit("--good-threshold must be between 0 and 100")

    from _harness import load_world
    from lead_sim import BANNED_ITEMS
    W = load_world()
    merged, moves, natures, typechart = (W["merged"], W["moves"], W["natures"],
                                         W["typechart"])
    if args.vs:
        targets = [n.strip() for n in args.vs.split(",") if n.strip()]
        missing = [n for n in targets if n not in merged]
        if missing:
            raise SystemExit(f"unknown Pokemon: {', '.join(missing)}")
    elif args.bring4 and args.vs_team:
        targets = _resolve_vs_team(args.vs_team[0], W["teams"], merged)
    else:
        targets = []
    if args.partner and args.partner not in merged:
        raise SystemExit(f"unknown Pokemon: {args.partner}")
    our_pair = [n.strip() for n in args.our.split(",") if n.strip()]
    if args.deep:
        if len(our_pair) != 2:
            raise SystemExit(f"--our needs exactly 2 Pokemon, lead first "
                             f"(got {len(our_pair)}: {our_pair})")
        unknown_our = [n for n in our_pair if n not in merged]
        if unknown_our:
            raise SystemExit(f"unknown Pokemon: {', '.join(unknown_our)}")
    if args.bring4:
        our6 = list(dict.fromkeys(our_pair))
        if not (3 <= len(our6) <= 6):
            raise SystemExit(f"--bring4 needs 3, 4, 5, or 6 distinct Pokemon "
                             f"in --our (got {len(our6)}: {our6})")
        unknown_our6 = [n for n in our6 if n not in merged]
        if unknown_our6:
            raise SystemExit(f"unknown Pokemon: {', '.join(unknown_our6)}")
    bench = [n.strip() for n in args.bench.split(",") if n.strip()]
    if bench:
        unknown_bench = [n for n in bench if n not in merged]
        if unknown_bench:
            raise SystemExit(f"unknown Pokemon: {', '.join(unknown_bench)}")
    vs_teams = []
    if args.multi_bring4:
        vs_team_names = list(W["teams"]) if args.vs_all_teams else args.vs_team
        for raw in vs_team_names:
            vs_teams.append(_resolve_vs_team(raw, W["teams"], merged))
        if args.min_enemies == 2 and len(vs_teams) < 2:
            # Still at the untouched default -- auto-clamp rather than make
            # a single-roster search jump through an extra flag for it.
            args.min_enemies = len(vs_teams)
        if not (1 <= args.min_enemies <= len(vs_teams)):
            raise SystemExit(f"--min-enemies must be between 1 and the "
                             f"number of --vs-team entries ({len(vs_teams)})")
        args.jobs, _jobs_warning = blas_limits.workers_advice(args.jobs)
    # --deep/--bring4 alone need no pool at all (a fixed pair/six); --switches
    # needs one UNLESS --bench already named the exact candidates to try.
    pool = (_pool(args, merged)
           if (not (args.deep or args.bring4) or (args.switches and not bench))
           else [])
    threshold = args.threshold / 100.0

    item_overrides = _parse_item_overrides(args.item)
    move_overrides = _parse_move_overrides(args.moves)
    excluded_items = frozenset() if args.allow_scarf else DEFAULT_EXCLUDED_ITEMS
    type_limits = _parse_type_limits(args.type_limit)
    strict_weak_types = _resolve_strict_weak_types(args.strict_weak_types)
    type_limits = _merge_strict_weak_types(type_limits, strict_weak_types)
    try:
        core_sizes = tuple(int(p.strip()) for p in args.core_sizes.split(",")
                           if p.strip())
    except ValueError:
        raise SystemExit(f"--core-sizes {args.core_sizes!r}: must be "
                         f"comma-separated integers")
    if not core_sizes or any(not (3 <= s <= 6) for s in core_sizes):
        raise SystemExit(f"--core-sizes {args.core_sizes!r}: each size must "
                         f"be between 3 and 6")
    banned = [i for i in (item_overrides or {}).values() if i in BANNED_ITEMS]
    if banned:
        raise SystemExit(f"--item: not legal in Regulation MB: {', '.join(banned)}")
    overridden = set(item_overrides or {}) | set(move_overrides or {})
    unknown = [n for n in overridden if n not in merged]
    if unknown:
        raise SystemExit(f"--item/--moves: unknown Pokemon: {', '.join(unknown)}")
    # A pin only means something if the Pokemon is actually searched -- add
    # anything named in --item/--moves that the pool (top-N or --team) didn't
    # already include, so e.g. "Gallade=Choice Scarf" is tested even when
    # Gallade wouldn't otherwise have made a --pool-size 40 cut.
    pool = pool + [n for n in overridden if n not in pool and n not in targets]
    if args.partner_item and args.partner_item in BANNED_ITEMS:
        raise SystemExit(f"--partner-item: {args.partner_item!r} is not legal "
                         "in Regulation MB")

    if args.deep:
        print(f"Deep dive: {' + '.join(our_pair)} vs {', '.join(targets)}\n")
    elif args.bring4:
        print(f"Bring-4 search: {' / '.join(our6)} vs {', '.join(targets)}\n")
    elif args.multi_bring4:
        print(f"Multi-bring4 search: {len(pool)} Pokemon vs "
             f"{len(vs_teams)} enemy teams")
        if args.jobs > 1 and len(vs_teams) > 1:
            print(f"workers  : {args.jobs} of {os.cpu_count()} cores")
            if _jobs_warning:
                print(f"WARNING  : {_jobs_warning}")
        print()
    else:
        print(f"Searching {len(pool)} Pokemon vs {', '.join(targets)}\n")

    if args.joint and not args.partner:
        # ~2ms per (our-pair, enemy-pair) combo, measured -- see --partner's
        # help text. Printed up front, same "say how long before it starts"
        # courtesy `lead_sweep.py`'s sweep stage and WORKFLOW.md's audit
        # estimates already give: this is the one search in this module
        # whose cost is quadratic in the pool rather than linear.
        n_our_pairs = len(pool) * (len(pool) - 1) // 2
        n_enemy_pairs = len(targets) * (len(targets) - 1) // 2
        est = n_our_pairs * n_enemy_pairs * 0.002
        print(f"Generating the pair: {n_our_pairs} candidate pairs x "
              f"{n_enemy_pairs} enemy pairs -- roughly {est:.0f}s.")
        if est > 30:
            print(f"  That's slow -- --pool-size narrows the {len(pool)}-Pokemon "
                 f"pool (pairs grow as N^2/2).\n")
        else:
            print()

    if args.deep:
        item1, item2, detail, summary = deep_dive(
            our_pair[0], our_pair[1], targets, merged, moves, natures,
            typechart, turns=args.turns, item_overrides=item_overrides,
            move_overrides=move_overrides, excluded_items=excluded_items)
        _print_deep(our_pair[0], our_pair[1], item1, item2, targets, detail,
                   summary, args.turns)
        if args.switches:
            bench_names = bench or [n for n in pool if n not in our_pair
                                    and n not in targets]
            switch_results = {}
            for (e1, e2), d in detail.items():
                if d["outcome"] != "loss":
                    continue
                s_rows, s_tried = switch_in_search(
                    our_pair[0], our_pair[1], (e1, e2), bench_names, merged,
                    moves, natures, typechart, turns=args.turns,
                    item_overrides=item_overrides, move_overrides=move_overrides,
                    excluded_items=excluded_items)
                switch_results[(e1, e2)] = (s_rows, s_tried)
            _print_switches(switch_results, len(bench_names))
    elif args.bring4:
        good_threshold = args.good_threshold / 100.0
        pair_rows, bring4_rows = bring4_search(
            our6, targets, merged, moves, natures, typechart,
            turns=args.turns, good_threshold=good_threshold,
            item_overrides=item_overrides, move_overrides=move_overrides,
            excluded_items=excluded_items,
            enforce_item_clause=args.unique_items)
        _print_bring4(pair_rows, bring4_rows, our6, targets, args.top,
                     args.turns, good_threshold)
        ranks = _parse_deep_dive_core(args.deep_dive_core)
        if not ranks:
            ranks = _prompt_deep_dive_ranks(len(bring4_rows), args.no_prompt)
        core_dives = []
        for rank in ranks:
            if rank > len(bring4_rows):
                raise SystemExit(f"--deep-dive-core {rank} but only "
                                 f"{len(bring4_rows)} bring-4(s) were found")
            dive = core_deep_dive(
                bring4_rows[rank - 1]["bring4"], [targets], merged, moves,
                natures, typechart, turns=args.turns,
                item_overrides=item_overrides, move_overrides=move_overrides,
                excluded_items=excluded_items,
                enforce_item_clause=args.unique_items)
            _print_core_deep_dive(dive)
            core_dives.append((rank, dive))
        if args.xlsx:
            fixed_items, fixed_moves = _fixed_sets_from_pair_rows(
                pair_rows, merged, moves, natures, typechart, targets, excluded_items)
            path = _write_bring4_xlsx(
                args.xlsx, bring4_rows, our6, targets, merged,
                moves, natures, typechart, item_overrides, move_overrides,
                excluded_items, core_dives,
                fixed_items=fixed_items, fixed_moves=fixed_moves)
            print(f"\nExcel workbook: {os.path.abspath(path)}")
        if args.teamsheet_json:
            # No core was dived (no --deep-dive-core, no interactive pick)
            # -- default to the top result, same "1 = top result"
            # convention --deep-dive-core itself uses.
            dive = core_dives[0][1] if core_dives else core_deep_dive(
                bring4_rows[0]["bring4"], [targets], merged, moves,
                natures, typechart, turns=args.turns,
                item_overrides=item_overrides, move_overrides=move_overrides,
                excluded_items=excluded_items,
                enforce_item_clause=args.unique_items)
            _write_teamsheet_json(args.teamsheet_json, dive)
    elif args.multi_bring4:
        good_threshold = args.good_threshold / 100.0
        coverage = multi_bring4_coverage(
            pool, vs_teams, merged, moves, natures, typechart,
            turns=args.turns, good_threshold=good_threshold,
            min_enemies=args.min_enemies, item_overrides=item_overrides,
            move_overrides=move_overrides, excluded_items=excluded_items,
            jobs=args.jobs)
        print(f"Candidate pool (appears in a good pair for >= "
             f"{args.min_enemies} of {len(vs_teams)} enemies): "
             f"{len(coverage['candidate_pool'])} of {len(pool)}\n")
        # Stage A (the pair-vs-each-enemy search) is already done at this
        # point, regardless of whether the team-of-6 sweep below can even
        # run -- "I want to at least see the results (i.e., high performing
        # pairs, results /15)" when a huge candidate pool makes the
        # exhaustive sweep infeasible. Always shown, not just on that
        # fallback, so a --multi-bring4 run never depends on Stage B
        # succeeding to show anything at all.
        _print_pair_summary(coverage, top=args.top)
        if args.beam:
            multi_rows = multi_bring4_beam(
                coverage, good_threshold=good_threshold,
                beam_width=args.beam_width, max_weak=args.max_weak,
                type_limits=type_limits, max_megas=args.max_megas,
                max_weak_types=args.max_weak_types, core_sizes=core_sizes,
                enforce_item_clause=args.unique_items)
            mode_label = f"beam, width {args.beam_width}"
        else:
            try:
                multi_rows = multi_bring4_exhaustive(
                    coverage, good_threshold=good_threshold,
                    max_candidates=args.max_candidates, max_weak=args.max_weak,
                    type_limits=type_limits, max_megas=args.max_megas,
                    max_weak_types=args.max_weak_types, core_sizes=core_sizes,
                    enforce_item_clause=args.unique_items)
                mode_label = "exhaustive"
            except ValueError as e:
                # "It should be very quick to compute the sets of 4 brings
                # for each [enemy] with the highest wins /15 ... this
                # should not take long" -- multi_bring4_beam already IS
                # that fast path (bounded by beam_width * pool_size per
                # growth step, not C(pool,6)), and it's built from the
                # exact same already-computed pair data Stage A just
                # produced -- no new racing. Auto-fallback instead of
                # erroring out with nothing to show.
                print(f"{e}\nFalling back to --beam automatically (width "
                     f"{args.beam_width}) so there's still a result.\n")
                multi_rows = multi_bring4_beam(
                    coverage, good_threshold=good_threshold,
                    beam_width=args.beam_width, max_weak=args.max_weak,
                    type_limits=type_limits, max_megas=args.max_megas,
                    max_weak_types=args.max_weak_types, core_sizes=core_sizes,
                    enforce_item_clause=args.unique_items)
                mode_label = f"beam, width {args.beam_width} (auto-fallback)"
        _print_multi_bring4(multi_rows, vs_teams, args.top, mode_label,
                            good_threshold, len(coverage["candidate_pool"]),
                            len(pool), merged, moves, natures, typechart,
                            item_overrides, move_overrides, excluded_items,
                            fixed_items=coverage["fixed_items"],
                            fixed_moves=coverage["fixed_moves"],
                            core_sizes=core_sizes)
        ranks = _parse_deep_dive_core(args.deep_dive_core)
        if not ranks:
            ranks = _prompt_deep_dive_ranks(len(multi_rows), args.no_prompt)
        core_dives = []
        for rank in ranks:
            if rank > len(multi_rows):
                raise SystemExit(f"--deep-dive-core {rank} but only "
                                 f"{len(multi_rows)} core(s) were found")
            dive = core_deep_dive(
                multi_rows[rank - 1]["core"], vs_teams, merged, moves,
                natures, typechart, turns=args.turns,
                item_overrides=item_overrides, move_overrides=move_overrides,
                excluded_items=excluded_items,
                enforce_item_clause=args.unique_items)
            _print_core_deep_dive(dive)
            core_dives.append((rank, dive))
        if args.xlsx:
            path = _write_multi_bring4_xlsx(
                args.xlsx, multi_rows, vs_teams, merged, moves, natures,
                typechart, item_overrides, move_overrides, excluded_items,
                coverage["fixed_items"], coverage["fixed_moves"], core_dives)
            print(f"\nExcel workbook: {os.path.abspath(path)}")
        if args.teamsheet_json:
            if not multi_rows:
                raise SystemExit("--teamsheet-json: no core was found to "
                                 "export -- widen the pool, lower "
                                 "--good-threshold/--min-enemies, relax "
                                 "--max-weak/--type-limit/--max-weak-types, "
                                 "or try --beam")
            # No core was dived (no --deep-dive-core, no interactive pick)
            # -- default to the top result, same "1 = top result"
            # convention --deep-dive-core itself uses.
            dive = core_dives[0][1] if core_dives else core_deep_dive(
                multi_rows[0]["core"], vs_teams, merged, moves, natures,
                typechart, turns=args.turns, item_overrides=item_overrides,
                move_overrides=move_overrides, excluded_items=excluded_items,
                enforce_item_clause=args.unique_items)
            _write_teamsheet_json(args.teamsheet_json, dive)
    elif args.speed:
        names = targets + [n for n in pool if n not in targets]
        rows = speed_tiers(names, targets, merged, moves, natures, typechart,
                           item_overrides=item_overrides,
                           excluded_items=excluded_items)
        _print_speed(rows, targets, args.top)
    elif args.joint and args.partner:
        rows = joint_pair_search(pool, targets, args.partner, merged, moves,
                                 natures, typechart, turns=args.turns,
                                 partner_item=args.partner_item or None,
                                 item_overrides=item_overrides,
                                 move_overrides=move_overrides,
                                 excluded_items=excluded_items)
        _print_joint(rows, targets, args.top, args.partner, args.turns)
    elif args.joint:
        rows = joint_pool_search(pool, targets, merged, moves, natures,
                                 typechart, turns=args.turns,
                                 item_overrides=item_overrides,
                                 move_overrides=move_overrides,
                                 excluded_items=excluded_items)
        _print_joint(rows, targets, args.top, "", args.turns)
    elif args.pairs:
        rows = pair_search(pool, targets, merged, moves, natures, typechart,
                           partner_name=args.chip_from or None,
                           partner_move_name=args.chip_move or None,
                           partner_item=args.partner_item or None,
                           item_overrides=item_overrides,
                           move_overrides=move_overrides,
                           excluded_items=excluded_items)
        _print_pairs(rows, targets, args.top, partner=args.chip_from,
                    move=args.chip_move)
    elif args.chip_from:
        rows = chip_then_ko(pool, targets, args.chip_from, args.chip_move,
                            merged, moves, natures, typechart,
                            partner_item=args.partner_item or None,
                            item_overrides=item_overrides,
                            move_overrides=move_overrides,
                            excluded_items=excluded_items)
        _print_chip(rows, targets, args.chip_from, args.chip_move, args.top)
    else:
        max_taken = args.max_taken / 100.0 if args.max_taken is not None else None
        rows = threshold_search(pool, targets, merged, moves, natures, typechart,
                                threshold=threshold,
                                item_overrides=item_overrides,
                                move_overrides=move_overrides,
                                max_taken=max_taken, outspeed=args.outspeed,
                                excluded_items=excluded_items)
        _print_threshold(rows, targets, threshold, args.top,
                         max_taken=max_taken, outspeed=args.outspeed)

    if args.csv and args.deep:
        import csv
        flat = []
        for (e1, e2), d in detail.items():
            row = {"our pair": " + ".join(our_pair), "enemy lead": f"{e1} + {e2}",
                  "outcome": d["outcome"], "turns used": d["turns_used"],
                  "tailwind outcome": d["tailwind_outcome"],
                  "tailwind safe": d["tailwind_safe"],
                  "protect safe": d["protect_safe"],
                  "if first enemy protects": d["protect_outcomes"]["E1"],
                  "if second enemy protects": d["protect_outcomes"]["E2"],
                  "ohko risk": "; ".join(
                      f"{r['attacker']}'s {r['move']} on {r['target']} "
                      f"({r['hi'] * 100:.0f}%)" for r in d["ohko_risk"])}
            for (atk, tgt), h in d["grid"]["ours"].items():
                row[f"our {atk}->{tgt} avg%"] = round(h.avg * 100, 1)
            for (atk, tgt), h in d["grid"]["theirs"].items():
                row[f"their {atk}->{tgt} avg%"] = round(h.avg * 100, 1)
            flat.append(row)
        with open(args.csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(flat[0]) if flat else [])
            w.writeheader()
            w.writerows(flat)
        print(f"\nFull table ({len(flat)} rows): {os.path.abspath(args.csv)}")
    elif args.csv and args.bring4:
        import csv
        enemy_tw = enemy_has_real_tailwind(targets, merged)
        our6_sets = _fixed_sets_for(our6, {}, {}, merged, moves, natures,
                                    typechart, targets, item_overrides,
                                    move_overrides, excluded_items)
        flat = []
        for b in bring4_rows:
            wr = b["worst_pair_row"]
            depth = bring4_pair_depth(b)
            pt = depth["pairs_total"]
            n_pairs = len(b["pair_rows"])
            row = {
                "bring4": " / ".join(b["bring4"]),
                "uncovered enemy pairs": "; ".join(
                    f"{e1}+{e2}" for e1, e2 in b["uncovered_enemy_pairs"]),
                "pairs good": b["pairs_good"], "pairs total": b["pairs_total"],
                "worst pair": " + ".join(b["worst_pair"]),
                "worst pair swept": wr["pairs_swept"],
                "worst pair traded": wr["pairs_traded"],
                "worst pair lost": wr["pairs_lost"],
                "worst pair no ko": wr["pairs_no_ko"],
                "worst pair tailwind safe": wr["pairs_tailwind_safe"],
                "worst pair protect safe": wr["pairs_protect_safe"],
                "worst pair total": wr["pairs_total"],
                "pairs beaten total": f"{depth['beaten_total']}/{n_pairs * pt}",
                "pairs beaten 3rd best": f"{depth['beaten_3rd']}/{pt}",
                "pairs beaten 4th best": f"{depth['beaten_4th']}/{pt}",
                "pairs beaten worst": f"{depth['beaten_worst']}/{pt}",
                "enemy has real tailwind": enemy_tw,
                "pairs tailwind safe total": f"{depth['tailwind_safe_total']}/{n_pairs * pt}",
                "pairs protect safe total": f"{depth['protect_safe_total']}/{n_pairs * pt}",
                "pairs clean win total": (f"{depth['clean_win_total']:.1f}/"
                                         f"{n_pairs * pt * 2:.0f}"),
            }
            _paste, token = _teamsheet_export_cells(
                b["bring4"], {n: our6_sets[n] for n in b["bring4"]}, merged)
            row["teamsheet base64"] = token
            for i, pr in enumerate(b["pair_rows"], start=1):
                row[f"pair {i}"] = " + ".join(pr["pair"])
                row[f"pair {i} beaten"] = (f"{pr['pairs_swept'] + pr['pairs_traded']}"
                                           f"/{pr['pairs_total']}")
            flat.append(row)
        with open(args.csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(flat[0]) if flat else [])
            w.writeheader()
            w.writerows(flat)
        print(f"\nFull table ({len(flat)} rows): {os.path.abspath(args.csv)}")
    elif args.csv and args.multi_bring4:
        import csv
        flat = []
        for r in multi_rows:
            core = r["core"]
            weak = member_weakness_summary(core, merged)
            by_type = sorted(((t, c) for t, c in weak["per_type"].items() if c > 0),
                             key=lambda tc: -tc[1])
            types_2plus = sum(1 for c in weak["per_type"].values() if c >= 2)
            row = {"core": " / ".join(core), "core size": r["core_size"],
                  "unused": " / ".join(r["unused"]),
                  "bottleneck enemy": r["worst_enemy_idx"] + 1,
                  "weak to 2+ types (members)": weak["weak_to_2plus"],
                  "weak to 1 type (members)": weak["weak_to_1"],
                  "weak to 0 types (members)": weak["weak_to_0"],
                  "types with 2+ weak members": types_2plus,
                  "weaknesses by type": "; ".join(
                      f"{t}:{c}" for t, c in by_type),
                  "per-member weak-type counts": ", ".join(
                      f"{n}={c}" for n, c in weak["per_member"].items())}
            for e_idx, pe in enumerate(r["per_enemy"], start=1):
                wr = pe["best_bring4_row"]["worst_pair_row"]
                depth = bring4_pair_depth(pe["best_bring4_row"])
                pt = depth["pairs_total"]
                n_pairs = len(pe["best_bring4_row"]["pair_rows"])
                row[f"enemy {e_idx}"] = ", ".join(pe["target_names"])
                row[f"enemy {e_idx} best bring4"] = " / ".join(pe["best_bring4"])
                row[f"enemy {e_idx} worst pair beaten"] = (
                    f"{wr['pairs_swept'] + wr['pairs_traded']}/{wr['pairs_total']}")
                row[f"enemy {e_idx} uncovered enemy pairs"] = "; ".join(
                    f"{e1}+{e2}" for e1, e2 in
                    pe["best_bring4_row"]["uncovered_enemy_pairs"])
                row[f"enemy {e_idx} pairs beaten total"] = (
                    f"{depth['beaten_total']}/{n_pairs * pt}")
                row[f"enemy {e_idx} pairs beaten 3rd best"] = f"{depth['beaten_3rd']}/{pt}"
                row[f"enemy {e_idx} pairs beaten 4th best"] = f"{depth['beaten_4th']}/{pt}"
                row[f"enemy {e_idx} pairs beaten worst"] = f"{depth['beaten_worst']}/{pt}"
                row[f"enemy {e_idx} has real tailwind"] = enemy_has_real_tailwind(
                    pe["target_names"], merged)
                row[f"enemy {e_idx} pairs tailwind safe total"] = (
                    f"{depth['tailwind_safe_total']}/{n_pairs * pt}")
                row[f"enemy {e_idx} pairs protect safe total"] = (
                    f"{depth['protect_safe_total']}/{n_pairs * pt}")
                row[f"enemy {e_idx} pairs clean win total"] = (
                    f"{depth['clean_win_total']:.1f}/{n_pairs * pt * 2:.0f}")
                teamsheet_bits = []
                b4_sets = {}
                core_resolved = r.get("item_clause_resolved_items") or {}
                for name in pe["best_bring4"]:
                    # Read the same FIXED set every printed number was
                    # computed from (`coverage["fixed_items"/"fixed_moves"]`,
                    # or -- when this core's own pool-wide items collided --
                    # `r["item_clause_resolved_items"]`, the core-scoped
                    # re-race's own set) rather than re-deriving one per
                    # enemy -- "for a team, the moves must stay the same".
                    if name in core_resolved:
                        item = core_resolved[name]
                        move_names = coverage["fixed_moves"].get(name)
                    elif name in coverage["fixed_items"]:
                        item = coverage["fixed_items"][name]
                        move_names = coverage["fixed_moves"].get(name)
                    else:
                        item, move_names, _weather = _answer_for(
                            name, merged, moves, natures, typechart,
                            pe["target_names"], item_overrides=item_overrides,
                            move_overrides=move_overrides,
                            excluded_items=excluded_items)
                    b4_sets[name] = {"item": item, "moves": move_names or []}
                    usage_by_move = dict(merged[name].get("moves_usage") or [])
                    moves_str = ", ".join(
                        f"{m} ({usage_by_move.get(m, 0.0):.0f}%)"
                        for m in (move_names or []))
                    teamsheet_bits.append(f"{name} @ {item or '-'}: {moves_str}")
                row[f"enemy {e_idx} teamsheet"] = "; ".join(teamsheet_bits)
                _paste, token = _teamsheet_export_cells(
                    pe["best_bring4"], b4_sets, merged)
                row[f"enemy {e_idx} teamsheet base64"] = token
            flat.append(row)
        with open(args.csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(flat[0]) if flat else [])
            w.writeheader()
            w.writerows(flat)
        print(f"\nFull table ({len(flat)} rows): {os.path.abspath(args.csv)}")
    elif args.csv:
        import csv
        flat = []
        for r in rows:
            if "pair" in r:
                row = {"name": " + ".join(r["pair"]), "item": None,
                      "item1": r["item1"], "item2": r["item2"]}
            else:
                row = {"name": r["name"], "item": r.get("item")}
            if "priority" in r:
                row["priority"] = r["priority"]
                row["priority move"] = r["priority_move"]
                row["speed"] = round(r["speed"], 1)
                row["target"] = r["name"] in targets
            elif "per_target" in r:
                for t, h in r["per_target"].items():
                    row[f"{t} move"] = h.move_name
                    row[f"{t} lo%"] = round(h.lo * 100, 1)
                    row[f"{t} avg%"] = round(h.avg * 100, 1)
                    row[f"{t} hi%"] = round(h.hi * 100, 1)
                row["worst %"] = round(r["worst_pct"] * 100, 1)
                row["meets all"] = r["meets_all"]
                if "incoming" in r:
                    for t, h in r["incoming"].items():
                        row[f"{t} incoming move"] = h.move_name
                        row[f"{t} incoming hi%"] = round(h.hi * 100, 1)
                    row["outspeeds all (natural)"] = all(r["outspeeds"].values())
                    row["outspeeds all (scarf)"] = all(r["outspeeds_scarf"].values())
            elif "finishes" in r:
                for t, (ko, h) in r["finishes"].items():
                    row[f"{t} KO"] = ko
                    row[f"{t} move"] = h.move_name
                    row[f"{t} lo%"] = round(h.lo * 100, 1)
                    row[f"{t} avg%"] = round(h.avg * 100, 1)
                    row[f"{t} hi%"] = round(h.hi * 100, 1)
                row["n_ko"] = r["n_ko"]
            elif "pairs_swept" in r:
                row["pairs swept"] = r["pairs_swept"]
                row["pairs traded"] = r["pairs_traded"]
                row["pairs lost"] = r["pairs_lost"]
                row["pairs no KO"] = r["pairs_no_ko"]
                row["pairs tailwind-safe"] = r["pairs_tailwind_safe"]
                row["pairs protect-safe"] = r["pairs_protect_safe"]
                row["pairs clean win total"] = round(r["pairs_clean_win_total"], 2)
                row["pairs total"] = r["pairs_total"]
            else:
                row["pairs clean"] = r["pairs_clean"]
                row["pairs trade"] = r["pairs_trade"]
                row["pairs no KO"] = r["pairs_no_ko"]
                row["pairs pinned"] = r["pairs_pinned"]
                row["pairs total"] = r["pairs_total"]
            flat.append(row)
        with open(args.csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(flat[0]) if flat else [])
            w.writeheader()
            w.writerows(flat)
        print(f"\nFull table ({len(flat)} rows): {os.path.abspath(args.csv)}")


if __name__ == "__main__":
    main()
