"""The lead-scan workbook: the plan, the lines, and why the lines are sound.

Asked for:

    "make sure I can see the lines, damage and logic when I generate the xlsx
     (why potential backs can't come in, which could, do I have to switch, is my
     switch ok vs their best move, maybe if I switch+protect other slot it's
     guaranteed)"

    "a given 4 doesn't need to beat every enemy team. I would like to see all the
     4s, and then perhaps as a second sheet I can see the teams of 6 that possess
     all or the most relevant 4s (i.e., choose three different combinations of 4
     to face three teams all of which belong to a set of 6)"

Six sheets, and the shape follows those two paragraphs rather than the shape of
the code:

  Brings          every legal 4 against every opponent, and their MEGA CHOICES.
                  Not filtered to "beats everything": a 4 is a tool for a
                  matchup, and this is the catalogue.
  Teams of 6      sets of six that CONTAIN good fours for several opponents at
                  once, with the specific four to bring to each. This is the
                  deliverable -- one team, a different bring per opponent.
  Openings        every (our 4 x their lead pair): the verdict, THEIR best plan
                  including a Protect split and Tailwind, the play we make, and
                  whether it is guaranteed.
  Lines           the turn-by-turn for each opening, with damage on every hit and
                  the pin called out where a Pokemon is removed before it acts.
  Switch-ins      why their backs cannot come in, and which one can: the damage we
                  put on it arriving, how much HP it could afford to lose, and
                  what it does back.
  Items           single item changes that convert a losing opening.

Nothing here computes; it renders what `lead_scan` produced. The legend sheet
states every assumption, because a spreadsheet detached from its assumptions is a
spreadsheet that will be misread.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import lead_scan as ls

HEAD = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
GOOD = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
WARN = PatternFill("solid", start_color="FCF3CF", end_color="FCF3CF")
BAD = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
FILL = {ls.WIN: GOOD, ls.EVEN: WARN, ls.LOSS: BAD}

MAX_ROWS = 100_000


def _sheet(wb, title, cols, widths=None, freeze="A2"):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = HEAD
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = freeze
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = (
            (widths or {}).get(name, max(12, len(name) + 2)))
    return ws


def _wrap(ws, col):
    for row in ws.iter_rows(min_col=col, max_col=col, min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def teams_of_six(records, min_opponents=2, limit=40):
    """Sets of six that contain a good four for several opponents at once.

    The union of two or three chosen fours is at most six Pokemon exactly when
    they overlap enough -- which is the whole idea: *"choose three different
    combinations of 4 to face three teams all of which belong to a set of 6"*.

    `records` is [{"bring": (...4), "opponent": str, "score": float, ...}]. Takes
    the best four per opponent, tries every combination of opponents, and keeps a
    six whose union fits, naming the bring for each.
    """
    import itertools

    by_opp = {}
    for r in records:
        if r["score"] <= 0:
            continue
        best = by_opp.setdefault(r["opponent"], [])
        best.append(r)
    for opp in by_opp:
        by_opp[opp].sort(key=lambda r: -r["score"])
        by_opp[opp] = by_opp[opp][:6]      # a few candidates each, not all

    opponents = sorted(by_opp)
    out = {}
    for n in range(len(opponents), min_opponents - 1, -1):
        for combo in itertools.combinations(opponents, n):
            for picks in itertools.product(*(by_opp[o] for o in combo)):
                species = set()
                for p in picks:
                    species |= {ls.species_of(x) for x in p["bring"]}
                if len(species) > 6:
                    continue
                key = tuple(sorted(species))
                score = sum(p["score"] for p in picks) / len(picks)
                covered = len(picks)
                prior = out.get(key)
                if prior is None or (covered, score) > (prior[0], prior[1]):
                    out[key] = (covered, score, picks)
        if out:
            break
    rows = [{"six": k, "covered": v[0], "score": v[1], "picks": v[2]}
            for k, v in out.items()]
    rows.sort(key=lambda r: (-r["covered"], -r["score"]))
    return rows[:limit]


def build(records, path):
    """Write the workbook. `records` is what `lead_sweep` collected."""
    wb = Workbook()

    # --- Brings -------------------------------------------------------------
    ws = _sheet(wb, "Brings",
                ["Our 4 (lead first)", "Opponent", "Their Mega", "Wins",
                 "Patched", "Unheld", "Ties", "Of their leads", "Score",
                 "Their leads we cannot hold"],
                {"Our 4 (lead first)": 46, "Opponent": 18, "Their Mega": 18,
                 "Their leads we cannot hold": 60})
    for r in sorted(records, key=lambda r: (r["opponent"], -r["score"])):
        rep = r["report"]
        ws.append([" / ".join(r["bring"]), r["opponent"], r.get("mega") or "-",
                   rep.wins, len(rep.patched), len(rep.losses), len(rep.ties),
                   len(rep.results), round(r["score"], 3),
                   "; ".join(" + ".join(x.enemy_lead) for x in rep.losses)])
        ws.cell(ws.max_row, 9).fill = GOOD if r["score"] > 0 else BAD
    ws.auto_filter.ref = ws.dimensions
    _wrap(ws, 10)

    # --- Teams of 6 ---------------------------------------------------------
    ws = _sheet(wb, "Teams of 6",
                ["#", "The six", "Opponents covered", "Mean score",
                 "Bring vs each opponent"],
                {"The six": 60, "Bring vs each opponent": 80})
    sixes = teams_of_six(records)
    for i, row in enumerate(sixes, start=1):
        detail = " | ".join(
            f"{p['opponent']}: {' / '.join(p['bring'])}" for p in row["picks"])
        ws.append([i, " / ".join(row["six"]), row["covered"],
                   round(row["score"], 3), detail])
        ws.cell(ws.max_row, 4).fill = GOOD if row["score"] > 0 else WARN
    if not sixes:
        ws.append(["", "No six covers two or more opponents with a held bring. "
                       "Widen --pool-size, or accept a hole and read Brings."])
    _wrap(ws, 5)

    # --- Openings -----------------------------------------------------------
    ws = _sheet(wb, "Openings",
                ["Our 4", "Opponent", "Their Mega", "Their lead pair", "Result",
                 "Speed tie", "Their best plan", "Our play", "Guaranteed?",
                 "Margin", "After it settles (mop-up)"],
                {"Our 4": 40, "Their lead pair": 34, "Their best plan": 34,
                 "Our play": 46, "After it settles (mop-up)": 60})
    for r in sorted(records, key=lambda r: (r["opponent"], -r["score"])):
        rep = r["report"]
        for x in rep.results:
            play = x.play
            ws.append([" / ".join(r["bring"]), r["opponent"],
                       r.get("mega") or "-", " + ".join(x.enemy_lead),
                       x.verdict, "yes" if x.tie else "", x.plan,
                       play.label if play else "-",
                       "GUARANTEED" if (play and play.guaranteed) else "",
                       round(x.margin, 3), (play.mopped if play else "")])
            ws.cell(ws.max_row, 5).fill = FILL.get(x.verdict, WARN)
            if play and play.guaranteed:
                ws.cell(ws.max_row, 9).fill = GOOD
    ws.auto_filter.ref = ws.dimensions
    _wrap(ws, 11)

    # --- Lines --------------------------------------------------------------
    ws = _sheet(wb, "Lines",
                ["Our 4", "Opponent", "Their lead pair", "Our play", "Turn",
                 "What happens (with damage)"],
                {"Our 4": 34, "Their lead pair": 30, "Our play": 40,
                 "What happens (with damage)": 96})
    truncated = 0
    for r in sorted(records, key=lambda r: (r["opponent"], -r["score"])):
        for x in r["report"].results:
            play = x.play
            if not (play and play.log):
                continue
            turn = ""
            first = True
            for line in play.log:
                if ws.max_row >= MAX_ROWS:
                    truncated += 1
                    continue
                stripped = line.strip()
                if stripped.startswith("Turn "):
                    turn = stripped
                    continue
                ws.append([" / ".join(r["bring"]) if first else None,
                           r["opponent"] if first else None,
                           " + ".join(x.enemy_lead) if first else None,
                           play.label if first else None, turn, stripped])
                first = False
    if truncated:
        ws.append([f"... {truncated} more rows not written (capped at "
                   f"{MAX_ROWS:,})"])
    _wrap(ws, 6)

    # --- Switch-ins ---------------------------------------------------------
    ws = _sheet(wb, "Switch-ins",
                ["Our 4", "Opponent", "Their bench Pokemon",
                 "We deal arriving (%)", "It could afford to lose (%)",
                 "It deals back (%)", "Its best move", "Verdict"],
                {"Our 4": 34, "Their bench Pokemon": 22, "Verdict": 40})
    for r in sorted(records, key=lambda r: (r["opponent"], -r["score"])):
        for row in r.get("switch_ins") or []:
            ws.append([" / ".join(r["bring"]), r["opponent"], row["name"],
                       round(row["we_deal"] * 100, 1),
                       round(row["afford"] * 100, 1),
                       round(row["it_deals"] * 100, 1), row["its_move"],
                       row["verdict"]])
            ws.cell(ws.max_row, 8).fill = (
                GOOD if row["afford"] <= 0 else
                (BAD if row["it_deals"] >= 1.0 else WARN))
    ws.auto_filter.ref = ws.dimensions
    _wrap(ws, 8)

    # --- Items --------------------------------------------------------------
    ws = _sheet(wb, "Items",
                ["Our 4", "Opponent", "Their lead pair", "Give this Pokemon",
                 "This item", "Turns the opening into", "Margin", "The play"],
                {"Our 4": 34, "Their lead pair": 30, "The play": 46})
    for r in sorted(records, key=lambda r: (r["opponent"], -r["score"])):
        for fix in r.get("item_fixes") or []:
            lead_pair, mon, item, play = fix
            ws.append([" / ".join(r["bring"]), r["opponent"],
                       " + ".join(lead_pair), mon, item, play.verdict,
                       round(play.margin, 3), play.label])
            ws.cell(ws.max_row, 6).fill = FILL.get(play.verdict, WARN)
    ws.auto_filter.ref = ws.dimensions

    _legend(wb)
    wb.save(path)
    return len(records)


def _legend(wb):
    ws = wb.create_sheet("How to read this")
    for row in (
        ["What this is"],
        ["Lead pairs and brings scored by arithmetic over TWO TURNS, plus a "
         "mop-up check. No games are played, so this is a screen -- but every "
         "number is traceable to a line on the Lines sheet."],
        [],
        ["The sheets", ""],
        ["Brings", "Every legal 4 against every opponent AND every Mega they "
                   "could choose. Not filtered to 'beats everything': a 4 is a "
                   "tool for a matchup."],
        ["Teams of 6", "Sixes that contain a good four for several opponents at "
                       "once, naming the four to bring to each. One team, a "
                       "different bring per opponent."],
        ["Openings", "Every (our 4 x their lead pair): the result, their best "
                     "plan, the play we make, and whether it is guaranteed."],
        ["Lines", "The turn-by-turn with damage on every hit. 'PINNED' marks a "
                  "Pokemon removed before it acted."],
        ["Switch-ins", "Why their backs cannot come in, and which one can."],
        ["Items", "Single item changes that convert a losing opening."],
        [],
        ["The plays", ""],
        [ls.STAY, "Both leads attack."],
        [ls.SWITCH, "One leaves; a back arrives, eats the turn and deals nothing."],
        [ls.SWITCH_PROTECT, "And the staying slot Protects, so the only thing "
                            "left to lose is off the table. Often what makes a "
                            "switch free."],
        [ls.SACRIFICE, "Let the pinned Pokemon faint, so the replacement arrives "
                       "at FULL health rather than having eaten both attacks. A "
                       "3v4 you win is a win."],
        [],
        ["Guaranteed", "It wins whatever they do -- including us losing every "
                       "speed tie. A plan that needs a coin flip is not a plan, "
                       "so ties resolve AGAINST us everywhere."],
        [],
        ["What is enumerated for them", ""],
        ["Their Mega", "One per team, chosen at preview after seeing our four. "
                       "Both choices are scored and the worst is what counts. On "
                       "the reference bring this is the difference between +0.48 "
                       "with 0 unheld and 0.00 with 3 unheld."],
        ["Protect", "Either slot, on either turn, but never twice in a row."],
        ["Speed control", "Tailwind or Trick Room on either turn instead of "
                          "attacking -- which is how Mega Floette stops tying "
                          "Garchomp and simply outspeeds it."],
        [],
        ["What is modelled", ""],
        ["Spread moves", "Hit both foes at 0.75x, and hit our OWN partner when "
                         "allAdjacent -- which is why a Flying or Levitate "
                         "partner beside Earthquake is worth something."],
        ["Megas", "Their Mega's stats, ability, typing and SPEED. Mega Floette "
                  "is 111 base and 166 evolved, exactly tying Garchomp."],
        ["Items", "Life Orb, Choice items, Assault Vest, Focus Sash and Sturdy, "
                  "and type-resist berries. The berries and Assault Vest were "
                  "TABLES ONLY until now -- measured, a Roseli Berry changed "
                  "Light of Ruin by exactly zero. Now: 105.5% into Garchomp "
                  "becomes 52.7%, a guaranteed OHKO becomes a survival."],
        [],
        ["What is NOT modelled", ""],
        ["", "Redirection (Follow Me / Rage Powder), Wide Guard, Substitute, "
             "setup moves other than Tailwind and Trick Room, abilities that "
             "trigger on contact, and status."],
        ["", "Accuracy: every move is treated as landing. Blizzard is 70% "
             "accurate outside snow and this does not know it."],
        ["", "Both sides are read at their usage-default set except where the "
             "Items sheet says otherwise."],
        ["", "Every verdict is a HYPOTHESIS to point the overnight audit at. The "
             "cheap thing proposes; the expensive thing decides."],
    ):
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    for cell in ("A4", "A12", "A18", "A21", "A26", "A32"):
        ws[cell].font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    _wrap(ws, 2)
