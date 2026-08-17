"""Find the overwhelming lead pair, one enemy at a time, with arithmetic only.

The worked example this is built from, and the acceptance test for it:

    "vs big 6, you only bring 4. There is one lead pair which has an
     overwhelming advantage: (focus sash) ninetales-alola + (life orb) garchomp
     with mega scizor and a flying/levitate pokemon in the back to ignore
     garchomp partner earthquake dmg. ... There is nothing which can really
     afford to switch in, so its a general pin given the spread damage output
     and speed tiers.

     1. Mega Charizard Y: outsped, Garchomp KOs with rock slide.
     2. Garchomp: outsped, Ninetales KOs with Blizzard.
     3. Kingambit, Basculegion: Garchomp earthquake + Ninetales-Alola Blizzard
        or Freeze Dry is very threatening, and will then collapse into a win by
        turn 2.
     4. Mega Floette, Whimsicott: either Garchomp earthquake + Ninetales
        blizzard to KO, or ... Garchomp switches to Scizor (or a specially
        defensive fairy resist who can outspeed/ohko in return, which Scizor can
        even in tailwind) while Ninetales-Alola uses Blizzard to bring Whimsicott
        to sash."

Read that list again and notice its SHAPE. It is six independent questions, each
answered by one of four verdicts, and none of them needs a game to be played:

    OUTSPED     we move first and one of ours guarantees the KO       (a pin)
    FOCUSED     both of ours together guarantee it this turn
    TWO TURNS   it dies across this turn and next before acting twice
    SWITCH      the back has something that walls it and out-trades
    PROBLEM     none of the above

That is exactly the arithmetic `pin.py` already does, applied across their whole
roster instead of one position -- and it costs a ThreatMatrix rather than a
search. The expensive pipeline plays games; this one adds numbers, so it can
sweep thousands of lead pairs in the time the audit spends on one pairing.

WHAT IT IS FOR. Narrowing. The full search rates a bring against 90 enemy
configurations by playing them out; before that is affordable you need to know
which few hundred of the hundreds of thousands of lead pairs are worth the
night. This answers "does this pair leave anything unanswered", across several
opponents at once, because:

    "There just needs to be an overlap across the different teams (bring 4 to
     each), with goodish type synergy, stats, damage output."

so the ranking is by the WORST opponent, not the mean. A pair that dismantles
five teams and has no answer to the sixth is not the pair.

WHICH OF THEIR POKEMON HOLDS THE MEGA SLOT CHANGES THE ANSWER, and this is the
sharpest known limitation. One Mega per team, chosen at preview, and the dataset
assigns it -- so a scan silently answers the question for ONE of their possible
choices. Measured on the reference case: base Mega Floette is speed 111 and its
Mega is 166, exactly tying Garchomp. On Big 6 as loaded, Charizard Y holds the
slot, Floette stays at 111 and Garchomp outspeeds it cleanly; if they Mega
Floette instead it is a coin flip, which is precisely the caution in the worked
example -- *"mega floette OHKOs garchomp if it wins speed tie"*. `mega_slots()`
enumerates their choices so a pair can be scanned against each, and `worst_of`
ranks on the worst, because they pick after seeing your team.

WHAT IT DELIBERATELY DOES NOT DO. It does not play the game, so it cannot see
Protect stalling, redirection, Wide Guard, a Substitute, setup, or anything
whose value is in a sequence rather than a sum. It reads the usage-default set
for both sides. It assumes we get to focus both attacks where we like, which the
enemy's own attacks may deny. Every verdict here is a HYPOTHESIS the audit
should be pointed at -- which is the division of labour, not a shortcoming: the
cheap thing proposes, the expensive thing decides.
"""
import re
from dataclasses import dataclass, field

from damage import hits_ally, is_spread_move
from pin import _incoming, _kills_outright, _moves_first, mega_bulk_factor
from species_data import base_form_name

# Verdicts, best first. Order matters: `scan` reports the FIRST that applies, so
# "we outspeed and OHKO it" is preferred over "we can grind it down", even
# though both are true, because the report should name the cleanest answer.
OUTSPED = "outsped"
FOCUSED = "focused"
TWO_TURNS = "two turns"
SWITCH = "switch"
PROBLEM = "PROBLEM"

ANSWERED = (OUTSPED, FOCUSED, TWO_TURNS, SWITCH)

# How good each verdict is, for the score. An outspeed-and-KO is worth more than
# a two-turn grind because it costs nothing and cannot be interrupted; a switch
# is worth least of the answers because it spends a turn and concedes momentum.
WEIGHT = {OUTSPED: 1.0, FOCUSED: 0.9, TWO_TURNS: 0.6, SWITCH: 0.45, PROBLEM: 0.0}

# What a verdict is worth when the enemy PRE-EMPTS: it moves before both of our
# leads and guarantees the KO on one of them. We still remove it, but we paid a
# Pokemon for the privilege, so it is a trade rather than a pin.
#
# This factor is why the score discriminates at all. Without it, 239 of 276
# swept pairs "answered every Pokemon" -- because focus-firing two attacks into
# one target while the other target politely does nothing answers almost
# anything. A screen that passes 87% of candidates is not narrowing.
PREEMPT_FACTOR = 0.35


@dataclass
class EnemyVerdict:
    """What happens to one of their Pokemon if it is on the field turn 1."""
    enemy: str
    verdict: str
    by: str = ""             # which of ours, and how
    incoming: float = 0.0    # what it does back, as a fraction of our HP
    note: str = ""
    # It moves before BOTH our leads and guarantees a KO on one of them. We may
    # still answer it, but not before it has taken a Pokemon.
    preempts: bool = False

    @property
    def answered(self) -> bool:
        return self.verdict in ANSWERED

    @property
    def weight(self) -> float:
        w = WEIGHT[self.verdict]
        return w * PREEMPT_FACTOR if self.preempts else w


@dataclass
class LeadReport:
    lead: tuple
    back: tuple
    opponent: str
    verdicts: list = field(default_factory=list)

    @property
    def problems(self):
        return [v for v in self.verdicts if not v.answered]

    @property
    def covered(self) -> int:
        return sum(1 for v in self.verdicts if v.answered)

    @property
    def score(self) -> float:
        """Mean verdict weight, zero if anything is unanswered.

        HARD zero, not a penalty. The question being asked is "can this lead
        hold up against ANY of theirs", and a pair with one hole does not answer
        it -- averaging the hole away is how a lead that loses to one Pokemon
        ends up recommended.
        """
        if not self.verdicts or self.problems:
            return 0.0
        return sum(v.weight for v in self.verdicts) / len(self.verdicts)


def _answers_from_back(matrix, back, enemy, ours_on_field):
    """A back Pokemon that walls `enemy` and out-trades it, or None.

    The role the user described, generalised: *"Scizor is a standin for a fairy
    resist who can switch in (bulky enough) and out-trade and/or outspeed
    floette and a partner."* So the test is not "resists the type" -- it is
    survives the hit, and then wins the exchange.
    """
    for b in back:
        # It has to live the switch-in. Mega bulk counts, since it evolves as it
        # lands.
        taken = matrix.threat(enemy, b).dmg_max * mega_bulk_factor(b)
        if taken >= 1.0:
            continue
        if not _kills_outright(matrix, b, enemy):
            continue
        # And it has to get to fire: either it is faster, or it survives a second
        # hit while it does.
        if _moves_first(matrix, b, enemy) or taken * 2 < 1.0:
            return b
    return None


def scan(matrix, battle, lead, back, opponent_name=""):
    """One lead pair against one enemy roster. Pure arithmetic, no simulation.

    `lead` and `back` are Combatants from `battle.p1`; the enemy roster is
    everything on `battle.p2`.
    """
    enemies = [c for c in matrix.theirs if not c.fainted]
    report = LeadReport(lead=tuple(c.name for c in lead),
                        back=tuple(c.name for c in back),
                        opponent=opponent_name)
    for e in enemies:
        worst_back = max((matrix.threat(e, o).dmg_max for o in lead),
                         default=0.0)
        # Does it take one of ours with it? Guaranteed KO, and ahead of BOTH of
        # our leads -- ahead of only one is not a pre-empt, since the other still
        # fires. This is the cost side the first version left out entirely.
        preempts = any(_kills_outright(matrix, e, o) and _moves_first(matrix, e, o)
                       for o in lead) and all(_moves_first(matrix, e, o)
                                              for o in lead)

        # 1. OUTSPED: one of ours moves first and guarantees the KO. The user's
        #    cases 1 and 2 -- "outsped, Garchomp KOs with rock slide".
        solo = [o for o in lead
                if _kills_outright(matrix, o, e) and _moves_first(matrix, o, e)]
        if solo:
            report.verdicts.append(EnemyVerdict(
                e.name, OUTSPED, by=f"{solo[0].name} {matrix.threat(solo[0], e).move}",
                incoming=worst_back, preempts=preempts))
            continue

        # 2. FOCUSED: both together remove it this turn. Case 3 -- "Garchomp
        #    earthquake + Ninetales-Alola Blizzard is very threatening".
        if len(lead) >= 2 and matrix.joint_kos(lead[0], lead[1], e,
                                               guaranteed=True):
            report.verdicts.append(EnemyVerdict(
                e.name, FOCUSED,
                by=f"{lead[0].name} + {lead[1].name}", incoming=worst_back,
                preempts=preempts))
            continue

        # 3. TWO TURNS: it dies across this turn and next, before acting twice.
        #    "will then collapse into a win by turn 2."
        t1 = _incoming(matrix, lead, e)
        t2 = _incoming(matrix, lead, e, only_before_it_acts=True)
        if t1 + t2 >= 1.0:
            report.verdicts.append(EnemyVerdict(
                e.name, TWO_TURNS,
                by=f"{lead[0].name} + {lead[1].name}", incoming=worst_back,
                preempts=preempts,
                note=f"{t1 * 100:.0f}% then {t2 * 100:.0f}%"))
            continue

        # 4. SWITCH: the back answers it. Case 4 -- "Garchomp switches to Scizor
        #    ... while Ninetales-Alola uses Blizzard".
        answer = _answers_from_back(matrix, back, e, lead)
        if answer is not None:
            report.verdicts.append(EnemyVerdict(
                e.name, SWITCH, by=answer.name, incoming=worst_back,
                preempts=preempts,
                note=f"walls it and answers with "
                     f"{matrix.threat(answer, e).move}"))
            continue

        report.verdicts.append(EnemyVerdict(
            e.name, PROBLEM, incoming=worst_back,
            note=f"nothing outspeeds it, focus fire leaves "
                 f"{max(0.0, 1.0 - t1) * 100:.0f}%, no back answers it"))
    return report


def _legal_default_sets(our4, world):
    """{name: {"item": ...}} pinning OUR bring to its most-used LEGAL item.

    `_harness.setup_battle` reads `make_team`'s raw usage default when no
    `sets` override is given, which does not filter Regulation MB's banned
    items (Assault Vest, the three Choice items) -- so a Pokemon whose #1
    recorded item happens to be one of those would otherwise field it in
    every cheap-stage call below. `default_item` already skips them; this
    just turns that into an override for every member of the bring.
    """
    return {name: {"item": default_item(name, world)} for name in our4}


def scan_bring(our4, enemy_roster, world, lead_index=(0, 1), opponent_name=""):
    """Convenience wrapper: build the position, scan it, return the report.

    `our4` is the bring in lead-first order, so `lead_index` almost never needs
    changing.
    """
    from _harness import setup_battle
    from threat import build_threat_matrix
    b, ms = setup_battle(list(our4), list(enemy_roster), world,
                         sets=_legal_default_sets(our4, world))
    matrix = build_threat_matrix(b, ms)
    ours = [c for c in b.p1.roster if not c.fainted]
    lead = [ours[i] for i in lead_index]
    back = [c for c in ours if c not in lead]
    return scan(matrix, b, lead, back, opponent_name)


def describe(report):
    """The user's own breakdown format: one numbered line per enemy."""
    lines = [f"{' + '.join(report.lead)}  (back: {', '.join(report.back)})"
             f"  vs {report.opponent}",
             f"  {report.covered}/{len(report.verdicts)} answered, "
             f"score {report.score:.2f}"]
    for i, v in enumerate(report.verdicts, start=1):
        bits = [f"  {i}. {v.enemy}: {v.verdict.upper()}"]
        if v.by:
            bits.append(f"by {v.by}")
        if v.note:
            bits.append(f"({v.note})")
        lines.append("  ".join(bits))
    return lines


def mega_slots(enemy_roster, world):
    """Their possible Mega choices: the names on this roster that can Mega.

    One Mega per team, and they choose it at PREVIEW -- after seeing your four.
    So a lead pair that beats a team when Charizard holds the slot and loses when
    Floette does has not beaten the team, and `worst_of` is the honest reading.
    """
    merged = world["merged"]
    out = []
    for name in enemy_roster:
        rec = merged.get(name) or {}
        if rec.get("mega_stats") or rec.get("is_mega") or name.startswith("Mega "):
            out.append(name)
    return out


def worst_of(reports):
    """The report that matters: their best answer to our pair.

    They choose their Mega, their bring and their lead after seeing our four, so
    a pair is worth what its WORST case is worth, not its average.
    """
    return min(reports, key=lambda r: (r.score, -len(r.problems))) if reports else None


def write_workbook(rows, path, opponents=(), pool_size=0):
    """The sweep as an .xlsx: a ranking sheet, a per-enemy detail sheet, a legend.

    The detail sheet is the one that matters. A score says a pair works; only the
    enemy-by-enemy breakdown says WHY, and that is the form the idea was
    expressed in -- six numbered lines, one per Pokemon.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
    good = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
    warn = PatternFill("solid", start_color="FCF3CF", end_color="FCF3CF")
    bad = PatternFill("solid", start_color="FADBD8", end_color="FADBD8")
    fill_for = {WIN: good, EVEN: warn, LOSS: bad,
                OUTSPED: good, FOCUSED: good, TWO_TURNS: warn, SWITCH: warn,
                PROBLEM: bad}

    def _header(ws, cols, freeze="A2"):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = head
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = freeze
        for idx, name in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(12, len(name) + 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead pairs"
    _header(ws, ["#", "Lead A", "Lead B", "Back A", "Back B", "Worst opponent",
                 "Wins", "Patched", "Unheld", "Ties", "Of their leads",
                 "Score", "Their leads we cannot hold"])
    for i, r in enumerate(rows, start=1):
        worst = r["worst"]
        ws.append([i, r["lead"][0], r["lead"][1], r["back"][0], r["back"][1],
                   worst.opponent, worst.wins, len(worst.patched),
                   len(worst.losses), len(worst.ties), len(worst.results),
                   round(r["score"], 3),
                   "; ".join(" + ".join(x.enemy_lead) for x in worst.losses)])
        ws.cell(ws.max_row, 12).fill = good if r["score"] > 0 else bad
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.auto_filter.ref = ws.dimensions

    det = wb.create_sheet("Their lead pairs")
    _header(det, ["Our lead", "Our back", "Opponent", "Their lead pair",
                  "Result", "Their best plan", "Speed tie", "We lose",
                  "They lose", "Margin", "Patch: switch", "Patch: to",
                  "Patch margin"])
    for r in rows:
        for rep in r["reports"]:
            for n, x in enumerate(rep.results, start=1):
                det.append([" + ".join(r["lead"]) if n == 1 else None,
                            " + ".join(r["back"]) if n == 1 else None,
                            rep.opponent if n == 1 else None,
                            " + ".join(x.enemy_lead), x.verdict, x.plan,
                            "yes" if x.tie else "", x.our_dead, x.their_dead,
                            round(x.margin, 3),
                            x.patch[0] if x.patch else None,
                            x.patch[1] if x.patch else None,
                            round(x.patch[2], 3) if x.patch else None])
                det.cell(det.max_row, 5).fill = fill_for.get(x.verdict, warn)
    det.column_dimensions["A"].width = 36
    det.column_dimensions["B"].width = 36
    det.column_dimensions["D"].width = 36

    legend = wb.create_sheet("How to read this")
    for row in (
        ["What this is"],
        ["Lead pairs ranked by whether ANY of the opponent's Pokemon is left "
         "unanswered. No games are played -- every verdict is arithmetic on a "
         "threat matrix, so this sweeps thousands of pairs in the time the "
         "overnight audit spends on one pairing."],
        [],
        ["Opponents", ", ".join(opponents)],
        ["Lead pool size", pool_size],
        [],
        ["Result", "Means"],
        [WIN, "After this turn and next they are down more Pokemon than we are "
              "(or level on Pokemon and ahead on HP lost)."],
        [EVEN, "Level within a quarter of a Pokemon."],
        [LOSS, "We are behind. One of these zeroes the pair's score."],
        [],
        ["Margin", "(their HP lost) - (our HP lost), in Pokemon-equivalents, "
                   "over two turns of both sides focus-firing."],
        ["Score", "Mean margin across all 15 of their lead pairs, or a HARD "
                  "ZERO if any single one beats us."],
        ["Why the race", "An earlier version asked only whether we could remove "
                         "each of their Pokemon, and 235 of 275 swept pairs "
                         "passed -- two attackers focused on one target while "
                         "its partner does nothing answers almost anything. "
                         "Pricing the exchange took it to 31 of 275."],
        ["Species clause", "Four distinct BASE species, at most one Mega. The "
                           "sweep's top answer was once Garchomp + Mega "
                           "Garchomp. A screen that recommends an illegal team "
                           "is worse than no screen."],
        [],
        ["Speed ties"],
        ["", "A tie is a coin flip and is NEVER counted as a win. The race "
             "resolves every tie AGAINST us, and a tied opening is sent to the "
             "patch search exactly as a loss is: a plan that needs to win a "
             "flip is a plan with a hole."],
        [],
        ["The patch (nested)"],
        ["", "An opening we do not win is not automatically a hole. If a BACK "
             "Pokemon converts it -- switch in turn 1 eating the damage, deal "
             "nothing, then attack from turn 2 -- the opening is HELD and the "
             "back slot has a job. That is the same two-turn pin question one "
             "level down, which is what makes the method recursive."],
        ["Unheld", "Openings we neither win nor patch. These are the holes, and "
                   "any one of them zeroes the score."],
        [],
        ["Their plan"],
        ["", "Their side may split: one attacks while the other Protects, "
             "chipping us while keeping a Pokemon whole. Every result takes "
             "THEIR BEST plan of {both attack, left protects, right protects}, "
             "so a claim that survives only the obliging column is not counted."],
        ["Ranked on the worst", "They choose their Mega, their four and their "
                               "lead AFTER seeing your four, so a pair is worth "
                               "what its worst case is worth."],
        [],
        ["Known limitations"],
        ["The Mega slot", "One Mega per team, chosen at preview. The dataset "
                          "assigns it, so a scan answers for ONE of their "
                          "choices. Reference case: base Mega Floette is speed "
                          "111 and its Mega is 166, exactly tying Garchomp -- "
                          "so which Pokemon holds their slot decides whether "
                          "that is a clean outspeed or a coin flip."],
        ["No sequences", "Protect stalling, redirection, Wide Guard, "
                         "Substitute and setup are all invisible to a sum."],
        ["Usage sets", "Both sides are read at their usage-default set."],
        ["Free focus fire", "It assumes we aim both attacks where we like, "
                            "which their own attacks may deny."],
        ["", "Every verdict is a HYPOTHESIS to point the audit at. The cheap "
             "thing proposes; the expensive thing decides."],
    ):
        legend.append(row)
    legend["A1"].font = Font(bold=True, size=13)
    for cell in ("A7", "A18"):
        legend[cell].font = Font(bold=True)
    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 94
    for row in legend.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return len(rows)


# --- the race ---------------------------------------------------------------
# Everything above answers "can we remove this ONE Pokemon", which turned out
# not to discriminate: 235 of 275 swept pairs answered every Pokemon on every
# opponent, because two attackers focused on one target while its partner does
# nothing answers almost anything. The missing half is the cost. What decides a
# matchup is the 2v2 EXCHANGE -- over this turn and the next, who runs out of
# Pokemon first -- and that is still only arithmetic.

WIN, EVEN, LOSS = "win", "even", "loss"


def _focus_target(matrix, attackers, defenders):
    """Which of `defenders` this side focuses: the one it removes soonest.

    A declared heuristic, not an optimisation. Ranked by combined guaranteed
    damage as a fraction of CURRENT hp, so "the one already in range" beats "the
    one we do more raw damage to" -- which is the focus-fire logic the whole idea
    rests on: chip into a partner's kill.
    """
    best = None
    for d in defenders:
        share = sum(matrix.threat(a, d).dmg_min for a in attackers)
        need = d.current_hp / (d.max_hp() or 1)
        # Fraction of the way to a removal, capped so an overkill does not beat a
        # clean kill on a second target.
        progress = min(1.0, share / need) if need > 0 else 1.0
        key = (progress, share)
        if best is None or key > best[0]:
            best = (key, d)
    return best[1] if best else None


def _turn(matrix, ours, theirs, hp):
    """Resolve one turn as arithmetic. Mutates `hp` (fractions remaining).

    Speed-ordered, and a Pokemon whose HP has already reached zero this turn
    does not act -- which is the pin, expressed as the only thing it really is:
    an attack that never happens.
    """
    live_ours = [c for c in ours if hp[id(c)] > 0]
    live_theirs = [c for c in theirs if hp[id(c)] > 0]
    if not live_ours or not live_theirs:
        return
    our_target = _focus_target(matrix, live_ours, live_theirs)
    their_target = _focus_target(matrix, live_theirs, live_ours)

    order = []
    for side, mine, foes, target in (("us", live_ours, live_theirs, our_target),
                                     ("them", live_theirs, live_ours, their_target)):
        for c in mine:
            # Sort key: priority first, then speed, then SIDE -- and the side
            # tiebreak resolves against us on purpose.
            #
            #     "Ideally your strategy does not rely on winning speed ties."
            #
            # A raw (priority, speed) sort leaves ties to list order, which put
            # our Pokemon first and quietly handed us every coin flip. Sorting
            # "them" ahead of "us" at equal priority and speed means a plan is
            # only credited for what it wins without the flip. 0 sorts before 1.
            t = matrix.threat(c, target) if target is not None else None
            order.append((t.priority if t else 0, c.stats.get("spe", 0),
                          0 if side == "them" else 1, side, c, target))
    order.sort(key=lambda row: (-row[0], -row[1], row[2]))

    for _prio, _spe, _tie, _side, actor, target in order:
        if hp[id(actor)] <= 0 or target is None:
            continue                      # removed before it acted: the pin
        if hp[id(target)] <= 0:
            continue                      # already gone; the attack is wasted
        dealt = matrix.threat(actor, target).dmg_min
        dealt *= mega_bulk_factor(target)
        hp[id(target)] = max(0.0, hp[id(target)] - dealt)


def race(matrix, ours, theirs, turns=2):
    """Who is ahead after `turns` of both sides focus-firing. Pure arithmetic.

    Returns (verdict, our_losses, their_losses, margin) where margin is
    (their HP lost) - (our HP lost), in Pokemon-equivalents.

    Two turns because that is the horizon the whole approach is built on:
    *"Really, this turn + next turn (very quick and simple arithmetic) is all
    that matters."*
    """
    hp = {id(c): 1.0 for c in list(ours) + list(theirs)}
    for _ in range(max(1, turns)):
        _turn(matrix, ours, theirs, hp)
    our_lost = sum(1.0 - hp[id(c)] for c in ours)
    their_lost = sum(1.0 - hp[id(c)] for c in theirs)
    our_dead = sum(1 for c in ours if hp[id(c)] <= 0)
    their_dead = sum(1 for c in theirs if hp[id(c)] <= 0)
    margin = their_lost - our_lost
    if their_dead > our_dead or (their_dead == our_dead and margin > 0.25):
        verdict = WIN
    elif our_dead > their_dead or margin < -0.25:
        verdict = LOSS
    else:
        verdict = EVEN
    return verdict, our_dead, their_dead, margin


@dataclass
class PairResult:
    enemy_lead: tuple
    verdict: str
    our_dead: int
    their_dead: int
    margin: float
    plan: str = "both attack"       # THEIR best plan, Protect split included
    tie: bool = False               # the opening hinges on a speed tie
    # The nested answer: if we do not win outright, which back converts it.
    # (leaving, arriving, margin) or None.
    patch: tuple | None = None
    log: list = field(default_factory=list)   # the turn-by-turn, when asked for
    play: object = None                       # the Play we make about it

    @property
    def held(self) -> bool:
        """We win it outright, or a back Pokemon converts it.

        The distinction the whole approach turns on: *"see if way to patch up
        losing case, which turns into a nested 2 turn pin question."* An opening
        we lose but can switch out of is not a hole -- it is a back slot with a
        job.
        """
        return self.verdict != LOSS or self.patch is not None


@dataclass
class RaceReport:
    """Our one lead pair against EVERY lead pair they could show."""
    lead: tuple
    back: tuple
    opponent: str
    results: list = field(default_factory=list)
    # True when --max-losses stopped the scan early. The score is already zero in
    # that case, but a reader needs to know the wins count is a partial tally.
    abandoned: bool = False

    @property
    def losses(self):
        """Openings we neither win nor patch. A patched loss is not a hole."""
        return [r for r in self.results if not r.held]

    @property
    def patched(self):
        return [r for r in self.results if r.verdict == LOSS and r.patch]

    @property
    def ties(self):
        """Openings that hinge on a coin flip. Never counted as wins."""
        return [r for r in self.results if r.tie]

    @property
    def wins(self):
        return sum(1 for r in self.results if r.verdict == WIN)

    @property
    def worst(self):
        return min(self.results, key=lambda r: r.margin) if self.results else None

    @property
    def score(self) -> float:
        """Mean margin, zeroed if ANY of their leads beats us.

        The question is the one that was asked for: *"a fixed lead vs each team
        that can withstand ANY of theirs"*. A pair that loses to one of their
        fifteen openings has not answered it.
        """
        if not self.results or self.losses:
            return 0.0
        return sum(r.margin for r in self.results) / len(self.results)


# --- speed ties, the nested salvage, and the HP budget -----------------------
#
#     "Ideally your strategy does not rely on winning speed ties. Such as,
#      Garchomp+Ninetales-Alola is good because Ninetales-Alola beats enemy
#      Garchomp first to avoid speed tie, and if the enemy Mega Floette is led and
#      speed ties, you can switch Garchomp into the back pokemon *specifically
#      prepared for that special case to make that rare tied/losing lead a win*,
#      who resists floette's attacks, and beats floette even after that damage
#      ... ninetales-alola blizzards twice t1/t2, scizor is switched in t1, then
#      t2 bullet punch threatens kill on them; sequence starting t2 is even if
#      they attack they are pinned, as nothing else can come in to survive and
#      then win (not enough HP after switch damage)"
#
# Three separate mechanisms in that paragraph, and the recursion is the point:
#
#   1. A tie is not a win. `speed_ties` names the openings that hinge on a flip
#      so they can be treated as losses and PATCHED rather than counted.
#   2. `salvage` is the patch: for an opening we do not win outright, is there a
#      back Pokemon that converts it -- switch in turn 1 eating the damage,
#      attack from turn 2? A back slot spent on exactly one bad opening is a
#      deliberate, cheap insurance policy.
#   3. Which is itself the same two-turn pin question one level down, and
#      `hp_budget` is the form it takes: for each of THEIR four possible
#      switch-ins, how much HP can it afford to lose and still win? If none can
#      afford the switch-in damage, the sequence is a pin. That is the "not
#      enough HP after switch damage" clause, made into a number.

# Their side may split: one attacks, one Protects, chipping us while keeping a
# Pokemon in reserve. A claim that survives only the both-attack column is not a
# claim -- "is this robust to one enemy attacking and one protecting to whittle
# down".
ENEMY_PLANS = ("both attack", "left protects", "right protects")


def speed_ties(matrix, ours, theirs):
    """(ours, theirs) pairs at EXACTLY equal effective speed. A coin flip.

    Read off the matrix rather than the raw stat, so Tailwind, Choice Scarf,
    paralysis and a Mega's post-transformation speed are all included: two
    Pokemon "tie" when neither `outspeeds` the other and neither has priority.
    """
    out = []
    for o in ours:
        for e in theirs:
            a, b = matrix.threat(o, e), matrix.threat(e, o)
            if a.priority == b.priority and not a.outspeeds and not b.outspeeds:
                out.append((o, e))
    return out


def _turn_with(matrix, ours, theirs, hp, plan="both attack", not_acting=()):
    """One turn, with an enemy plan and a set of Pokemon that cannot attack.

    `not_acting` is how a switch is modelled: the slot spent its turn coming in,
    so it takes damage and deals none. `plan` lets one of theirs Protect, which
    means it neither takes damage nor deals it.
    """
    live_ours = [c for c in ours if hp[id(c)] > 0]
    live_theirs = [c for c in theirs if hp[id(c)] > 0]
    if not live_ours or not live_theirs:
        return
    protecting = set()
    if plan == "left protects" and live_theirs:
        protecting.add(id(live_theirs[0]))
    elif plan == "right protects" and len(live_theirs) > 1:
        protecting.add(id(live_theirs[1]))

    attackable = [c for c in live_theirs if id(c) not in protecting]
    our_target = _focus_target(matrix, live_ours, attackable or live_theirs)
    their_target = _focus_target(matrix, live_theirs, live_ours)

    order = []
    for side, mine, target in (("us", live_ours, our_target),
                               ("them", live_theirs, their_target)):
        for c in mine:
            t = matrix.threat(c, target) if target is not None else None
            order.append((t.priority if t else 0, c.stats.get("spe", 0),
                          0 if side == "them" else 1, side, c, target))
    order.sort(key=lambda row: (-row[0], -row[1], row[2]))

    for _p, _s, _tie, side, actor, target in order:
        if hp[id(actor)] <= 0 or target is None:
            continue
        if any(actor is x for x in not_acting):
            continue                       # switched in this turn
        if id(actor) in protecting:
            continue                       # spent its turn Protecting
        if side == "us" and id(target) in protecting:
            continue                       # the attack is denied
        if hp[id(target)] <= 0:
            continue
        hp[id(target)] = max(0.0, hp[id(target)]
                             - matrix.threat(actor, target).dmg_min
                             * mega_bulk_factor(target))


def _outcome(hp, ours, theirs):
    our_dead = sum(1 for c in ours if hp[id(c)] <= 0)
    their_dead = sum(1 for c in theirs if hp[id(c)] <= 0)
    margin = (sum(1.0 - hp[id(c)] for c in theirs)
              - sum(1.0 - hp[id(c)] for c in ours))
    if their_dead > our_dead or (their_dead == our_dead and margin > 0.25):
        return WIN, our_dead, their_dead, margin
    if our_dead > their_dead or margin < -0.25:
        return LOSS, our_dead, their_dead, margin
    return EVEN, our_dead, their_dead, margin


def race_robust(matrix, ours, theirs, turns=2):
    """The race, taking their BEST plan against us -- including a Protect split.

    Returns (verdict, our_dead, their_dead, margin, plan) for their best plan.
    """
    worst = None
    for plan in ENEMY_PLANS:
        hp = {id(c): 1.0 for c in list(ours) + list(theirs)}
        for _ in range(max(1, turns)):
            _turn_with(matrix, ours, theirs, hp, plan=plan)
        verdict, od, td, margin = _outcome(hp, ours, theirs)
        if worst is None or margin < worst[3]:
            worst = (verdict, od, td, margin, plan)
    return worst


def salvage(matrix, lead, back, theirs, turns=3):
    """Can a BACK Pokemon convert an opening we do not win? The nested question.

    For each (which of our leads leaves, which back replaces it): the switch-in
    eats turn 1's damage and deals none, the stayer attacks, and from turn 2 both
    attack. Three turns, because the switch spends one -- *"ninetales-alola
    blizzards twice t1/t2, scizor is switched in t1, then t2 bullet punch
    threatens kill"*.

    Returns (leaving, arriving, verdict, margin, plan) for the best patch found,
    or None. Their best plan is used throughout, Protect split included.
    """
    best = None
    for leaving in lead:
        stayer = [c for c in lead if c is not leaving]
        for arriving in back:
            ours = stayer + [arriving]
            worst = None
            for plan in ENEMY_PLANS:
                hp = {id(c): 1.0 for c in ours + list(theirs)}
                for i in range(max(1, turns)):
                    _turn_with(matrix, ours, theirs, hp, plan=plan,
                               not_acting=(arriving,) if i == 0 else ())
                verdict, _od, _td, margin = _outcome(hp, ours, theirs)
                if worst is None or margin < worst[1]:
                    worst = (verdict, margin, plan)
            verdict, margin, plan = worst
            if verdict == WIN and (best is None or margin > best[3]):
                best = (leaving, arriving, verdict, margin, plan)
    return best


def hp_budget(matrix, ours, theirs_bench, turns=2):
    """For each of their possible switch-ins: how much HP can it afford to lose?

    *"it may be helpful to think in terms of how much HP can each given potential
    back pokemon out of 4 afford to lose and still win vs yours, i.e., can they
    switch in, take damage and win, if not then it is a pin."*

    Returns [(name, afford, verdict)] where `afford` is the fraction of its max
    HP it can arrive already missing and still not lose the exchange. A negative
    number means it cannot switch in at all: even at full health our two turns
    remove it, which is the "not enough HP after switch damage" pin.
    """
    rows = []
    for b in theirs_bench:
        # What we strip while it comes in and before it can answer.
        cost = (_incoming(matrix, ours, b)
                + _incoming(matrix, ours, b, only_before_it_acts=True))
        answers = _kills_outright(matrix, b, ours[0]) if ours else False
        for o in ours[1:]:
            answers = answers or _kills_outright(matrix, b, o)
        afford = 1.0 - cost
        if afford <= 0:
            verdict = "cannot switch in"
        elif not answers:
            verdict = "survives but cannot win"
        else:
            verdict = "can switch in and win"
        rows.append((b.name, afford, verdict))
    rows.sort(key=lambda r: -r[1])
    return rows


def enemy_hp_budget(our4, enemy_roster, world, turns=2):
    """`hp_budget` for a whole bring: their bench against OUR back two.

    The back two, not the lead, because the question this answers arrives after
    the patch has been made -- Scizor is in, and now what of theirs can come in
    against it. Convenience wrapper so a CLI does not have to build a battle.
    """
    from _harness import setup_battle
    from threat import build_threat_matrix
    b, ms = setup_battle(list(our4), list(enemy_roster), world,
                         sets=_legal_default_sets(our4, world))
    matrix = build_threat_matrix(b, ms)
    ours = [c for c in b.p1.roster if not c.fainted]
    established = [ours[0], ours[2]] if len(ours) > 2 else ours
    bench = [c for c in b.p2.roster if not c.fainted][2:]
    return hp_budget(matrix, established, bench, turns=turns)


# --- committing to a MOVE ----------------------------------------------------
#
#     "the idea was ideally exerting spread damage (e.g. Garchomp Rock Slide +
#      Ninetales-Alola Blizzard) which is the ideal way to maximise our damage
#      output (damageslop) and pin; if we focus one of theirs we can be punished
#      and it's no longer a solved lead. I don't know if this is robust to the
#      simulator; I don't see how Garchomp+Dragonite can beat Whimsicott+Mega
#      Floette."
#
# That doubt was correct, and the flaw was deeper than the verdict. Everything
# above reads `matrix.threat(a, d)`, which is "a's BEST MOVE against d", computed
# independently per target. So a single attacker could be credited with Earthquake
# on one foe and Rock Slide on the other IN THE SAME TURN, and a spread move --
# the whole point -- could never hit two Pokemon, because the matrix has no
# concept of a move at all, only of a per-target damage number.
#
# Measured on exactly the position that was doubted, Garchomp against
# Whimsicott + Mega Floette:
#
#     Rock Slide   SPREAD   Whimsicott  35.8%   Floette  58.0%
#     Earthquake   SPREAD   Whimsicott  35.5%   Floette 115.3%
#
# One committed move removes Mega Floette before it acts and chips Whimsicott by
# a third. And Earthquake is `allAdjacent`, so it hits our own partner -- except
# Dragonite is Dragon/Flying and immune, which is the "flying/levitate pokemon in
# the back to ignore garchomp partner earthquake dmg" principle, priced. The old
# model reached the right verdict for the wrong reason: its focus heuristic had
# Garchomp using Rock Slide on Whimsicott, so Floette lived and Light of Ruin
# killed Garchomp.

def move_plans(actor, moveset, foes, allies, typechart, field, battle):
    """Every damaging move `actor` has, with what it does to EVERYONE it hits.

    Returns [(move, {id: damage_fraction}, priority)] where the dict covers both
    foes for a spread move and the ALLY too for an `allAdjacent` one. Worst-roll
    damage, as a fraction of each target's max HP.
    """
    from damage import damage_roll, defensive_stat, effective_stat

    def dmg(move, defender, n_targets):
        physical = move.category == "Physical"
        ak, dk = ("atk", "def") if physical else ("spa", "spd")
        atk = effective_stat(actor.stats[ak], actor.stages[ak])
        if actor.item == "Choice Band" and physical:
            atk *= 1.5
        if actor.item == "Choice Specs" and not physical:
            atk *= 1.5
        lo, _hi, _avg, _eff = damage_roll(
            50, move.power, atk, defensive_stat(defender, dk, move), actor,
            defender, move, typechart, weather=field.weather,
            num_targets_hit=n_targets)
        return (lo / (defender.max_hp() or 1)) * mega_bulk_factor(defender)

    out = []
    for move, _usage in moveset:
        if move.category == "Status" or not move.power:
            continue
        spread = is_spread_move(move.target)
        live_foes = [f for f in foes if not f.fainted]
        if spread:
            targets = list(live_foes)
            if hits_ally(move.target):
                targets += [a for a in allies if a is not actor and not a.fainted]
            n = max(1, len(targets))
            hits = {id(t): dmg(move, t, n) for t in targets}
            out.append((move, hits, move.priority, True))
        else:
            # A single-target move is a CHOICE of target, so each aim is its own
            # plan -- which is exactly the punishable thing a spread move is not.
            for t in live_foes:
                out.append((move, {id(t): dmg(move, t, 1)}, move.priority, False))
    return out


def _apply(hp, tid, dealt, sashed):
    """Damage one target, honouring Focus Sash / Sturdy. Returns HP removed.

    A sash survivor sits at a hair above zero rather than at zero, so it counts
    as ALIVE for the pin logic -- which is the point: a Pokemon on 1 HP still
    gets its turn, and a plan that assumed it was gone is wrong.
    """
    before = hp.get(tid, 0.0)
    if before <= 0:
        return 0.0
    after = before - dealt
    if after <= 0 and tid in sashed and before >= 1.0:
        after = 0.01
    hp[tid] = max(0.0, after)
    return before - hp[tid]


def race_speed(c, field, side):
    """Effective speed AS IT WILL BE: Mega stats, weather, Tailwind, Scarf.

    Ordering on `c.stats["spe"]` read the BASE form, so Mega Floette sorted at
    111 when the thing that shows up is 166 -- *"you must account for the fact
    that mega floette speed ties Garchomp when it mega evolves"*. That tie is
    invisible unless the ordering uses the evolved speed.
    """
    from copy import copy

    from engine import effective_speed
    view = c
    if getattr(c, "is_mega_pick", False) and not c.mega_evolved and c.mega_stats:
        view = copy(c)
        view.stats = dict(c.mega_stats)
        if c.mega_ability:
            view.ability = c.mega_ability
    return effective_speed(view, field, side)


def sash_ids(combatants):
    """Who survives a would-be KO at 1 HP, from full health."""
    return {id(c) for c in combatants
            if c.item == "Focus Sash" or c.ability == "Sturdy"}


def setup_moves(moveset):
    """Their turn-1 setup options that change the SPEED order, by name.

    Tailwind is the one that matters and it was completely invisible: `move_plans`
    skips Status moves, so *"whimsicott can use tailwind which lets mega floette
    outspeed (no speed tie) and easily KO Garchomp"* was a line the race could
    not represent at all. Trick Room is the mirror image and is listed for the
    same reason.
    """
    found = []
    for move, _usage in moveset or []:
        if move.name in ("Tailwind", "Trick Room"):
            found.append(move.name)
    return found


def race_bring(our4, enemy_roster, world, opponent_name="", turns=2,
               want_logs=False):
    """Our bring-4, lead first, against all 15 of their lead pairs.

    Still scored on the calibrated threat-matrix race (`race_robust`) -- this is
    the CHEAP narrowing stage the module's own docstring describes, and it stays
    arithmetic on purpose. What changed: the turn-by-turn LINES, when asked for,
    used to come from `move_race` -- a second, hand-rolled damage/turn-resolution
    model with its own move-selection heuristic (see the warning that used to sit
    on `_best_joint`: it would Earthquake its own partner for 48% to gain 23% net
    on a foe). Now the lines are read off an actual `Battle.run_turn`, played
    through `lead_sim`, so what you see explaining the score is the same engine
    the detailed report and the workbook use -- one damage model, not two.
    """
    import itertools

    import lead_sim as sim
    from _harness import setup_battle
    from threat import build_threat_matrix
    our_sets = _legal_default_sets(our4, world)
    b, ms = setup_battle(list(our4), list(enemy_roster), world, sets=our_sets)
    matrix = build_threat_matrix(b, ms)
    ours = [c for c in b.p1.roster if not c.fainted]
    lead, back = ours[:2], ours[2:]
    theirs = [c for c in b.p2.roster if not c.fainted]
    report = RaceReport(lead=tuple(c.name for c in lead),
                        back=tuple(c.name for c in back),
                        opponent=opponent_name)
    for pair in itertools.combinations(theirs, 2):
        verdict, od, td, margin, plan = race_robust(matrix, lead, pair,
                                                   turns=turns)
        log = []
        if want_logs:
            enemy4 = [c.name for c in pair] + [c.name for c in theirs
                                               if c not in pair]
            battle, movesets, _s = sim.build_position(
                [c.name for c in ours], enemy4, world, our_sets=our_sets,
                optimise=False)
            _v, _o, _t, _m, _d, log = sim.race(
                battle, movesets, turns=turns, breadth="cheap", want_log=True)
            log = log or []
        tie = bool(speed_ties(matrix, lead, pair))
        patch = None
        if verdict == LOSS or tie:
            # The nested question. A tie is patched for the same reason a loss
            # is: a plan that needs to win a coin flip is a plan with a hole.
            got = salvage(matrix, lead, back, pair)
            if got is not None:
                patch = (got[0].name, got[1].name, got[3])
        report.results.append(PairResult(
            enemy_lead=tuple(c.name for c in pair), verdict=verdict,
            our_dead=od, their_dead=td, margin=margin, plan=plan, tie=tie,
            patch=patch, log=log or []))
    report.results.sort(key=lambda r: r.margin)
    return report


def describe_race(report, limit=6):
    lines = [f"{' + '.join(report.lead)}  (back: {', '.join(report.back)})"
             f"  vs {report.opponent}",
             f"  {report.wins} wins, {len(report.patched)} patched, "
             f"{len(report.losses)} unheld of {len(report.results)} of their "
             f"lead pairs   score {report.score:+.2f}"]
    if report.ties:
        lines.append(f"  speed ties (never counted as wins): "
                     + "; ".join(" + ".join(r.enemy_lead) for r in report.ties))
    for r in report.results[:limit]:
        bits = [f"    {r.verdict.upper():5s} {' + '.join(r.enemy_lead):40s} "
                f"margin {r.margin:+.2f}"]
        if r.plan != "both attack":
            bits.append(f"[{r.plan}]")
        if r.tie:
            bits.append("TIE")
        if r.patch:
            bits.append(f"-> switch {r.patch[0]} to {r.patch[1]} "
                        f"({r.patch[2]:+.2f})")
        lines.append("  ".join(bits))
    return lines


def species_of(name: str) -> str:
    """The species-clause identity: "Mega Garchomp" and "Garchomp" are one entry.

    The sweep produced `Garchomp + Mega Garchomp` as its top pair, with
    `Dragonite + Mega Dragonite` and `Mega Garchomp / back: Garchomp` behind it.
    All illegal -- species clause counts the BASE form, and a Mega is the same
    Pokemon holding a stone. A screen that recommends an illegal team is worse
    than no screen, because its answer looks actionable.
    """
    return base_form_name(name) or name


def legal_bring(names) -> bool:
    """Four DISTINCT species. Also rejects two Megas, which no team can field."""
    species = [species_of(n) for n in names]
    if len(set(species)) != len(species):
        return False
    return sum(1 for n in names if n.startswith("Mega ")) <= 1


def mega_variants(enemy_roster, world):
    """Their roster reordered once per Mega they could choose.

    One Mega per team, chosen at PREVIEW -- after seeing your four -- so a scan
    that answers for one of their choices answers the wrong question:

        "The mega choice should not be set in stone. They could mega either, and
         it must be robust to both."

    `setup_battle` resolves the slot to the first Mega-capable name on the roster,
    so putting each candidate first in turn enumerates their options. Returns
    [(mega_name, roster)]; a single unnamed variant when they have no choice.
    """
    megas = mega_slots(enemy_roster, world)
    if len(megas) < 2:
        return [(megas[0] if megas else None, list(enemy_roster))]
    out = []
    for m in megas:
        out.append((m, [m] + [n for n in enemy_roster if n != m]))
    return out


def race_all_megas(our4, enemy_roster, world, opponent_name="", turns=2,
                   want_logs=False):
    """`race_bring` against EVERY Mega they could pick. Returns the worst.

    Measured on the reference bring, Ninetales-Alola + Garchomp vs Big 6: with
    Charizard Y holding the slot it reads as a comfortable lead, and with Mega
    Floette holding it the lead LOSES -- Floette's Mega speed is 166, exactly our
    Garchomp, the tie resolves against us, and Light of Ruin removes Garchomp
    before it acts. Blizzard takes only 27% off the Mega, so it lives easily. The
    "overwhelming lead" reading does not survive their other Mega choice.
    """
    reports = []
    for mega_name, roster in mega_variants(enemy_roster, world):
        label = (f"{opponent_name} (their Mega: {mega_name})" if mega_name
                 else opponent_name)
        reports.append(race_bring(our4, roster, world, opponent_name=label,
                                  turns=turns, want_logs=want_logs))
    worst = min(reports, key=lambda r: (r.score, -len(r.losses)))
    return worst, reports


# --- what we can actually DO about a bad opening -----------------------------
#
#     "for your play to be guaranteed and avoid rolling speed ties (my play can
#      never rely on winning a speed tie, I must have another play, assume I lose
#      speed ties), your play needs to be e.g., to switch in something on
#      Garchomp's slot that can help KO Mega Floette or just let Garchomp faint to
#      get the new pokemon in at 100% HP, assuming its still a winning line
#      afterwards (3v4 for instance)"
#
#     "I don't think this uses the full simulator logic ... It MUST be properly
#      linked. This is meant to be a useful and accurate tool, and there is no
#      room for a redundant, useless simulation - use the full battle simulator
#      logic."
#
# The four plays used to be four hand-rolled search branches, because the old
# turn resolver (`move_turn`, since removed) never itself considered Protect as
# an option and had no idea what a fainted Pokemon's replacement should be --
# so SWITCH+PROTECT and SACRIFICE had to be special-cased from outside. Neither
# needs a special case against a real `Battle`: `lead_sim.candidate_joints`
# always offers Protect as one of the stayer's options and picks it when it
# scores best, and `Battle._replace_fainted` sends in the strategically-best
# bench Pokemon (not just "next in the list") the instant a lead faints. So
# STAY already covers "stay and attack" AND "stay, and one Protects" AND "one
# faints, the best replacement arrives at 100% next turn" -- whichever of those
# the engine finds is actually best -- and SWITCH covers "switch, and the
# stayer either attacks or Protects, whichever is better". Two searches, not
# four, and both are searches the simulator itself runs rather than arithmetic
# guessing what a real turn would do.
#
# The kind label is still detected from what actually happened -- SWITCH_PROTECT
# and SACRIFICE are not distinct SEARCHES anymore, but they are still distinct
# OUTCOMES worth naming in a report, and the played-out log says which one
# occurred.

STAY = "stay"                     # both leads attack (Protect is one of their options)
SWITCH = "switch"                 # one leaves, a back arrives and eats the turn
SWITCH_PROTECT = "switch+protect"  # ...and the staying slot Protected, so it was free
SACRIFICE = "sacrifice"           # a lead fainted; the best replacement arrived at 100%

PLAYS = (STAY, SWITCH, SWITCH_PROTECT, SACRIFICE)


@dataclass
class Play:
    """One thing we can do about an opening, and how it turns out."""
    kind: str
    leaving: str = ""
    arriving: str = ""
    verdict: str = ""
    margin: float = 0.0
    their_plan: str = ""
    log: list = field(default_factory=list)
    mopped: str = ""          # what happens after, once the dust settles

    @property
    def label(self):
        if self.kind == SACRIFICE:
            return f"let {self.leaving} faint, bring {self.arriving} in at 100%"
        if self.kind == SWITCH_PROTECT:
            return (f"switch {self.leaving} -> {self.arriving}, "
                    f"Protect the other slot")
        if self.kind == SWITCH:
            return f"switch {self.leaving} -> {self.arriving}"
        return "stay in and attack"

    @property
    def guaranteed(self):
        """It wins whatever they do, INCLUDING losing every speed tie.

        `lead_sim.race` takes their best strategy over the whole enumeration and
        `Battle` resolves every exact speed tie against us on its own rule, so a
        WIN here is a guarantee in the sense that matters: it does not need a
        coin flip to land.
        """
        return self.verdict == WIN


def _refine_play_kind(kind, leaving, arriving, log):
    """Read what ACTUALLY happened off the played-out log, so a Play is labelled
    by its real outcome rather than by which search produced it.

    A SWITCH whose stayer Protected turn 1 is reported as SWITCH_PROTECT; a STAY
    (or SWITCH) during which one of our leads fainted and was replaced is
    reported as SACRIFICE, with `leaving`/`arriving` filled in from the
    replacement line Battle itself writes. Best-effort text matching against the
    engine's own log lines -- if nothing matches, the original kind stands.
    """
    text = "\n".join(log or [])
    if kind == SWITCH:
        # The stayer is whichever of our active Pokemon isn't `leaving` (which
        # just switched out and cannot itself Protect); its Protect line reads
        # "<name> protects itself!" and always sits before "--- Turn 2 ---".
        turn1_text = text.split("--- Turn 2 ---")[0]
        for line in turn1_text.splitlines():
            stripped = line.strip()
            if (stripped.endswith("protects itself!")
                    and not stripped.startswith(arriving)):
                return SWITCH_PROTECT, leaving, arriving
        return kind, leaving, arriving
    m = re.search(r"p1 sends in (.+?) \(replacing fainted (.+?)\)", text)
    if m:
        return SACRIFICE, m.group(2), m.group(1)
    return kind, leaving, arriving


def plays_for(battle, movesets, lead_names, back_names, turns=2,
              want_log=False, breadth="full"):
    """Every play available against one specific enemy opening, best first.

    `battle` must already be built with the enemy pair being tested as its
    active `p2` (see `full_report` / `item_fixes`, which build a fresh position
    per opening via `lead_sim.build_position`). STAY and each SWITCH are raced
    through `lead_sim.race`, i.e. `Battle.run_turn` end to end -- the real
    simulator decides who Protects, who gets focused, whether a Focus Sash
    holds, and which bench Pokemon is the best reply to a faint. Ranked by
    verdict then margin, so a guaranteed win sorts above a good-looking gamble.
    """
    import lead_sim as sim

    out = []
    v, _od, _td, m, plan, log = sim.race(battle, movesets, turns=turns,
                                         breadth=breadth, want_log=want_log)
    kind, leaving, arriving = _refine_play_kind(STAY, "", "", log)
    out.append(Play(kind=kind, leaving=leaving, arriving=arriving, verdict=v,
                    margin=m, their_plan=plan, log=log or []))

    for leaving in lead_names:
        for arriving in back_names:
            v2, _o, _t, m2, plan2, log2 = sim.race(
                battle, movesets, turns=turns, breadth=breadth,
                our_switch=(leaving, arriving), want_log=want_log)
            kind2, lv2, ar2 = _refine_play_kind(SWITCH, leaving, arriving, log2)
            out.append(Play(kind=kind2, leaving=lv2, arriving=ar2, verdict=v2,
                            margin=m2, their_plan=plan2, log=log2 or []))

    rank = {WIN: 0, EVEN: 1, LOSS: 2}
    out.sort(key=lambda p: (rank.get(p.verdict, 3), -p.margin))
    return out


def mop_up(our_names, their_names, world, sets=None, turns=2):
    """After the opening: do our remaining Pokemon beat theirs, on the real
    simulator?

        "It's also important that your backs can mop up and win vs remaining
         pokemon after the 2 turn sequences."

    An approximation in the same place the old version was one: it re-fights at
    fresh full HP rather than reading true HP off the specific played-out line,
    because `plays_for` does not keep battle state per candidate play. What
    changed is the fight itself -- `Battle.run_turn`, not hand-rolled arithmetic.
    Counting only Pokemon that are actually still around is still the point: an
    opening that trades evenly and leaves us with nothing that beats their last
    is not a win, and the two-turn window alone cannot see that.
    """
    import lead_sim as sim

    ours = list(our_names)[:2]
    theirs = list(their_names)[:2]
    if not ours:
        return LOSS, -1.0, "nothing left on our side"
    if not theirs:
        return WIN, 1.0, "their side is wiped"
    if len(ours) < 2 or len(theirs) < 2:
        return EVEN, 0.0, (f"{' + '.join(ours)} vs {' + '.join(theirs)}: too "
                           f"few Pokemon left on one side to simulate a 2v2")
    battle, movesets, _s = sim.build_position(
        ours, theirs, world, our_sets=sets, optimise=(sets is None))
    v, _od, _td, m, plan, _log = sim.race(battle, movesets, turns=turns,
                                          breadth="cheap", want_log=False)
    text = (f"{' + '.join(ours)} vs {' + '.join(theirs)}: {v}"
            f" (margin {m:+.2f}, their best: {plan})")
    return v, m, text


def switch_in_table(ms, typechart, field, battle, ours_out, their_bench,
                    turns=2):
    """Who of theirs can come in against what we have out, and why not.

    The reporting form of the HP budget: for each of their bench Pokemon, the
    guaranteed damage we put on it across the switch-in turn and the next before
    it acts, what it can do back, and the verdict.
    """
    rows = []
    for b in their_bench:
        cost = 0.0
        for turn in range(max(1, turns)):
            plans = {c: move_plans(c, ms.get(c.name) or [], [b], ours_out,
                                   typechart, field, battle)
                     for c in ours_out}
            plans = {c: p for c, p in plans.items() if p}
            if not plans:
                break
            best = 0.0
            for _actor, ps in plans.items():
                best_here = max((h.get(id(b), 0.0) for _m, h, _p, _s in ps),
                                default=0.0)
                best += best_here
            cost += best
        back = 0.0
        move_name = "-"
        for _m, hits, _p, _s in [x for c in ours_out
                                 for x in move_plans(b, ms.get(b.name) or [],
                                                     ours_out, [b], typechart,
                                                     field, battle)]:
            worst = max(hits.values(), default=0.0)
            if worst > back:
                back, move_name = worst, _m.name
        afford = 1.0 - cost
        if afford <= 0:
            verdict = "CANNOT come in -- dies before it acts"
        elif back >= 1.0:
            verdict = "comes in and KOs one of ours"
        else:
            verdict = "survives but does not win the exchange"
        rows.append({"name": b.name, "we_deal": cost, "afford": afford,
                     "it_deals": back, "its_move": move_name,
                     "verdict": verdict})
    rows.sort(key=lambda r: -r["afford"])
    return rows


# --- items ------------------------------------------------------------------
#
#     "It will be highly useful to include items, such as life orb on Garchomp or
#      Dragonite if it gives them sufficient damage, focus sash on Ninetales-Alola
#      if it helps an edge case where it needs to survive an OHKO. Or, you could
#      use a roseli berry (makes one incoming fairy attack damage halved) which
#      could allow Garchomp+Ninetales-Alola to win the aforementioned
#      Whimsicott+Mega Floette lead."
#
# The berry example was unusable when it was suggested, and not because of this
# module: `optimize_sets.TYPE_RESIST_BERRY` has existed for as long as the salvage
# flow, and the salvage flow has been handing berries out, but `damage.py` never
# halved anything for them. Measured before the fix: a Roseli Berry changed Light
# of Ruin by exactly zero. After it, into Garchomp (Dragon/Ground, so Fairy is
# super-effective and the berry applies): 105.5% -> 52.7%. A guaranteed OHKO
# becomes a survival, which is the whole lead.

# Offence, survival, and the type-resist berry for the type that is killing us.
# Deliberately short: the point is to find the ONE item that converts a losing
# opening, not to re-run set optimisation.
# NOT LEGAL IN REGULATION MB, and they were in here: Assault Vest, Choice Band
# and Choice Specs. An item search that proposes an illegal item is worse than no
# item search, because its answer looks actionable. Removed, and the Assault Vest
# damage reduction added to damage.py alongside the resist berries stays only
# because a berry needed the same code path -- nothing offers the Vest now.
ITEM_CANDIDATES = ("Life Orb", "Focus Sash", "Leftovers", "Sitrus Berry")


def killing_types(ms, typechart, field, battle, ours, theirs):
    """The move types actually doing the damage to us, worst first.

    Used to pick which type-resist berry is worth trying, rather than trying
    eighteen of them.
    """
    tally = {}
    for e in theirs:
        for move, hits, _p, _s in move_plans(e, ms.get(e.name) or [], ours,
                                             theirs, typechart, field, battle):
            worst = max(hits.values(), default=0.0)
            if worst > tally.get(move.move_type, 0.0):
                tally[move.move_type] = worst
    return [t for t, _v in sorted(tally.items(), key=lambda kv: -kv[1])]


def _optimised_sets(our4, enemy_roster, world):
    """Our4's best legal moveset and item against this whole enemy roster,
    computed once and reused across every lead pair.

    Optimising per lead pair would be both needlessly expensive and wrong: a
    real team is built once, before team preview, not re-drafted after seeing
    which two of six the opponent happens to lead with. `lead_sim.optimised_*`
    already filters Regulation MB's banned items (Assault Vest, Choice Band /
    Specs / Scarf), so nothing built through this helper can carry one.
    """
    import lead_sim as sim
    moves = sim.optimised_movesets(list(our4), world,
                                   enemy_names=list(enemy_roster))
    items = sim.optimised_items(list(our4), world,
                                enemy_names=list(enemy_roster))
    sets = {}
    for name in our4:
        item, item_moves = items.get(name) or (None, None)
        spec = {"moves": item_moves or moves.get(name)}
        if item:
            spec["item"] = item
        sets[name] = {k: v for k, v in spec.items() if v}
    return sets


def item_fixes(our4, enemy_roster, world, their_lead, turns=2, limit=3):
    """Items that turn a losing opening into a held one. [(mon, item, play)].

    Tries each candidate on each of our four, one at a time -- a single change,
    so the answer is actionable rather than a wholesale re-spread -- plus the
    type-resist berry for whichever type is actually killing us. Every trial is
    played out on the real simulator (`plays_for`, `Battle.run_turn`), at CHEAP
    breadth: this is a search among many candidates, not the final verdict, and
    a candidate worth taking should be re-checked at full breadth (which
    `full_report` already does for whatever bring is eventually chosen).
    """
    from optimize_sets import TYPE_RESIST_BERRY

    import lead_sim as sim
    if not all(n in enemy_roster for n in their_lead):
        return []
    base_sets = _optimised_sets(our4, enemy_roster, world)
    enemy4 = list(their_lead) + [n for n in enemy_roster if n not in their_lead]
    lead, back = list(our4[:2]), list(our4[2:])

    def run(overrides):
        sets = {k: dict(v) for k, v in base_sets.items()}
        for name, ov in overrides.items():
            sets.setdefault(name, {}).update(ov)
        battle, movesets, _s = sim.build_position(our4, enemy4, world,
                                                   our_sets=sets, optimise=False)
        return plays_for(battle, movesets, lead, back, turns=turns,
                         breadth="cheap")[0]

    base = run({})
    if base.verdict == WIN:
        return []

    battle0, movesets0, _s0 = sim.build_position(
        our4, enemy4, world, our_sets=base_sets, optimise=False)
    ours0 = battle0.p1.roster
    theirs0 = battle0.p2.roster[:2]
    berries = [TYPE_RESIST_BERRY[t] for t in
               killing_types(movesets0, battle0.typechart, battle0.field,
                             battle0, ours0[:2], theirs0)
               if t in TYPE_RESIST_BERRY][:2]

    found = []
    for mon in our4:
        for item in list(ITEM_CANDIDATES) + berries:
            play = run({mon: {"item": item}})
            if play.verdict == WIN or (base.verdict == LOSS
                                       and play.verdict == EVEN):
                found.append((mon, item, play))
    found.sort(key=lambda t: (0 if t[2].verdict == WIN else 1, -t[2].margin))
    return found[:limit]


def full_report(our4, enemy_roster, world, opponent_name="", turns=2,
                want_logs=True, mega_name=None, plays=True, max_losses=0,
                sets=None):
    """One bring against one opponent (one Mega choice), with everything.

    For each of their lead pairs (every combination of two from `enemy_roster`):
    the best PLAY we have (stay, switch, switch+Protect, sacrifice), whether it
    is guaranteed, and the MOP-UP -- do our remaining Pokemon beat their
    remaining Pokemon once the opening resolves. Also builds the reference
    position (their roster in its own order) for the switch-in table.

    ONE ENGINE. Every opening is played on a real `Battle`, built fresh per pair
    via `lead_sim.build_position` (so weather, Intimidate on send-out, item and
    the Mega slot are all resolved by the engine exactly as they are in a game)
    and raced with `lead_sim.race`, i.e. `Battle.run_turn` end to end -- not a
    second, hand-rolled damage model. `sets` fixes our moves/items instead of
    optimising them fresh; `item_fixes`/`recommended_items` use this to ask
    "what if this one held a different item" without re-deriving everything.
    """
    import itertools

    import lead_sim as sim
    theirs_all = list(enemy_roster)
    if sets is None:
        sets = _optimised_sets(our4, theirs_all, world)

    lead, back = list(our4[:2]), list(our4[2:])
    label = (f"{opponent_name} (their Mega: {mega_name})" if mega_name
             else opponent_name)
    report = RaceReport(lead=tuple(lead), back=tuple(back), opponent=label)

    # The reference position, in the roster's OWN order -- their stated leads
    # first, the rest as bench -- for the switch-in table and for callers that
    # want a single position representing "the bring as given".
    ref_battle, ref_movesets, _s = sim.build_position(
        our4, theirs_all, world, our_sets=sets, optimise=False,
        enemy_mega=mega_name)

    for pair in itertools.combinations(theirs_all, 2):
        # EARLY EXIT. The score is zeroed by a single unheld opening, so a bring
        # already five down is not a candidate and finishing the rest tells us
        # nothing we will act on. Much the cheapest speed-up available.
        if max_losses and len(report.losses) >= max_losses:
            report.abandoned = True
            break
        enemy4 = list(pair) + [n for n in theirs_all if n not in pair]
        battle, movesets, _s2 = sim.build_position(
            our4, enemy4, world, our_sets=sets, optimise=False,
            enemy_mega=mega_name)
        tie = any(race_speed(o, battle.field, "p1")
                 == race_speed(e, battle.field, "p2")
                 for o in battle.p1.active if o is not None
                 for e in battle.p2.active if e is not None)
        if plays:
            options = plays_for(battle, movesets, lead, back, turns=turns,
                                want_log=want_logs)
            best = options[0]
        else:
            # SWEEP MODE: the stay-in baseline only, at CHEAP breadth. The
            # relationship to the detailed pass is MONOTONE -- adding plays and
            # more of their strategy space can only improve an opening's
            # verdict, never worsen it -- so the sweep's score is a lower bound
            # and the detailed report's is the real one. Same engine either way
            # ("ONE ENGINE": the sweep and the workbook used to disagree because
            # the sweep scored on committed-move arithmetic while the workbook
            # played the threat-matrix race; now both are `Battle.run_turn`, at
            # different breadths, so they can never contradict each other).
            v, _o, _t, m, plan, log = sim.race(
                battle, movesets, turns=turns, breadth="cheap",
                want_log=want_logs)
            best = Play(kind=STAY, verdict=v, margin=m, their_plan=plan,
                        log=log or [])
        # The mop-up: whatever of ours is left (everyone except whoever left, if
        # anyone did) against whatever of theirs is left (everyone not in this
        # opening's pair).
        if plays:
            their_rest = [n for n in theirs_all if n not in pair]
            our_rest = [n for n in our4 if n != best.leaving]
            v_mop, _m_mop, text = mop_up(our_rest, their_rest, world,
                                         sets=sets, turns=turns)
            best.mopped = f"{v_mop}: {text}"
        report.results.append(PairResult(
            enemy_lead=tuple(pair), verdict=best.verdict,
            our_dead=0, their_dead=0, margin=best.margin,
            plan=best.their_plan, tie=tie,
            # A patch only counts if it actually SALVAGES the opening. Setting
            # it whenever the best play was a switch made `held` true for
            # openings the switch still loses, so the report read "5 patched, 0
            # unheld" beside five LOSS rows.
            patch=((best.leaving, best.arriving, best.margin)
                   if best.kind != STAY and best.verdict != LOSS else None),
            log=best.log, play=best))
    report.results.sort(key=lambda r: r.margin)
    return report, ref_movesets, ref_battle


def default_item(name, world):
    """The most-used LEGAL item for this Pokemon.

    Stated as a function so the workbook can SHOW the baseline it is deviating
    from. NOT quite `make_team`'s raw default: `make_team` (and everything
    downstream of it, like `_harness.setup_battle`) trusts `items_usage`'s #1
    entry outright, which is not filtered against Regulation MB's banned items
    -- that table is logged across every format, banned or not, so a Pokemon
    whose most-used item happens to be a Choice item or Assault Vest would
    otherwise field one with nobody having asked for it (Hydreigon's #1 item is
    Choice Scarf). `_legal_default_sets` turns this into an explicit override
    for every `setup_battle` call in this module's cheap stage; `lead_sim`'s own
    optimiser filters separately for the detailed stage. Measured, otherwise:
    Ninetales-Alola gets Never-Melt Ice (29%), Garchomp Life Orb (61%),
    Rotom-Wash Sitrus Berry (31%).
    """
    from lead_sim import BANNED_ITEMS
    usage = (world["merged"].get(name) or {}).get("items_usage") or []
    for item, _pct in usage:
        if item and item != "Other" and item not in BANNED_ITEMS:
            return item
    return None


def _held_count(our4, enemy_roster, world, sets, mega_name=None, turns=2):
    """How many of their lead pairs we hold (win outright, or salvage by
    switching), at CHEAP simulator breadth. A ranking tool for item candidates,
    not a verdict -- `full_report` (full breadth) is the real count.
    """
    import itertools

    import lead_sim as sim
    theirs_all = list(enemy_roster)
    lead, back = list(our4[:2]), list(our4[2:])
    n = 0
    for pair in itertools.combinations(theirs_all, 2):
        enemy4 = list(pair) + [x for x in theirs_all if x not in pair]
        battle, movesets, _s = sim.build_position(
            our4, enemy4, world, our_sets=sets, optimise=False,
            enemy_mega=mega_name)
        best = plays_for(battle, movesets, lead, back, turns=turns,
                         breadth="cheap")[0]
        if best.verdict != LOSS:
            n += 1
    return n


def recommended_items(our4, enemy_roster, world, turns=2, mega_name=None):
    """Per Pokemon: keep the optimised item, or the one that holds more openings.

        "By default give every pokemon their most popular item, assuming no
         better item for a specific case."

    The baseline is `_optimised_sets`'s choice (which already starts from the
    usage-popular item and only departs from it when a matchup earns the
    departure), and a swap has to EARN its place too: it is only reported when
    it strictly increases the number of their lead pairs we hold. One change at
    a time, so the answer stays actionable.

    Returns [(mon, from_item, to_item, held_before, held_after)] for the swaps
    worth making, best first. An empty list means the chosen items are right.
    """
    from optimize_sets import TYPE_RESIST_BERRY

    import lead_sim as sim
    theirs_all = list(enemy_roster)
    base_sets = _optimised_sets(our4, theirs_all, world)

    base = _held_count(our4, theirs_all, world, base_sets,
                       mega_name=mega_name, turns=turns)
    battle0, movesets0, _s0 = sim.build_position(
        our4, theirs_all, world, our_sets=base_sets, optimise=False,
        enemy_mega=mega_name)
    ours0 = battle0.p1.roster
    theirs0 = battle0.p2.roster[:2]
    berries = [TYPE_RESIST_BERRY[t] for t in
               killing_types(movesets0, battle0.typechart, battle0.field,
                             battle0, ours0[:2], theirs0)
               if t in TYPE_RESIST_BERRY][:2]

    out = []
    for mon in our4:
        current = (base_sets.get(mon) or {}).get("item") or default_item(mon, world)
        # A Mega is locked to its stone; swapping it is not a legal suggestion.
        if current and current.endswith("ite"):
            continue
        for item in list(ITEM_CANDIDATES) + berries:
            if item == current:
                continue
            sets = {k: dict(v) for k, v in base_sets.items()}
            sets.setdefault(mon, {}).update({"item": item})
            got = _held_count(our4, theirs_all, world, sets,
                              mega_name=mega_name, turns=turns)
            if got > base:
                out.append((mon, current, item, base, got))
    out.sort(key=lambda t: -t[4])
    return out
