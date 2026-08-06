"""
Excel export -- gives full visibility into how each recommendation was
reached (not just the winning pick) plus the actual turn-by-turn gameplan.

Workbook layout:
  - "Recommendations": one row per enemy team, the winning lead/back + result
  - "<Team> - All Combos": EVERY bring4/lead2 combination that was searched
    for that team, sorted best-first -- this is the transparency piece: you
    can see exactly how many alternatives were tried and how the winner
    compares to everything else, not just take the #1 pick on faith.
  - "<Team> - Gameplan": the actual turn-by-turn play-by-play of the
    recommended composition's battle against that team's toughest lead --
    every move, target, damage number, flinch, switch, and field effect.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WIN_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
BEST_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths=None):
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)


def _safe_sheet_name(name: str) -> str:
    for ch in '[]:*?/\\':
        name = name.replace(ch, "")
    return name[:31]


def build_workbook(team_reports: dict, out_path: str):
    """team_reports: {team_name: {
        'toughest_lead': (name1, name2),
        'enemy_back': [name1, name2],
        'all_combos': [(bring4, winner, turns), ...],   # sorted best-first
        'gameplan_events': [...],                          # battle.events for the winning combo
        'gameplan_log': str,                                # battle.log.dump() text version
    }}"""
    wb = Workbook()

    # --- Recommendations sheet ---
    ws = wb.active
    ws.title = "Recommendations"
    ws.append(["Enemy Team", "Enemy Toughest Lead", "Enemy Back", "Our Lead", "Our Back",
               "Result", "Turns", "Combos That Win", "Combos Tested"])
    _style_header(ws)
    for team_name, rep in team_reports.items():
        combos = rep["all_combos"]
        best = combos[0]
        wins = sum(1 for _, w, _ in combos if w == "p1")
        row = [team_name, "/".join(rep["toughest_lead"]), "/".join(rep["enemy_back"]),
               "/".join(best[0][:2]), "/".join(best[0][2:]), best[1], best[2], wins, len(combos)]
        ws.append(row)
        r = ws.max_row
        fill = WIN_FILL if best[1] == "p1" else LOSS_FILL
        for cell in ws[r]:
            cell.fill = fill
    ws.freeze_panes = "A2"
    _autosize(ws)

    # --- Per-team sheets ---
    for team_name, rep in team_reports.items():
        # All Combos
        ws = wb.create_sheet(_safe_sheet_name(f"{team_name} Combos"))
        ws.append(["Lead 1", "Lead 2", "Back 1", "Back 2", "Result", "Turns"])
        _style_header(ws)
        for i, (bring4, winner, turns) in enumerate(rep["all_combos"]):
            ws.append([bring4[0], bring4[1], bring4[2], bring4[3], winner, turns])
            r = ws.max_row
            fill = BEST_FILL if i == 0 else (WIN_FILL if winner == "p1" else
                                              (LOSS_FILL if winner == "p2" else None))
            if fill:
                for cell in ws[r]:
                    cell.fill = fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)

        # Per-enemy-lead breakdown (comprehensive mode only): one row per possible
        # enemy lead, with our best answer and result against each.
        if rep.get("per_lead"):
            ws = wb.create_sheet(_safe_sheet_name(f"{team_name} All Leads"))
            ws.append(["Enemy Lead", "Enemy Back", "Our Lead", "Our Back", "Result", "Turns",
                       "Our Combos Winning", "Combos Tested"])
            _style_header(ws)
            for enemy_lead, d in rep["per_lead"].items():
                b = d["best"]
                ws.append(["/".join(enemy_lead), "/".join(d["enemy_bring4"][2:]),
                           "/".join(b[0][:2]), "/".join(b[0][2:]), b[1], b[2],
                           d["wins"], len(d["all_combos"])])
                r = ws.max_row
                fill = WIN_FILL if b[1] == "p1" else (LOSS_FILL if b[1] == "p2" else None)
                if fill:
                    for cell in ws[r]:
                        cell.fill = fill
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            _autosize(ws)

            # One gameplan sheet per enemy lead.
            for enemy_lead, d in rep["per_lead"].items():
                title = _safe_sheet_name(f"{team_name} vs {enemy_lead[0][:8]}-{enemy_lead[1][:8]}")
                ws = wb.create_sheet(title)
                _write_gameplan(ws, d["gameplan_events"])

        # Gameplan (turn-by-turn)
        ws = wb.create_sheet(_safe_sheet_name(f"{team_name} Gameplan"))
        _write_gameplan(ws, rep["gameplan_events"])

    wb.save(out_path)
    return out_path


def _write_gameplan(ws, events):
        ws.append(["Turn", "Side", "Event", "Actor", "Move/Detail", "Target", "Damage",
                   "Type Eff", "Target HP After", "Notes"])
        _style_header(ws)
        for ev in events:
            notes = []
            if ev.get("fainted"):
                notes.append("FAINTED")
            if ev.get("mega_evolved"):
                notes.append("Mega Evolved")
            if ev.get("weather_after"):
                notes.append(f"weather={ev['weather_after']}")
            if ev.get("trick_room_now") is not None:
                notes.append(f"trick_room={'ON' if ev['trick_room_now'] else 'OFF'}")

            target_hp = f"{ev['target_hp_after']}/{ev['target_max_hp']}" if "target_hp_after" in ev else ""
            ws.append([
                ev.get("turn", ""), ev.get("side", ""), ev.get("event", ""),
                ev.get("actor", ev.get("out", "")),
                ev.get("move", ev.get("incoming", "")),
                ev.get("target", ""),
                ev.get("damage", ""),
                ev.get("type_eff", ""),
                target_hp,
                "; ".join(notes),
            ])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)
