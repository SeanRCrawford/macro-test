"""Turn a team-search cache into a workbook you can actually read.

The search prints a ranked table and stores its results as JSON. Neither is a
good way to answer the question a player has after an overnight run, which is
never "what is the number" but "WHY is it that number, and what do I change".
Answering that needs the losing turns themselves, so this exports four sheets
at descending zoom:

  Teams      one row per team: mean exploitability across its matchups
  Matchups   one row per (team vs opponent): the bring that rated best
  Candidates every bring that was audited, not just the winner -- the second
             place bring is the one you switch to, so it has to be visible
  Turns      every audited turn: our play, their punish, and the gap

The Turns sheet is the transparency piece and the reason the search records
per-turn detail at all: a rating you cannot trace back to a specific turn where
a specific play gets answered by a specific move is a number to take on faith.

Reads the cache dict written by tools/search_teams.py. Rows from a `quick`-tier
run carry no audit, and are exported with blank rating columns rather than
being dropped -- "this pairing was searched but not rated" is information.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOOD_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
BAD_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")

# Matches robustness.SEVERE: a best response gaining more than this is a play
# worth reconsidering. Imported rather than restated where possible.
try:
    from robustness import SEVERE
except ImportError:                                   # pragma: no cover
    SEVERE = 60.0

MILD = SEVERE / 2.0

# Excel's hard limit is 1,048,576 rows; openpyxl raises above it and the run
# dies AFTER the search, which is the worst possible time. `--audit-all
# --brings 90` across eight opponents is 8100 audited lines a pairing, so the
# Turns sheet projects to ~1.2M rows -- over the limit -- and the Lines sheet to
# 65k, which fits but takes minutes to autosize. Cap well below the limit: a
# sheet nobody can scroll is not more useful than a sheet that says where it
# stopped.
MAX_SHEET_ROWS = 100_000

# Sheets that survive `--workbook light`. The three omitted ones (Lines,
# Candidates, Turns) are the diagnosis sheets, and they are exactly the three
# whose row count scales with `brings x their configs`.
LIGHT_SHEETS = ("Teams", "Plan", "Matchups", "Best lines", "Team sheets",
                "How to read this")


class _Discarded:
    """Absorbs `.fill = ...` / `.alignment = ...` on a row that was not written."""
    __slots__ = ("fill", "alignment", "value", "font")


class _Capped:
    """A worksheet that stops appending at MAX_SHEET_ROWS and says so.

    Wrapping rather than checking at each call site: the four exploding sheets
    append from three levels of nested loop, and a `break` deep inside one of
    them is easy to get subtly wrong (and easy to forget in the next sheet).
    """

    def __init__(self, ws, limit=None):
        self._ws = ws
        # Read at construction, not bound as a default argument: a test that
        # lowers the cap to check the truncation note would otherwise be
        # silently ignored, and pass by building the whole sheet.
        self._limit = limit or MAX_SHEET_ROWS
        self.truncated = 0

    def __getattr__(self, name):
        return getattr(self._ws, name)

    def __getitem__(self, key):
        return self._ws[key]

    def append(self, row):
        if self._ws.max_row >= self._limit:
            self.truncated += 1
            return
        self._ws.append(row)

    def cell(self, *a, **kw):
        # Once truncated, callers are still styling "the row I just appended",
        # which no longer exists. Hand back a scratch object so the fill lands
        # nowhere instead of recolouring the last row that did get written.
        if self.truncated:
            return _Discarded()
        return self._ws.cell(*a, **kw)

    def finish(self):
        if self.truncated:
            self._ws.append([f"... and {self.truncated} more rows, not written: "
                             f"this sheet is capped at {self._limit:,} rows. "
                             f"Re-run the export with a narrower cache, or use "
                             f"--workbook light and read the Best lines sheet."])
        return self._ws


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws):
    for i, col in enumerate(ws.columns, start=1):
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 55)


def _shade(cell, value):
    """Green / amber / red by how much a good player gains. Lower is better."""
    if value is None:
        return
    cell.fill = GOOD_FILL if value <= MILD else (WARN_FILL if value <= SEVERE else BAD_FILL)


def rows_of(cache_data):
    """The pairing records in a cache dict, oldest key order preserved."""
    return [v for v in cache_data.values()
            if isinstance(v, dict) and v.get("ours")]


def _line_totals(row):
    """(lines, wins, adjusted_total) over every audited line of a pairing.

    Counts, not rates. "0.63" does not tell you whether that is 57 wins out of
    90 or a rate over two sampled leads, and under --audit-all the denominator
    is the thing that makes the number trustworthy.
    """
    lines = wins = 0
    adjusted = 0.0
    for cand in row.get("candidates") or []:
        for lead in cand.get("audit") or []:
            lines += 1
            if (lead.get("outcome") or "").lower() == "win":
                wins += 1
            adjusted += lead.get("adjusted_value") or 0.0
    return lines, wins, adjusted


def committed_plan(row):
    """The ONE bring-4/lead-2 to commit to against this opponent, and its record.

    At team preview you choose your four and your lead pair BEFORE seeing which
    four they bring, let alone which two they lead. So a plan is a single
    committed choice measured across everything they might do -- not the best
    line found per enemy configuration, which is what the sheets used to show
    and which no player can actually execute.

    Picks the candidate with the most wins across the enemy configurations it
    was audited against, breaking ties on the lower mean punish. Returns
    (candidate, wins, total, adjusted_total) or None.
    """
    best = None
    for cand in row.get("candidates") or []:
        lines = cand.get("audit") or []
        if not lines:
            continue
        wins = sum(1 for ln in lines
                   if (ln.get("outcome") or "").lower() == "win")
        adjusted = sum(ln.get("adjusted_value") or 0.0 for ln in lines)
        punish = _mean([ln.get("mean_exploitability") for ln in lines]) or 0.0
        rank = (-wins, punish)
        if best is None or rank < best[0]:
            best = (rank, cand, wins, len(lines), adjusted)
    if best is None:
        return None
    _rank, cand, wins, total, adjusted = best
    return cand, wins, total, adjusted


def _mean(values):
    values = [v for v in values if v is not None]
    return sum(values) / len(values) if values else None


def _teams_sheet(wb, rows):
    ws = wb.active
    ws.title = "Teams"
    ws.append(["Team", "Adjusted wins", "Wins vs punisher", "Clean wins",
               "Lines won", "Lines audited", "Adjusted wins (count)",
               "Mean exploitability", "Games won", "Games played",
               "Worst matchup", "Worst value", "Severe turns",
               "Matchups rated", "Matchups searched", "Read with care"])
    _style_header(ws)

    by_team = {}
    for r in rows:
        by_team.setdefault(r["ours"], []).append(r)

    def _rank_key(kv):
        # Wins that hold up first, punishability as the tie-break. Ranking on
        # exploitability alone puts a team that loses everything on top.
        adjusted = _mean([x.get("adjusted_win_rate") for x in kv[1]])
        exploit = _mean([x.get("exploitability") for x in kv[1]])
        return (-(adjusted if adjusted is not None else -1.0),
                exploit if exploit is not None else float("inf"))

    ranked = sorted(by_team.items(), key=_rank_key)
    for name, rs in ranked:
        rated = [x for x in rs if x.get("exploitability") is not None]
        mean = _mean([x["exploitability"] for x in rated])
        worst = max(rated, key=lambda x: x["exploitability"]) if rated else None
        wins = sum(x.get("solver_wins") or 0 for x in rs)
        played = sum(x.get("solver_total") or 0 for x in rs)
        # Exploitability is measured against each turn's equilibrium, so a
        # position that is simply LOST has nothing left to punish and rates
        # near zero. Without the win count beside it, a team that loses
        # everything can top a sheet headed "least punishable".
        caution = ("loses most games -- low rating here means the line was "
                   "played well, not that the team is good"
                   if played and wins / played < 0.5 else None)
        adjusted = _mean([x.get("adjusted_win_rate") for x in rs])
        robust = _mean([x.get("robust_win_rate") for x in rs])
        clean = _mean([x.get("reliable_wins") for x in rs])
        n_lines = n_wins = 0
        adj_total = 0.0
        for x in rs:
            li, wi, ad = _line_totals(x)
            n_lines += li
            n_wins += wi
            adj_total += ad
        ws.append([name,
                   round(adjusted, 3) if adjusted is not None else None,
                   round(robust, 3) if robust is not None else None,
                   round(clean, 3) if clean is not None else None,
                   f"{n_wins} / {n_lines}" if n_lines else None,
                   n_lines or None,
                   f"{adj_total:.1f} / {n_lines}" if n_lines else None,
                   round(mean, 1) if mean is not None else None,
                   wins or None, played or None,
                   worst["theirs"] if worst else None,
                   round(worst["exploitability"], 1) if worst else None,
                   sum(x.get("severe_turns") or 0 for x in rs),
                   len(rated), len(rs), caution])
        # Adjusted wins is a rate where HIGH is good, the opposite of every
        # other shaded column, so it gets its own scale rather than _shade.
        if adjusted is not None:
            ws.cell(ws.max_row, 2).fill = (
                GOOD_FILL if adjusted >= 0.6
                else (WARN_FILL if adjusted >= 0.3 else BAD_FILL))
        _shade(ws.cell(ws.max_row, 8), mean)
        _shade(ws.cell(ws.max_row, 12), worst["exploitability"] if worst else None)
        if caution:
            ws.cell(ws.max_row, 16).fill = BAD_FILL
    _autosize(ws)


def _matchups_sheet(wb, rows):
    ws = wb.create_sheet("Matchups")
    ws.append(["Team", "Opponent", "Best bring (lead first)", "Adjusted wins",
               "Wins vs punisher", "Lines won", "Lines audited",
               "Adjusted wins (count)", "Exploitability",
               "Severe turns", "Their hardest lead", "Worst turn",
               "Our play", "Punished by", "Games won", "Games played"])
    _style_header(ws)
    for r in sorted(rows, key=lambda x: (x["ours"],
                                         -(x.get("adjusted_win_rate") or -1))):
        wt = r.get("worst_turn") or {}
        ws.append([
            r["ours"], r["theirs"],
            " / ".join(r["bring"]) if r.get("bring") else None,
            round(r["adjusted_win_rate"], 3)
            if r.get("adjusted_win_rate") is not None else None,
            round(r["robust_win_rate"], 3)
            if r.get("robust_win_rate") is not None else None,
            (f"{_line_totals(r)[1]} / {_line_totals(r)[0]}"
             if _line_totals(r)[0] else None),
            _line_totals(r)[0] or None,
            (f"{_line_totals(r)[2]:.1f} / {_line_totals(r)[0]}"
             if _line_totals(r)[0] else None),
            round(r["exploitability"], 1) if r.get("exploitability") is not None else None,
            r.get("severe_turns"),
            " / ".join(r.get("hardest_lead") or []) or None,
            f"T{wt['turn']}" if wt.get("turn") else None,
            wt.get("our_play"), wt.get("punished_by"),
            r.get("solver_wins"), r.get("solver_total"),
        ])
        _shade(ws.cell(ws.max_row, 9), r.get("exploitability"))
    _autosize(ws)


def _candidates_sheet(wb, rows):
    """Every audited bring, so the runner-up is visible next to the winner."""
    ws = wb.create_sheet("Candidates")
    ws.append(["Team", "Opponent", "Rank", "Bring (lead first)", "Exploitability",
               "Severe turns", "Turns rated", "Screen worst margin",
               "Games won", "Games played"])
    _style_header(ws)
    any_rows = False
    for r in sorted(rows, key=lambda x: (x["ours"], x["theirs"])):
        for i, cand in enumerate(r.get("candidates") or [], start=1):
            any_rows = True
            ws.append([
                r["ours"], r["theirs"], i,
                " / ".join(cand.get("bring") or []) or None,
                round(cand["exploitability"], 1)
                if cand.get("exploitability") is not None else None,
                cand.get("severe_turns"), cand.get("rated_turns"),
                round(cand["worst_margin"], 1)
                if cand.get("worst_margin") is not None else None,
                cand.get("solver_wins"), cand.get("solver_total"),
            ])
            _shade(ws.cell(ws.max_row, 5), cand.get("exploitability"))
    if not any_rows:
        ws.append(["No per-candidate detail in this cache "
                   "(quick tier, or a run from before candidates were recorded)."])
    _autosize(ws)


def _plan_sheet(wb, rows):
    """THE deliverable: one committed bring/lead per opponent, and its record.

    You choose your four and your lead pair at preview, before seeing theirs.
    So the only number that describes a playable plan is: commit to this, and
    win X of their N configurations. A perfect answer to an opponent is X == N
    (90/90 when the run used --audit-all or --effort exhaustive).

    Sorted worst record first, because the matchup you cannot answer is the one
    that decides whether the team is worth taking.
    """
    ws = wb.create_sheet("Plan")
    ws.append(["Team", "Opponent", "COMMIT: bring (lead first)",
               "Wins", "Of", "Win rate", "Adjusted wins", "Mean punish",
               "Worst punish", "Severe turns", "Losing to"])
    _style_header(ws)
    any_rows = False
    entries = []
    for r in rows:
        plan = committed_plan(r)
        if plan is None:
            continue
        any_rows = True
        cand, wins, total, adjusted = plan
        lines = cand.get("audit") or []
        punish = _mean([ln.get("mean_exploitability") for ln in lines]) or 0.0
        worst = max((ln.get("worst_exploitability") or 0.0) for ln in lines)
        severe = sum(ln.get("severe_turns") or 0 for ln in lines)
        losses = [" / ".join(ln.get("enemy_bring") or ln.get("lead") or [])
                  for ln in lines
                  if (ln.get("outcome") or "").lower() != "win"]
        entries.append((wins / total if total else 0.0, r, cand, wins, total,
                        adjusted, punish, worst, severe, losses))

    for rate, r, cand, wins, total, adjusted, punish, worst, severe, losses \
            in sorted(entries, key=lambda e: (e[0], -e[6])):
        ws.append([
            r["ours"], r["theirs"], " / ".join(cand.get("bring") or []),
            wins, total, round(rate, 3), round(adjusted, 1),
            round(punish, 1), round(worst, 1), severe,
            # Naming the configurations that beat this plan is the actionable
            # part: they are what a different bring, or a different team, has
            # to answer.
            "; ".join(losses[:6]) + ("  ..." if len(losses) > 6 else "") or None,
        ])
        ws.cell(ws.max_row, 6).fill = (
            GOOD_FILL if rate >= 0.9 else (WARN_FILL if rate >= 0.7 else BAD_FILL))
        _shade(ws.cell(ws.max_row, 8), punish)
    if not any_rows:
        ws.append(["No audited plans in this cache "
                   "(the quick tier does not audit lines)."])
    _autosize(ws)


def _best_lines_sheet(wb, rows):
    """The plan: ONE committed bring per opponent, and how it fares against all
    of their configurations.

    Everything else in this workbook is diagnosis. This is the answer, and it
    is deliberately a single lead per opponent -- you commit at preview, so a
    sheet that switched lead depending on what they brought would be showing a
    plan nobody can play.
    """
    ws = _Capped(wb.create_sheet("Best lines"))
    ws.append(["Opponent", "Team", "COMMIT: bring (lead first)", "Record",
               "Adjusted", "Their bring", "Result", "Worst punish", "Turn",
               "Play this", "If they answer with", "What happens", "KOs",
               "HP after"])
    _style_header(ws)
    any_rows = False
    for r in sorted(rows, key=lambda x: (x["ours"], x["theirs"])):
        plan = committed_plan(r)
        if plan is None:
            ws.append([r["theirs"], r["ours"],
                       " / ".join(r.get("bring") or []) or None,
                       "not audited", None, None, None, None, None, None,
                       None, None, None, None])
            continue
        cand, wins, total, adjusted = plan
        any_rows = True
        # Winning lines first, then the worst offenders -- the losses are the
        # reason the record is not X/X and are what a plan has to answer.
        lines = sorted(cand.get("audit") or [],
                       key=lambda ln: ((ln.get("outcome") or "").lower() != "win",
                                       -(ln.get("mean_exploitability") or 0)))
        header_done = False
        for ln in lines:
            outcome = (ln.get("outcome") or "").lower()
            worst = ln.get("worst_exploitability")
            # One row per turn, and ONE row when there are no turns to show.
            # `enumerate(x) or [...]` never fell through -- a generator is
            # always truthy -- so a line with no stored turns silently vanished
            # from this sheet. That was invisible while every line carried its
            # turns and became the whole sheet the moment --detail started
            # trimming them: under `light` the winning lines are exactly the
            # ones without turns, so the plan's wins disappeared and only its
            # losses were listed.
            turns = list(enumerate(ln.get("turns") or [])) or [(0, {})]
            for i, t in turns:
                ws.append([
                    r["theirs"] if not header_done else None,
                    r["ours"] if not header_done else None,
                    " / ".join(cand.get("bring") or [])
                    if not header_done else None,
                    f"{wins} / {total}" if not header_done else None,
                    f"{adjusted:.1f} / {total}" if not header_done else None,
                    " / ".join(ln.get("enemy_bring") or ln.get("lead") or [])
                    if i == 0 else None,
                    outcome.upper() if i == 0 else None,
                    round(worst, 1) if (i == 0 and worst is not None) else None,
                    t.get("turn"), t.get("our_play"),
                    t.get("punished_by") or ("no punish available"
                                             if t.get("no_punish") else None),
                    "\n".join(t.get("events") or []) or None,
                    ", ".join(t.get("kos") or []) or None,
                    ", ".join(t.get("hp_after") or []) or None,
                ])
                if i == 0:
                    ws.cell(ws.max_row, 7).fill = (
                        GOOD_FILL if outcome == "win" else BAD_FILL)
                ws.cell(ws.max_row, 12).alignment = Alignment(wrap_text=True,
                                                              vertical="top")
                header_done = True
    if not any_rows:
        ws.append(["No audited lines in this cache "
                   "(the quick tier does not audit lines)."])
    _autosize(ws.finish())


def _lines_sheet(wb, rows):
    """EVERY audited line: one row per (our bring x their bring).

    The Best lines sheet answers "what do I play"; this answers "and what
    happens against everything else they could have brought". Under
    --audit-all / --effort exhaustive that is all 90 of their bring-4s, so the
    lines that lose sit next to the ones that win rather than being summarised
    away into a single team average.

    Adjusted value is this line's own contribution to the team's adjusted-wins
    number, on the same 0-1 scale, so a team total can be traced back to the
    lines that produced it.
    """
    ws = _Capped(wb.create_sheet("Lines"))
    ws.append(["Team", "Opponent", "Our bring (lead first)",
               "Their bring (lead first)", "Lead likelihood", "Result",
               "Adjusted value", "Mean punish", "Worst punish",
               "Expected loss", "Severe turns", "Turns"])
    _style_header(ws)
    any_rows = False
    for r in sorted(rows, key=lambda x: (x["ours"], x["theirs"])):
        for cand in r.get("candidates") or []:
            for lead in cand.get("audit") or []:
                any_rows = True
                outcome = (lead.get("outcome") or "").lower()
                ws.append([
                    r["ours"], r["theirs"],
                    " / ".join(cand.get("bring") or []) or None,
                    " / ".join(lead.get("enemy_bring")
                               or lead.get("lead") or []) or None,
                    round(lead["probability"], 3)
                    if lead.get("probability") is not None else None,
                    outcome.upper() or None,
                    round(lead["adjusted_value"], 3)
                    if lead.get("adjusted_value") is not None else None,
                    round(lead["mean_exploitability"], 1)
                    if lead.get("mean_exploitability") is not None else None,
                    round(lead["worst_exploitability"], 1)
                    if lead.get("worst_exploitability") is not None else None,
                    round(lead["mean_expected_loss"], 1)
                    if lead.get("mean_expected_loss") is not None else None,
                    lead.get("severe_turns"), lead.get("line_turns"),
                ])
                ws.cell(ws.max_row, 6).fill = (
                    GOOD_FILL if outcome == "win"
                    else (WARN_FILL if outcome == "draw" else BAD_FILL))
                _shade(ws.cell(ws.max_row, 8), lead.get("mean_exploitability"))
    if not any_rows:
        ws.append(["No audited lines in this cache "
                   "(the quick tier does not audit lines)."])
    _autosize(ws.finish())


def _turns_sheet(wb, rows):
    """Every audited turn. The evidence behind every number in the other sheets."""
    ws = _Capped(wb.create_sheet("Turns"))
    ws.append(["Team", "Opponent", "Bring (lead first)", "Their bring",
               "Lead likelihood", "Line result", "Turn", "Exploitability",
               "Expected loss", "Regret", "Equilibrium", "Our worst case",
               "Our play", "Their best answer", "Tied replies"])
    _style_header(ws)
    any_rows = False
    for r in sorted(rows, key=lambda x: (x["ours"], x["theirs"])):
        for cand in r.get("candidates") or []:
            for lead in cand.get("audit") or []:
                for t in lead.get("turns") or []:
                    any_rows = True
                    ws.append([
                        r["ours"], r["theirs"],
                        " / ".join(cand.get("bring") or []) or None,
                        " / ".join(lead.get("enemy_bring")
                                   or lead.get("lead") or []) or None,
                        round(lead["probability"], 3)
                        if lead.get("probability") is not None else None,
                        (lead.get("outcome") or "").upper() or None,
                        t.get("turn"),
                        round(t["exploitability"], 1)
                        if t.get("exploitability") is not None else None,
                        round(t["expected_loss"], 1)
                        if t.get("expected_loss") is not None else None,
                        round(t["regret"], 1) if t.get("regret") is not None else None,
                        round(t["equilibrium"], 1)
                        if t.get("equilibrium") is not None else None,
                        round(t["worst_case"], 1)
                        if t.get("worst_case") is not None else None,
                        t.get("our_play"),
                        t.get("punished_by") or ("no punish available"
                                                 if t.get("no_punish") else None),
                        t.get("tied_replies"),
                    ])
                    _shade(ws.cell(ws.max_row, 8), t.get("exploitability"))
                    _shade(ws.cell(ws.max_row, 9), t.get("expected_loss"))
                    outcome = (lead.get("outcome") or "").lower()
                    if outcome == "win":
                        ws.cell(ws.max_row, 6).fill = GOOD_FILL
                    elif outcome in ("loss", "unresolved"):
                        ws.cell(ws.max_row, 6).fill = BAD_FILL
    if not any_rows:
        ws.append(["No per-turn detail in this cache "
                   "(the quick tier does not audit lines)."])
    _autosize(ws.finish())


def _legend_sheet(wb):
    ws = wb.create_sheet("How to read this")
    ws.append(["Column", "What it means"])
    _style_header(ws)
    for row in [
        ("Plan sheet",
         "The deliverable. One COMMITTED bring-4/lead-2 per opponent and its "
         "record against every configuration of theirs that was audited. You "
         "choose your four and your lead at preview, before seeing theirs, so "
         "this is the only number describing something you can actually play. "
         "A perfect answer is Wins == Of -- 90/90 when the run used "
         "--audit-all or --effort exhaustive. Anything less, and 'Losing to' "
         "names the brings that beat the plan."),
        ("Adjusted wins",
         "THE RANKING NUMBER. Share of their plausible leads where our line "
         "WINS against an opponent who punishes correctly every turn, with "
         "each win discounted by how punishable the winning line was: a win "
         "whose worst turn hands back a whole Pokemon (180 points) counts as "
         "zero, an unpunishable win counts fully. Higher is better. It needs "
         "BOTH halves to be good, which is why it replaces ranking on "
         "exploitability alone."),
        ("Wins vs punisher",
         "The same share WITHOUT the punishment discount -- did the line win, "
         "yes or no, against an opponent answering optimally every turn. Not a "
         "win count against our own bot. Hitting the turn cap with both sides "
         "alive counts as NOT a win: a line that cannot close is not a win."),
        ("Clean wins",
         "Wins with no severe turn anywhere in the line. The strictest column: "
         "these are the lines that win without ever handing a good player "
         "more than a third of a Pokemon."),
        ("What happens",
         "On the Best lines sheet: the engine's own log for that turn -- the "
         "damage numbers, who fainted, weather and item triggers, in the order "
         "they resolved. This is the match itself, not a summary of it."),
        ("Their best answer",
         "Blank, shown as 'no punish available', when nothing they can do "
         "gains more than a couple of points. Naming a 'punish' there would be "
         "reporting the arbitrary winner of a tie as if it were a read -- "
         "measured on a real line, 46 of their 62 replies scored identically "
         "on one turn."),
        ("Tied replies",
         "How many of their actions scored the same as the one named. A big "
         "number means the position is insensitive to what they do, usually "
         "because we protected, and the 'punish' is a tie-break."),
        ("Lines won / audited",
         "Counts, not rates. Under --audit-all (and at --effort exhaustive) "
         "the denominator is all 90 of their bring-4s, so '57 / 90' means the "
         "line was played against every bring they could make. At the cheaper "
         "tiers it is a sample of their most plausible leads instead, and the "
         "denominator says so."),
        ("Adjusted wins (count)",
         "The same discount applied to counts rather than a rate, written as "
         "'X / N' against the number of lines audited -- so under --audit-all "
         "it reads directly as 'this many wins out of 90, after paying for how "
         "punishable each one was'."),
        ("Adjusted value",
         "On the Lines sheet: one line's own contribution to Adjusted wins, "
         "0 to 1. Zero means the line lost or was too punishable to count."),
        ("Line result",
         "On the Turns sheet: whether the line those turns belong to ended in "
         "a WIN, LOSS, DRAW, or UNRESOLVED (turn cap reached)."),
        ("Exploitability",
         "Equilibrium value of the turn minus the worst case of the play we "
         "actually made. It is what a good player GAINS by answering us "
         "correctly. 0 = unpunishable. Lower is better. This is the ranking "
         "number, and it is not a win rate."),
        ("Regret",
         "How much better the safest available play would have been. "
         "Exploitability compares against equilibrium (which may require "
         "mixing); regret compares against the best single play, so it is the "
         "part that was avoidable by choosing differently."),
        ("Equilibrium",
         "Value of the turn if both sides play optimally. Points, not "
         "probability: a KO is worth roughly 180."),
        ("Our worst case",
         "Value of the play we chose against their best reply. Equilibrium "
         "minus this is the exploitability."),
        ("Severe turns",
         f"Turns where a best-responding opponent gains more than {SEVERE:g} "
         "points -- about a third of a Pokemon. These are the plays to change."),
        ("Lead likelihood",
         "How plausible that opening is for them, from the preview screen. "
         "Ratings are weighted by it so a lead no good player brings cannot "
         "drag a team down."),
        ("Games won / played",
         "The older win-count verification against our own scripted opponent. "
         "A biased measure of team strength, which is why the ranking does not "
         "use it -- but do not ignore it either, for the reason in the next "
         "row."),
        ("Low rating, few wins",
         "Read these two columns TOGETHER. Exploitability is measured against "
         "each turn's equilibrium, so a position that is simply lost has "
         "nothing left for the opponent to gain and rates near zero. A team "
         "that loses most of its games while rating well was played well in a "
         "bad spot -- it is not a good team. The 'Read with care' column flags "
         "any team winning under half its games."),
        ("Colours",
         f"Green up to {MILD:g}, amber up to {SEVERE:g}, red above."),
    ]:
        ws.append(list(row))
        ws.cell(ws.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    return ws


def _team_sheets_sheet(wb, sheets):
    """What each generated team actually IS: members, item, ability, EVs, moves.

    Without this the workbook names "gen03" and there is no way to see which
    six Pokemon that is, let alone which four moves each runs -- so a promising
    result cannot be taken to a game without re-running generation.
    """
    ws = wb.create_sheet("Team sheets")
    ws.append(["Team", "Pokemon", "Types", "Item", "Ability", "Nature",
               "EVs (hp/atk/def/spa/spd/spe)", "Moves", "Sets"])
    _style_header(ws)
    for name, members in (sheets or {}).items():
        for i, m in enumerate(members):
            ws.append([name if i == 0 else None, m.get("pokemon"),
                       m.get("types"), m.get("item"), m.get("ability"),
                       m.get("nature"), m.get("evs"),
                       ", ".join(m.get("moves") or []), m.get("sets")])
    if not sheets:
        ws.append(["No team sheets for this run.",
                   "Generated teams write <shortlist>_sheets.json; pass it to "
                   "--export-sheets, or this run used library teams whose sets "
                   "are already in the data files."])
    _autosize(ws)


def audited_line_count(cache_data):
    """How many audited lines the cache holds, across every pairing.

    The one number that predicts whether a full workbook is buildable: Lines
    gets a row each, and Turns gets `line_turns` rows each.
    """
    return sum(len(cand.get("audit") or [])
               for row in rows_of(cache_data)
               for cand in row.get("candidates") or [])


# Above this many audited lines the diagnosis sheets stop being readable and
# start being slow: 20k lines is already a 20k-row Lines sheet and a Turns sheet
# in the high hundreds of thousands. `--workbook auto` drops to the light set
# here rather than spending twenty minutes producing something unopenable.
AUTO_LIGHT_LINES = 20_000


def build_workbook(cache_data, out_path, team_sheets=None, workbook="full"):
    """Write the workbook. Returns the number of pairing rows exported.

    `workbook` is "full", "light", or "auto":

      full   every sheet (subject to the per-sheet row cap)
      light  LIGHT_SHEETS only -- the plan and the committed lines, without the
             three sheets whose size scales with `brings x their configs`
      auto   light when the cache holds more than AUTO_LIGHT_LINES audited
             lines, full otherwise

    Reported as: an exhaustive run's cache "will be absolutely impossible to
    generate an xlsx based on". The Best lines sheet is the deliverable; Lines,
    Candidates and Turns are the evidence behind it, and at exhaustive settings
    the evidence is larger than any spreadsheet can hold.
    """
    rows = rows_of(cache_data)
    if workbook == "auto":
        workbook = ("light" if audited_line_count(cache_data) > AUTO_LIGHT_LINES
                    else "full")
    wanted = set(LIGHT_SHEETS) if workbook == "light" else None

    def want(name):
        return wanted is None or name in wanted

    wb = Workbook()
    # Teams always runs: it owns the default sheet, and leaving that untitled
    # sheet behind would put an empty "Sheet" first in every light workbook.
    _teams_sheet(wb, rows)
    if want("Plan"):
        _plan_sheet(wb, rows)
    if want("Matchups"):
        _matchups_sheet(wb, rows)
    if want("Best lines"):
        _best_lines_sheet(wb, rows)
    if want("Lines"):
        _lines_sheet(wb, rows)
    if want("Candidates"):
        _candidates_sheet(wb, rows)
    if want("Turns"):
        _turns_sheet(wb, rows)
    if want("Team sheets"):
        _team_sheets_sheet(wb, team_sheets)
    if want("How to read this"):
        _legend_sheet(wb)
    wb.save(out_path)
    return len(rows)
