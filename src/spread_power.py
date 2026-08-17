"""Who wins the damageslop war: spread output x how long you survive.

The idea, as it was put to me:

    "Spread damage appears to be highly valuable, as this game is what I like to
     call a 'damageslop' format. Spread damage, given it attacks two enemies at
     0.75x, is a way to maximise damage output, and chip enemies into kills by
     partner. If your spread attacker is bulky and not threatened by
     supereffective damage, it can stay on the field and trade favourably vs the
     enemy (they cannot match its output). It also is simple because there is no
     target selection, so is harder to punish and harder to switch in on."

Two halves, and the product is the point.

OUTPUT. A spread move hits both foes at 0.75x, so its turn total is 1.5x a
single-target hit of the same power -- before any of the multipliers that make
the good ones absurd. The ones that matter are all modelled in `damage.py`
already and are simply switched on here: SELF-GENERATED WEATHER (Drought sun on
Heat Wave, Drizzle rain on Muddy Water), the -ate abilities (Pixilate turns
Hyper Voice into a 1.2x Fairy move, and then Fairy Feather multiplies it again),
Liquid Voice, Sheer Force, Aura.

SURVIVAL. Output you do not get to repeat is worth one turn of output. So the
score multiplies by how many hits the attacker takes to fall:

    hits_to_ko = 1 / mean(incoming physical, incoming special)

against a generic attacker -- 125 in the relevant offensive stat, 100 BP.

NEUTRAL THROUGHOUT, in both directions. Every figure divides out the actual type
multiplier, so the table ranks the Pokemon rather than the matchup: a Fire move
into a benchmark that happens to resist it is not evidence about Charizard. The
type chart is what the rest of the system is for; this answers the narrower
question of who brings the most raw, repeatable, untargetable damage.

WHAT THIS IS NOT. It is a screen, not a verdict. It does not know that the
opponent has a Fairy for your Dragon, that Wide Guard exists, that your spread
move hits your own partner (Earthquake, Muddy Water, Blizzard), or that Prankster
Tailwind changes who repeats first. Nor does it price the ally-hitting downside:
a `allAdjacent` move is scored on its two FOES only, so Earthquake looks like
Heat Wave here and does not play like it. Read it as "who is worth building
around", then let the real search argue.
"""
from damage import (Combatant, MoveInfo, damage_roll, defensive_stat,
                    effective_stat, hits_ally, is_spread_move,
                    move_from_showdown)
from engine import WEATHER_SETTERS

# The benchmark defender: base 100 across HP/Def/SpD, neutral nature, no item,
# no ability. Level 50, 0 EVs, so it is a fixed yardstick rather than a
# plausible Pokemon -- the point is that every row is divided by the same thing.
BENCH_BASE = 100
BENCH_LEVEL = 50

# The benchmark attacker used for the incoming half: 125 in the offensive stat,
# 100 BP. Chosen to be a real VGC threat rather than an average one -- "how many
# hits do I take" is a question about the hits you actually face.
INCOMING_STAT = 125
INCOMING_POWER = 100

# Spread moves below this are not the phenomenon being measured. Dazzling Gleam
# at 80 is; Snarl at 55 and Icy Wind at 55 are speed control and chip, and they
# would crowd the table with rows nobody is building a game plan around.
MIN_SPREAD_POWER = 70


def _stat_at_50(base, evs=0, boost=1.0):
    """Showdown's level-50 stat formula, neutral nature."""
    return int((((2 * base + 31 + evs // 4) * BENCH_LEVEL) // 100 + 5) * boost)


def _bench_hp():
    return int(((2 * BENCH_BASE + 31) * BENCH_LEVEL) // 100 + BENCH_LEVEL + 10)


def benchmark_defender():
    """A neutral 100/100/100 target, and deliberately TYPELESS.

    The first version typed it Normal, which quietly broke the table: a Normal
    benchmark is IMMUNE to nothing but is hit 0x by nothing either -- while the
    incoming probe, also Normal, did 0 damage to every Ghost-type, so thirteen
    Ghosts came out with infinite survivability and swept the top of the
    ranking. An empty type list makes `type_multiplier` return 1.0 by
    construction, in both directions, which is what "neutral" was supposed to
    mean. STAB is unaffected -- that reads the ATTACKER's types.
    """
    hp = _bench_hp()
    stat = _stat_at_50(BENCH_BASE)
    return Combatant(
        name="Benchmark", types=[], ability="", item="",
        stats={"hp": hp, "atk": stat, "def": stat, "spa": stat, "spd": stat,
               "spe": stat},
        current_hp=hp,
    )


# A move type that is in no typechart, so `damageTaken.get(...)` misses for
# every defending type and `_TYPE_MAPPING[0]` returns exactly 1.0. That is what
# TYPELESS means here, and it is exact rather than searched-for: no immunity, no
# resistance, and no STAB either, since it matches no attacker's typing.
TYPELESS = "Typeless"


def neutral_probe_type(defender, typechart):
    """The typeless probe, and its multiplier. Always (TYPELESS, 1.0).

    Kept as a function because the first two versions were not this. The first
    used a fixed Normal probe, which does 0 damage to Ghost -- all thirteen
    Ghost-types scored infinite survivability and swept the table. The second
    SEARCHED the chart for a type this defender happens to take at 1.0x, which
    is right for almost everyone and still not typeless: a Pokemon with no
    neutral type fell back to "least resisted", so its row was measured on a
    different yardstick from every other row.

    `defender` and `typechart` are unused and stay in the signature so the
    exactness is checkable -- see the test that asserts 1.0 for every Pokemon in
    the dataset.
    """
    return TYPELESS, 1.0


def spread_moves(record, moves_db, min_power=MIN_SPREAD_POWER,
                 foes_only=False):
    """Every damaging spread move this Pokemon is known to run, above `min_power`.

    Returns (MoveInfo, raw_showdown_dict) pairs, because two things that decide
    whether a move belongs in this table live only in the raw entry.

    Reads the same usage list the rest of the system reads, so a Pokemon that
    does not actually carry Heat Wave does not get credit for it.

    SUICIDE MOVES ARE EXCLUDED, and this is not a detail. The first run of the
    table put Metagross Explosion at #1 and Snorlax Self-Destruct at #2, scoring
    250 and 200 BP multiplied by four to six turns of survivability -- for a
    move that faints the user on use. The score's whole premise is REPEATABLE
    output, and `selfdestruct` is the exact negation of it.

    `foes_only` drops `allAdjacent` moves, which hit your own partner. Earthquake
    and Sludge Wave are real damageslop, but they constrain the partner slot in a
    way Heat Wave and Hyper Voice do not, so it is worth being able to ask the
    question without them.
    """
    out = []
    seen = set()
    for name, _pct in record.get("moves_usage") or []:
        if name == "Other":
            continue
        key = name.lower().replace(" ", "").replace("-", "").replace("'", "")
        if key not in moves_db or key in seen:
            continue
        seen.add(key)
        raw = moves_db[key]
        if raw.get("selfdestruct"):
            continue
        move = move_from_showdown(raw)
        if move.category == "Status" or (move.power or 0) < min_power:
            continue
        if not is_spread_move(move.target):
            continue
        if foes_only and hits_ally(move.target):
            continue
        out.append((move, raw))
    return out


def _neutral_damage(attacker, defender, move, typechart, weather, targets):
    """Average damage as a FRACTION of the defender's max HP, type divided out.

    `targets` is how many Pokemon the move hits, which is what applies the 0.75x
    spread penalty inside `damage_roll`.
    """
    physical = move.category == "Physical"
    atk_key = "atk" if physical else "spa"
    def_key = "def" if physical else "spd"
    atk_stat = effective_stat(attacker.stats[atk_key], attacker.stages[atk_key])
    if attacker.item == "Choice Band" and physical:
        atk_stat *= 1.5
    if attacker.item == "Choice Specs" and not physical:
        atk_stat *= 1.5
    def_stat = defensive_stat(defender, def_key, move)
    _lo, _hi, avg, eff = damage_roll(
        BENCH_LEVEL, move.power, atk_stat, def_stat, attacker, defender, move,
        typechart, weather=weather, num_targets_hit=targets)
    if not eff:
        return 0.0
    return (avg / eff) / (defender.stats["hp"] or 1)


def self_weather(combatant):
    """The weather this Pokemon brings on its own, or None.

    Sun on Heat Wave is not a lucky matchup, it is the same Pokemon: Drought
    fires the moment it lands. Reads the MEGA ability too, since Mega Charizard
    Y's whole identity is that the Mega is the one with Drought.
    """
    for ability in (combatant.mega_ability, combatant.ability):
        if ability and ability in WEATHER_SETTERS:
            return WEATHER_SETTERS[ability]
    return None


def incoming_per_hit(combatant, typechart, weather=None):
    """Mean fraction of `combatant`'s HP a generic 100 BP hit removes.

    The average of a physical and a special hit from 125 in the matching stat,
    neutral. Averaging the two rather than taking the worse one is deliberate:
    the question is how long it lasts against a metagame, and a Pokemon with one
    good defence and one terrible one really does survive some matchups.
    """
    probe_type, probe_mult = neutral_probe_type(combatant, typechart)
    total = 0.0
    for category, key in (("Physical", "def"), ("Special", "spd")):
        move = MoveInfo(name=f"Benchmark {category}", power=INCOMING_POWER,
                        move_type=probe_type, category=category, target="normal")
        def_stat = defensive_stat(combatant, key, move)
        attacker = benchmark_defender()
        attacker.stats[("atk" if key == "def" else "spa")] = INCOMING_STAT
        _lo, _hi, avg, eff = damage_roll(
            BENCH_LEVEL, INCOMING_POWER, INCOMING_STAT, def_stat, attacker,
            combatant, move, typechart, weather=weather, num_targets_hit=1)
        eff = eff or probe_mult
        total += (avg / eff) / (combatant.max_hp() or 1)
    return total / 2.0


def rate(name, merged, natures, moves_db, typechart, item=None,
         min_power=MIN_SPREAD_POWER, foes_only=False):
    """One row of the table, or None if this Pokemon runs no real spread move.

    Returns a dict: the best spread move, its per-target and two-target neutral
    output, how many hits the attacker takes, and the product.
    """
    from combatants import make_combatant
    record = merged.get(name)
    if not record:
        return None
    attacker = make_combatant(name, merged, natures, item=item)
    moves = spread_moves(record, moves_db, min_power, foes_only)
    if not moves:
        return None
    weather = self_weather(attacker)
    defender = benchmark_defender()

    # The MEGA's offence, where there is one: a Mega Charizard Y rated on base
    # Charizard's 109 SpA is not the Pokemon anyone is talking about.
    view = attacker
    if attacker.mega_stats:
        from copy import deepcopy
        view = deepcopy(attacker)
        view.stats = dict(attacker.mega_stats)
        if attacker.mega_ability:
            view.ability = attacker.mega_ability
        if attacker.mega_types:
            view.types = list(attacker.mega_types)

    best = None
    for move, raw in moves:
        per_target = _neutral_damage(view, defender, move, typechart, weather, 2)
        if best is None or per_target > best[1]:
            best = (move, per_target, raw)
    move, per_target, raw = best
    hits = incoming_per_hit(attacker, typechart, weather)
    hits_to_ko = (1.0 / hits) if hits > 0 else float("inf")
    two_target = per_target * 2.0
    return {
        "name": name,
        "move": move.name,
        "type": move.move_type,
        "power": move.power,
        "weather": weather,
        "ability": view.ability,
        # Earthquake / Sludge Wave hit our own partner: real output, but it
        # constrains the slot next to it, which the score does not price.
        "hits_ally": hits_ally(move.target),
        # Eruption and Water Spout scale with the user's CURRENT hp, so this
        # row is a full-health number and decays as the game goes on -- the
        # opposite of the repeatability the score is meant to reward.
        "hp_scaled": bool(raw.get("basePowerCallback")),
        "per_target": per_target,
        "two_target": two_target,
        "incoming": hits,
        "hits_to_ko": hits_to_ko,
        "score": two_target * hits_to_ko,
    }


def table(names, merged, natures, moves_db, typechart,
          min_power=MIN_SPREAD_POWER, foes_only=False):
    """Every Pokemon in `names` that runs a real spread move, best score first."""
    rows = []
    for name in names:
        row = rate(name, merged, natures, moves_db, typechart,
                   min_power=min_power, foes_only=foes_only)
        if row is not None:
            rows.append(row)
    rows.sort(key=lambda r: -r["score"])
    return rows


def write_workbook(rows, path, scope="", min_power=MIN_SPREAD_POWER,
                   foes_only=False):
    """The whole table as an .xlsx, so it can be sorted and checked by hand.

    Two sheets: the ranking, and a legend that states every assumption. The
    legend exists because the numbers are only meaningful next to their
    definition -- "67.0%" is a claim about a specific benchmark, and a
    spreadsheet detached from that is a spreadsheet that will be misread.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", start_color="2C3E50", end_color="2C3E50")
    good = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
    warn = PatternFill("solid", start_color="FCF3CF", end_color="FCF3CF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Spread attackers"
    cols = ["#", "Pokemon", "Spread move", "Type", "BP", "Weather", "Ability",
            "1 target %", "2 targets %", "Taken per hit %", "Hits to KO",
            "SCORE", "Hits ally", "Full-HP only"]
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "C2"

    for i, r in enumerate(rows, start=1):
        ws.append([
            i, r["name"], r["move"], r["type"], r["power"],
            r["weather"] or "-", r["ability"] or "-",
            round(r["per_target"] * 100, 1), round(r["two_target"] * 100, 1),
            round(r["incoming"] * 100, 1), round(r["hits_to_ko"], 2),
            round(r["score"], 2),
            "yes" if r["hits_ally"] else "", "yes" if r["hp_scaled"] else "",
        ])
        # Shade the score column so the top of the table reads at a glance, and
        # amber the two caveat flags rather than hiding them in a text column.
        ws.cell(ws.max_row, 12).fill = good if r["score"] >= 4.0 else warn
        for col in (13, 14):
            if ws.cell(ws.max_row, col).value:
                ws.cell(ws.max_row, col).fill = warn

    for idx, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(11, len(name) + 2)
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.auto_filter.ref = ws.dimensions

    legend = wb.create_sheet("How to read this")
    for row in (
        ["What this is"],
        ["Spread attackers ranked by REPEATABLE, UNTARGETABLE damage: how much "
         "a spread move removes across both foes, times how many hits the "
         "attacker survives to keep doing it."],
        [],
        ["Scope", scope],
        ["Minimum spread power", min_power],
        ["allAdjacent moves", "excluded" if foes_only else "included, flagged"],
        [],
        ["Column", "Means"],
        ["1 target %", "Neutral damage as a % of the benchmark's max HP, one "
                       "target, with the 0.75x spread penalty applied."],
        ["2 targets %", "The same doubled: what the turn actually removes. "
                        "1.5x a single-target hit of the same power."],
        ["Taken per hit %", f"Mean % of its own HP removed by a typeless "
                            f"{INCOMING_POWER} BP hit from a {INCOMING_STAT} "
                            f"offensive stat, averaged over physical and "
                            f"special."],
        ["Hits to KO", "1 / taken. How many of those it survives."],
        ["SCORE", "2 targets % x hits to KO."],
        ["Hits ally", "allAdjacent: Earthquake and Sludge Wave hit your own "
                      "partner, which constrains the slot beside them in a way "
                      "the score does not price."],
        ["Full-HP only", "Eruption and Water Spout scale with the user's "
                         "CURRENT HP, so the row is a full-health number and "
                         "decays -- the opposite of the repeatability the "
                         "score rewards."],
        [],
        ["The benchmark"],
        ["Defender", f"base {BENCH_BASE} HP / {BENCH_BASE} Def / {BENCH_BASE} "
                     f"SpD at level {BENCH_LEVEL}, 0 EVs, and TYPELESS."],
        ["Typeless, both ways", "Outgoing damage hits a defender with no types, "
                                "so the multiplier is 1.0 by construction. "
                                "Incoming uses a move type that is in no "
                                "typechart, so it is 1.0 against every real "
                                "Pokemon -- no immunity, no resistance, no "
                                "STAB. The table ranks the Pokemon, not the "
                                "matchup."],
        ["STAB is kept", "It reads the ATTACKER's own typing, so it is a "
                         "property of the Pokemon rather than of the matchup."],
        ["Megas", "Rated on the Mega's stats, ability and typing where the "
                  "Pokemon has one -- Charizard Y's Drought is the whole "
                  "point of Charizard Y."],
        [],
        ["What is excluded"],
        ["Self-destruct moves", "Explosion and Self-Destruct are not repeatable "
                                "output. They came out #1 and #2 before this."],
        [],
        ["What this does NOT know"],
        ["", "Wide Guard, which is the direct counter to a spread plan."],
        ["", "Whether the opponent resists your one spread type."],
        ["", "Redirection (Follow Me / Rage Powder) and Prankster Tailwind."],
        ["", "Items and EV spreads beyond the usage defaults."],
        ["", "It is a screen for who is worth building around, not a verdict."],
    ):
        legend.append(row)
    legend["A1"].font = Font(bold=True, size=13)
    for cell in ("A8", "A18", "A25", "A28"):
        legend[cell].font = Font(bold=True)
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 96
    for row in legend.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return len(rows)
