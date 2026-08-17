"""The stage-2 cache and the workbook must stay a size a person can open.

Reported: "the stage2-only is generating jsons which are 200MB-1.5GB. I should
not be generating a JSON this large in this case ... it will be absolutely
impossible to generate an xlsx based on this."

Both halves are real and they are separate problems.

THE CACHE. The per-turn record is ~890 bytes (of which `events` is 479 and
`hp_after` 187), an 18-turn audited line is 10.6 kB, and `--audit-all
--brings 90` audits 90 x 90 = 8100 lines a pairing. That is 86 MB a pairing and
~690 MB across eight opponents, which reproduces the report.

THE WORKBOOK. Those same 8100 lines a pairing give the Turns sheet roughly
1.2 million rows across eight opponents, above Excel's 1,048,576-row limit, so
the export does not merely take a long time -- it raises, after the search has
already spent the night.

The tests below pin the two mechanisms that bound each. Note in particular
`test_light_is_bounded_when_every_line_is_bad`, which is the one that caught
the first attempt: "light" was a PREDICATE ("keep it if the line lost or had a
severe turn") and measured 0% saving on a real thorough pairing, because all 24
of its lines qualified. A predicate cannot bound anything when the expensive
case is the one that satisfies it.
"""
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from openpyxl import load_workbook  # noqa: E402

import export_search  # noqa: E402
from export_search import (AUTO_LIGHT_LINES, LIGHT_SHEETS,  # noqa: E402
                           audited_line_count, build_workbook)
from matchup_search import (DETAIL_LEVELS, LIGHT_LINE_BUDGET,  # noqa: E402
                            _detailed_lines)


class _Report:
    """The fields `_detailed_lines` reads off a LineReport."""

    def __init__(self, outcome, mean_exploitability, severe_count=0):
        self.outcome = outcome
        self.mean_exploitability = mean_exploitability
        self.severe_count = severe_count


def _per_lead(reports):
    return [((f"L{i}a", f"L{i}b"), 0.1, r) for i, r in enumerate(reports)]


class DetailBudget(unittest.TestCase):
    def test_full_keeps_everything_and_summary_keeps_nothing(self):
        per_lead = _per_lead([_Report("win", 5.0) for _ in range(40)])
        self.assertEqual(len(_detailed_lines(per_lead, "full")), 40)
        self.assertEqual(_detailed_lines(per_lead, "summary"), set())

    def test_light_is_bounded_when_every_line_is_bad(self):
        """The measurement that rejected the first implementation.

        Forty lines, every one a loss with a severe turn -- the shape of an
        exhaustive run against a matchup that does not work. The predicate
        version kept all forty and saved nothing. A budget keeps six.
        """
        per_lead = _per_lead([_Report("loss", 90.0, severe_count=3)
                              for _ in range(40)])
        kept = _detailed_lines(per_lead, "light")
        self.assertEqual(len(kept), LIGHT_LINE_BUDGET)

    def test_light_is_bounded_when_every_line_is_clean(self):
        """And the other extreme: forty clean wins still cost only six."""
        per_lead = _per_lead([_Report("win", 1.0) for _ in range(40)])
        self.assertEqual(len(_detailed_lines(per_lead, "light")),
                         LIGHT_LINE_BUDGET)

    def test_light_keeps_the_losses_first(self):
        """The budget is spent where the diagnosis is. Two losses among many
        wins must both survive, and the wins that survive alongside them are
        the most punished ones."""
        reports = [_Report("win", float(i)) for i in range(20)]
        reports[7] = _Report("loss", 3.0)
        reports[13] = _Report("loss", 2.0)
        kept = _detailed_lines(_per_lead(reports), "light")
        self.assertIn(7, kept)
        self.assertIn(13, kept)
        # The remaining four go to the worst-punished wins: 19, 18, 17, 16.
        self.assertEqual(kept, {7, 13, 19, 18, 17, 16})

    def test_a_short_audit_is_untouched(self):
        """Below the budget, "light" and "full" are the same thing -- the
        cheap tiers must not lose detail they were not costing anything for."""
        per_lead = _per_lead([_Report("win", 5.0) for _ in range(4)])
        self.assertEqual(_detailed_lines(per_lead, "light"),
                         _detailed_lines(per_lead, "full"))

    def test_every_level_is_reachable(self):
        for level in DETAIL_LEVELS:
            self.assertIsInstance(_detailed_lines(_per_lead([]), level), set)


def _line(outcome="win", turns=2, exploitability=10.0):
    return {
        "lead": ["A", "B"], "enemy_bring": ["A", "B", "C", "D"],
        "probability": 0.3, "mean_exploitability": exploitability,
        "mean_expected_loss": exploitability / 2,
        "worst_exploitability": exploitability, "adjusted_value": 0.5,
        "severe_turns": 0, "outcome": outcome, "final_margin": 12.0,
        "line_turns": turns,
        "turns": [{"turn": i + 1, "exploitability": exploitability,
                   "expected_loss": 1.0, "regret": 0.5, "equilibrium": 5.0,
                   "worst_case": 4.0, "severe": False, "tied_replies": 1,
                   "our_play": "X move", "events": ["something happened"],
                   "kos": [], "hp_after": ["X 100/100"],
                   "punished_by": "Y move", "no_punish": False}
                  for i in range(turns)],
    }


def _cache(pairings=1, lines_per_candidate=2, candidates=1, turns=2):
    data = {}
    for p in range(pairings):
        cand = [{"bring": ["A", "B", "C", "D"], "exploitability": 20.0,
                 "adjusted_win_rate": 0.5, "robust_win_rate": 0.5,
                 "reliable_wins": 0.4, "severe_turns": 0, "rated_turns": turns,
                 "worst_margin": -5.0, "solver_wins": 70, "solver_total": 90,
                 "hardest_lead": ["A", "B"], "worst_turn": None,
                 "audit": [_line(turns=turns)
                           for _ in range(lines_per_candidate)]}
                for _ in range(candidates)]
        data[f"k{p}"] = {"ours": f"Ours{p}", "theirs": "Theirs",
                         "bring": ["A", "B", "C", "D"], "exploitability": 20.0,
                         "candidates": cand}
    return data


class WorkbookSize(unittest.TestCase):
    def _build(self, data, **kw):
        path = os.path.join(tempfile.mkdtemp(), "out.xlsx")
        build_workbook(data, path, **kw)
        return load_workbook(path)

    def test_audited_line_count_totals_the_cache(self):
        self.assertEqual(
            audited_line_count(_cache(pairings=3, candidates=2,
                                      lines_per_candidate=5)), 30)

    def test_light_drops_the_sheets_that_scale(self):
        wb = self._build(_cache(), workbook="light")
        self.assertEqual(set(wb.sheetnames), set(LIGHT_SHEETS))
        for gone in ("Lines", "Candidates", "Turns"):
            self.assertNotIn(gone, wb.sheetnames)

    def test_light_still_answers_the_question(self):
        """Dropping sheets must not drop the ANSWER. Best lines is the sheet
        that says what to commit to, so it has to survive the trim."""
        wb = self._build(_cache(), workbook="light")
        self.assertIn("Best lines", wb.sheetnames)
        rows = list(wb["Best lines"].iter_rows(values_only=True))
        self.assertGreater(len(rows), 1)
        self.assertTrue(any(r[2] and "A / B / C / D" in str(r[2])
                            for r in rows[1:]))

    def test_full_keeps_them(self):
        wb = self._build(_cache(), workbook="full")
        for present in ("Lines", "Candidates", "Turns"):
            self.assertIn(present, wb.sheetnames)

    def test_auto_stays_full_for_a_small_cache(self):
        wb = self._build(_cache(), workbook="auto")
        self.assertIn("Turns", wb.sheetnames)

    def test_auto_goes_light_for_a_big_one(self):
        self.assertGreater(AUTO_LIGHT_LINES, 0)
        old = export_search.AUTO_LIGHT_LINES
        export_search.AUTO_LIGHT_LINES = 5
        try:
            wb = self._build(_cache(lines_per_candidate=6, turns=0),
                             workbook="auto")
        finally:
            export_search.AUTO_LIGHT_LINES = old
        self.assertNotIn("Turns", wb.sheetnames)
        self.assertIn("Best lines", wb.sheetnames)

    def test_a_line_with_no_stored_turns_still_appears(self):
        """The bug that made --detail unusable: `enumerate(turns) or [...]`
        never fell through, because a generator is always truthy, so a line
        whose turns were trimmed vanished from Best lines entirely. Under
        `light` the trimmed lines are the WINS, so the plan's record read as
        all losses."""
        data = _cache()
        for cand in data["k0"]["candidates"]:
            for ln in cand["audit"]:
                ln["turns"] = []
        wb = self._build(data, workbook="light")
        rows = list(wb["Best lines"].iter_rows(values_only=True))[1:]
        results = [r[6] for r in rows if r[6]]
        self.assertEqual(results, ["WIN", "WIN"])

    def test_a_sheet_stops_at_the_cap_and_says_so(self):
        """Over the cap, the export must degrade to a note rather than raise
        after the search has already spent the night."""
        old = export_search.MAX_SHEET_ROWS
        export_search.MAX_SHEET_ROWS = 20
        try:
            wb = self._build(_cache(lines_per_candidate=200, turns=0),
                             workbook="full")
        finally:
            export_search.MAX_SHEET_ROWS = old
        rows = list(wb["Lines"].iter_rows(values_only=True))
        self.assertLessEqual(len(rows), 21)
        self.assertIn("not written", str(rows[-1][0]))


if __name__ == "__main__":
    unittest.main()
