"""The workbook export must survive every cache shape the search can produce.

A run that took all night must not fail at the last step because one pairing
was rated at a tier that computes no audit, so the tests below are mostly about
partial and missing data rather than the happy path.
"""
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from export_search import MILD, SEVERE, build_workbook, rows_of  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


def _turn(turn, exploitability):
    return {"turn": turn, "exploitability": exploitability, "regret": 1.0,
            "equilibrium": 10.0, "worst_case": 10.0 - exploitability,
            "severe": exploitability > SEVERE, "expected_loss": exploitability / 2,
            "our_play": f"A move{turn}", "punished_by": "B answer"}


def _candidate(bring, exploitability, turns=2):
    return {
        "bring": bring, "exploitability": exploitability, "severe_turns": 1,
        "rated_turns": turns, "worst_margin": -12.0, "solver_wins": 80,
        "solver_total": 90, "hardest_lead": bring[:2],
        "worst_turn": {"turn": 1, "exploitability": exploitability,
                       "our_play": "A move1", "punished_by": "B answer"},
        "adjusted_win_rate": 0.5, "robust_win_rate": 0.5, "reliable_wins": 0.25,
        "audit": [{"lead": bring[:2], "probability": 0.4,
                   "mean_exploitability": exploitability, "severe_turns": 1,
                   "outcome": "win", "final_margin": 42.0, "line_turns": turns,
                   "enemy_bring": bring[:2] + ["P", "Q"],
                   "mean_expected_loss": exploitability / 2,
                   "worst_exploitability": exploitability,
                   "adjusted_value": 0.5,
                   "turns": [_turn(i + 1, exploitability) for i in range(turns)]}],
    }


def _rated_row(ours, theirs, exploitability):
    cands = [_candidate(["A", "B", "C", "D"], exploitability),
             _candidate(["A", "B", "C", "E"], exploitability + 20)]
    return {"ours": ours, "theirs": theirs, "bring": cands[0]["bring"],
            "exploitability": exploitability, "severe_turns": 1,
            "adjusted_win_rate": 0.5, "robust_win_rate": 0.5,
            "reliable_wins": 0.25,
            "solver_wins": 80, "solver_total": 90,
            "hardest_lead": ["X", "Y"], "worst_turn": cands[0]["worst_turn"],
            "candidates": cands}


def _quick_row(ours, theirs):
    """What the quick tier stores: searched, but never rated."""
    return {"ours": ours, "theirs": theirs, "bring": ["A", "B", "C", "D"],
            "exploitability": None, "severe_turns": None,
            "solver_wins": 88, "solver_total": 90, "candidates": []}


def _build(cache):
    path = os.path.join(tempfile.mkdtemp(), "out.xlsx")
    count = build_workbook(cache, path)
    return load_workbook(path), count


class TestExportSearch(unittest.TestCase):

    def test_ranks_teams_by_mean_exploitability_lowest_first(self):
        cache = {"k1": _rated_row("Sun", "Rain", 90.0),
                 "k2": _rated_row("Sun", "Sand", 110.0),
                 "k3": _rated_row("NAIC", "Rain", 10.0),
                 "k4": _rated_row("NAIC", "Sand", 30.0)}
        wb, count = _build(cache)
        self.assertEqual(count, 4)
        ws = wb["Teams"]
        names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        self.assertEqual(names, ["NAIC", "Sun"])
        self.assertAlmostEqual(ws.cell(2, 8).value, 20.0)    # NAIC mean
        self.assertAlmostEqual(ws.cell(3, 8).value, 100.0)   # Sun mean
        # The worst matchup is named, since that is the one to go and fix.
        self.assertEqual(ws.cell(2, 11).value, "Sand")

    def test_every_audited_turn_reaches_the_turns_sheet(self):
        wb, _ = _build({"k": _rated_row("Sun", "Rain", 40.0)})
        ws = wb["Turns"]
        # 2 candidates x 1 lead x 2 turns
        self.assertEqual(ws.max_row - 1, 4)
        header = [c.value for c in ws[1]]
        self.assertIn("Their best answer", header)
        self.assertIn("Our play", header)
        answers = {ws.cell(r, 14).value for r in range(2, ws.max_row + 1)}
        self.assertEqual(answers, {"B answer"})

    def test_runner_up_candidates_are_kept_not_just_the_winner(self):
        wb, _ = _build({"k": _rated_row("Sun", "Rain", 40.0)})
        ws = wb["Candidates"]
        self.assertEqual(ws.max_row - 1, 2)
        self.assertEqual([ws.cell(r, 5).value for r in range(2, 4)], [40.0, 60.0])

    def test_unrated_rows_are_exported_blank_rather_than_dropped(self):
        """A searched-but-unrated pairing is information, not an error."""
        wb, count = _build({"k1": _rated_row("Sun", "Rain", 40.0),
                            "k2": _quick_row("Sun", "Sand")})
        self.assertEqual(count, 2)
        ws = wb["Matchups"]
        self.assertEqual(ws.max_row - 1, 2)
        opponents = {ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
        self.assertEqual(opponents, {"Rain", "Sand"})
        # The Teams sheet counts it as searched but not rated.
        teams = wb["Teams"]
        self.assertEqual(teams.cell(2, 14).value, 1)   # rated
        self.assertEqual(teams.cell(2, 15).value, 2)   # searched

    def test_cache_with_no_audit_anywhere_still_writes_a_workbook(self):
        wb, count = _build({"k": _quick_row("Sun", "Rain")})
        self.assertEqual(count, 1)
        # Sheets exist and say why they are empty rather than being absent.
        self.assertIn("does not audit", wb["Turns"].cell(2, 1).value)
        self.assertIn("No per-candidate detail", wb["Candidates"].cell(2, 1).value)

    def test_non_pairing_entries_in_the_cache_are_ignored(self):
        """The cache is a plain dict and may hold other bookkeeping."""
        cache = {"k": _rated_row("Sun", "Rain", 40.0),
                 "meta": {"note": "not a pairing"}, "junk": 7}
        self.assertEqual(len(rows_of(cache)), 1)
        _wb, count = _build(cache)
        self.assertEqual(count, 1)

    def test_empty_cache_produces_an_empty_but_valid_workbook(self):
        wb, count = _build({})
        self.assertEqual(count, 0)
        self.assertEqual(wb["Teams"].max_row, 1)
        self.assertIn("How to read this", wb.sheetnames)

    def test_a_team_that_loses_everything_is_flagged_despite_a_good_rating(self):
        """The trap this column exists for.

        Exploitability is measured against each turn's equilibrium, so a LOST
        position has nothing left to punish and rates near zero. Observed live:
        a team beating 0/6 configurations rated 8. Without the flag it would
        top a table headed "least punishable".
        """
        losing = _rated_row("Doomed", "King", 8.0)
        losing["solver_wins"], losing["solver_total"] = 0, 6
        healthy = _rated_row("Fine", "King", 45.0)     # 80/90 by default
        wb, _ = _build({"a": losing, "b": healthy})
        ws = wb["Teams"]
        flags = {ws.cell(r, 1).value: ws.cell(r, 16).value
                 for r in range(2, ws.max_row + 1)}
        # It still SORTS first -- the rating is what it is -- but it cannot be
        # read without the warning and the win count sitting next to it.
        self.assertEqual(ws.cell(2, 1).value, "Doomed")
        self.assertIsNotNone(flags["Doomed"])
        self.assertIn("loses most games", flags["Doomed"])
        self.assertIsNone(flags["Fine"])
        self.assertEqual(ws.cell(2, 9).value, None)   # 0 wins, shown as blank
        self.assertEqual(ws.cell(2, 10).value, 6)

    def test_every_audited_line_reaches_the_lines_sheet(self):
        """Under --audit-all this is 90 rows per bring; none may be dropped."""
        wb, _ = _build({"k": _rated_row("Sun", "Rain", 40.0)})
        ws = wb["Lines"]
        self.assertEqual(ws.max_row - 1, 2)          # 2 candidates x 1 line
        header = [c.value for c in ws[1]]
        self.assertIn("Their bring (lead first)", header)
        self.assertIn("Adjusted value", header)
        self.assertIn("Result", header)
        # Their FULL bring is named, not just the lead pair -- the whole point
        # of auditing backs is being able to tell the four apart.
        self.assertEqual(ws.cell(2, 4).value, "A / B / P / Q")
        self.assertEqual(ws.cell(2, 6).value, "WIN")
        self.assertAlmostEqual(ws.cell(2, 7).value, 0.5)

    def test_line_counts_appear_beside_the_rates(self):
        """A rate without its denominator cannot be trusted."""
        wb, _ = _build({"k": _rated_row("Sun", "Rain", 40.0)})
        ws = wb["Teams"]
        header = [c.value for c in ws[1]]
        self.assertEqual(header[4:7],
                         ["Lines won", "Lines audited", "Adjusted wins (count)"])
        # Written against the denominator, not as a bare decimal.
        self.assertEqual(ws.cell(2, 5).value, "2 / 2")
        self.assertEqual(ws.cell(2, 6).value, 2)
        self.assertEqual(ws.cell(2, 7).value, "1.0 / 2")   # 0.5 + 0.5

    def test_a_losing_line_is_shown_not_hidden(self):
        row = _rated_row("Sun", "Rain", 40.0)
        row["candidates"][0]["audit"][0]["outcome"] = "loss"
        row["candidates"][0]["audit"][0]["adjusted_value"] = 0.0
        wb, _ = _build({"k": row})
        ws = wb["Lines"]
        results = {ws.cell(r, 6).value for r in range(2, ws.max_row + 1)}
        self.assertEqual(results, {"LOSS", "WIN"})
        self.assertEqual(wb["Teams"].cell(2, 5).value, "1 / 2")

    def test_a_turn_with_no_real_punish_says_so_instead_of_naming_one(self):
        """The nonsense this removes: a 46-way tie reported as a read."""
        row = _rated_row("Sun", "Rain", 40.0)
        turn = row["candidates"][0]["audit"][0]["turns"][0]
        turn["punished_by"], turn["no_punish"] = None, True
        turn["tied_replies"] = 46
        wb, _ = _build({"k": row})
        ws = wb["Turns"]
        answers = [ws.cell(r, 14).value for r in range(2, ws.max_row + 1)]
        self.assertIn("no punish available", answers)
        tied = [ws.cell(r, 15).value for r in range(2, ws.max_row + 1)]
        self.assertIn(46, tied)

    def test_severity_shading_thresholds_bracket_the_severe_constant(self):
        wb, _ = _build({"a": _rated_row("Low", "X", MILD - 1),
                        "b": _rated_row("High", "X", SEVERE + 1)})
        ws = wb["Teams"]
        fills = {ws.cell(r, 1).value: ws.cell(r, 8).fill.start_color.rgb
                 for r in range(2, ws.max_row + 1)}
        self.assertNotEqual(fills["Low"], fills["High"])


if __name__ == "__main__":
    unittest.main()
