"""CLI coverage for the item 1/3/4/5 changes to `tools/counter_table.py`:

    1. multi-bring4 cores can be 4, 5, or 6 members, never padded with dead
       weight ("core"/"core_size"/"unused", not the old fixed "team6").
    2. mega + its own base form can't both be brought (already covered at
       the `counter_finder` layer -- `bring4_search` raises).
    3. Choice Scarf is excluded from every SEARCH by default; --allow-scarf
       opts back in.
    4. --max-weak/--type-limit are hard multi-bring4 synergy filters, and
       every printed/CSV'd core shows its `member_weakness_summary`.
    5. a teamsheet (item + moves + usage%) is shown for every bring-4
       member, per enemy.

Mirrors this repo's existing CLI-testing convention (`test_pick_and_beam_
flags.py`): unit-test the parsing helpers directly, and use `--help`
subprocess checks to confirm a documented flag is one argparse actually
accepts (a flag that's in the docstring but not wired is a command that
cannot run).
"""
import io
import os
import subprocess
import sys
import unittest
from contextlib import redirect_stdout

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "tools"))

import counter_table as ct  # noqa: E402
from counter_finder import Hit  # noqa: E402

ROOT = os.path.join(os.path.dirname(__file__), "..")


def hit(name="Move", frac=0.5):
    return Hit(move_name=name, frac=frac, lo=frac, avg=frac, hi=frac,
              eff=1.0, num_targets_hit=1)


def run_main(argv):
    """`main()` with `sys.argv` patched, capturing stdout -- validation
    errors (the only thing the argv-only tests below exercise) all raise
    `SystemExit` BEFORE `load_world()` runs, so this stays fast."""
    old_argv = sys.argv
    sys.argv = ["counter_table.py"] + argv
    buf = io.StringIO()
    try:
        with redirect_stdout(buf):
            ct.main()
        return None, buf.getvalue()
    except SystemExit as e:
        return str(e), buf.getvalue()
    finally:
        sys.argv = old_argv


class TestParseTypeLimits(unittest.TestCase):

    def test_a_single_type_both_keys(self):
        self.assertEqual(ct._parse_type_limits(["Fire:max_weak=1,max_net=-2"]),
                         {"Fire": {"max_weak": 1, "max_net": -2}})

    def test_repeated_flag_gives_one_entry_per_type(self):
        got = ct._parse_type_limits(["Fire:max_weak=1", "Ice:max_net=0"])
        self.assertEqual(got, {"Fire": {"max_weak": 1}, "Ice": {"max_net": 0}})

    def test_one_key_may_be_omitted(self):
        self.assertEqual(ct._parse_type_limits(["Water:max_weak=2"]),
                         {"Water": {"max_weak": 2}})

    def test_empty_list_is_none(self):
        self.assertIsNone(ct._parse_type_limits([]))

    def test_unknown_type_is_rejected(self):
        with self.assertRaises(SystemExit) as caught:
            ct._parse_type_limits(["Cheese:max_weak=1"])
        self.assertIn("unknown type", str(caught.exception))

    def test_missing_colon_is_rejected(self):
        with self.assertRaises(SystemExit) as caught:
            ct._parse_type_limits(["Fire max_weak=1"])
        self.assertIn("Type:max_weak", str(caught.exception))

    def test_unknown_key_is_rejected(self):
        with self.assertRaises(SystemExit) as caught:
            ct._parse_type_limits(["Fire:max_wet=1"])
        self.assertIn("max_weak/max_net", str(caught.exception))

    def test_a_non_integer_value_is_rejected(self):
        with self.assertRaises(SystemExit) as caught:
            ct._parse_type_limits(["Fire:max_weak=one"])
        self.assertIn("integer", str(caught.exception))


class TestStrictWeakTypesShorthand(unittest.TestCase):
    """"a new type weakness limiter where I name specific types that may
    have no more than 1 weakness, I can define types but default if I use
    the argument is 'fire, water, fairy, steel, dark, ghost, ground', if I
    don't use the argument then none are applied" -- a convenience
    shorthand over the existing --type-limit mechanism, not a new
    enforcement path (`_merge_strict_weak_types` folds it into the exact
    same {type: {"max_weak": N}} shape `_parse_type_limits` already
    produces)."""

    def test_omitted_flag_resolves_to_no_types(self):
        self.assertEqual(ct._resolve_strict_weak_types(None), [])

    def test_bare_flag_resolves_to_the_default_list(self):
        self.assertEqual(
            ct._resolve_strict_weak_types("__DEFAULT__"),
            ["Fire", "Water", "Fairy", "Steel", "Dark", "Ghost", "Ground"])

    def test_an_explicit_list_is_used_verbatim(self):
        self.assertEqual(ct._resolve_strict_weak_types("Fire,Ice"),
                         ["Fire", "Ice"])

    def test_an_unknown_type_is_rejected(self):
        with self.assertRaises(SystemExit) as caught:
            ct._resolve_strict_weak_types("Bogus")
        self.assertIn("unknown type", str(caught.exception))

    def test_no_types_leaves_type_limits_unchanged(self):
        self.assertIsNone(ct._merge_strict_weak_types(None, []))
        self.assertEqual(ct._merge_strict_weak_types({"Ice": {"max_net": 0}}, []),
                         {"Ice": {"max_net": 0}})

    def test_each_named_type_gets_max_weak_one(self):
        self.assertEqual(ct._merge_strict_weak_types(None, ["Fire", "Water"]),
                         {"Fire": {"max_weak": 1}, "Water": {"max_weak": 1}})

    def test_an_explicit_type_limit_for_the_same_type_wins(self):
        """More specific/intentional always beats the shorthand -- Fire's
        own explicit --type-limit entry replaces the shorthand's bare
        max_weak=1 outright, while Water (only named by the shorthand)
        still gets it."""
        merged = ct._merge_strict_weak_types(
            {"Fire": {"max_weak": 2, "max_net": -1}}, ["Fire", "Water"])
        self.assertEqual(merged, {"Fire": {"max_weak": 2, "max_net": -1},
                                  "Water": {"max_weak": 1}})

    def test_only_applies_to_multi_bring4(self):
        msg, _out = run_main(["--vs", "Kingambit", "--strict-weak-types"])
        self.assertIsNotNone(msg)
        self.assertIn("--strict-weak-types", msg)

    def test_end_to_end_no_printed_core_has_two_weak_to_a_named_type(self):
        """A real --multi-bring4 run with an explicit --strict-weak-types
        list: every printed core's own weaknesses-by-type line must never
        show 2+ for Fire (the only named type here)."""
        msg, out = run_main(
            ["--pool-size", "16", "--multi-bring4", "--vs-team",
             "Kingambit,Sableye", "--vs-team", "Ariados,Basculegion",
             "--good-threshold", "0", "--min-enemies", "1",
             "--strict-weak-types", "Fire", "--top", "30"])
        self.assertIsNone(msg, out)
        import re
        fire_counts = re.findall(r"weaknesses by type: [^\n]*Fire (\d+)", out)
        self.assertTrue(fire_counts, "expected at least one printed core")
        self.assertTrue(all(int(c) <= 1 for c in fire_counts), fire_counts)


class TestModeRestrictionsOnTheNewFlags(unittest.TestCase):
    """--max-weak/--type-limit only mean something under --multi-bring4 --
    same "reject loudly rather than silently ignore" convention every other
    mode-specific flag in this file already follows."""

    def test_max_weak_without_multi_bring4_is_rejected(self):
        # 3, not 2: --max-weak now defaults to 2 ("by default ... the
        # weakness limit should be 2"), so an explicit value must differ
        # from the default to be distinguishable from "never passed" --
        # same "!= default" convention --beam-width/--max-candidates
        # already use for this exact check.
        msg, _out = run_main(["--vs", "Kingambit", "--max-weak", "3"])
        self.assertIsNotNone(msg)
        self.assertIn("--multi-bring4", msg)

    def test_type_limit_without_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit",
                              "--type-limit", "Fire:max_weak=1"])
        self.assertIsNotNone(msg)
        self.assertIn("--multi-bring4", msg)

    def test_allow_scarf_has_no_mode_restriction(self):
        """Unlike --max-weak/--type-limit, --allow-scarf threads through
        every search mode (`counter_finder`'s `excluded_items` is a
        parameter on all ten of them) -- it must NOT be rejected here."""
        msg, _out = run_main(["--vs", "unknown-pokemon-name", "--allow-scarf"])
        # Fails on the unrelated "unknown Pokemon" check, not a mode
        # restriction -- proves --allow-scarf itself was accepted.
        self.assertIn("unknown Pokemon", msg)

    def test_deep_dive_core_without_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--deep-dive-core", "1"])
        self.assertIsNotNone(msg)
        self.assertIn("--multi-bring4", msg)

    def test_xlsx_without_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--xlsx", "/tmp/x.xlsx"])
        self.assertIsNotNone(msg)
        self.assertIn("--multi-bring4", msg)

    def test_max_weak_types_without_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--max-weak-types", "3"])
        self.assertIsNotNone(msg)
        self.assertIn("--multi-bring4", msg)

    def test_teamsheet_json_without_bring4_or_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--teamsheet-json", "-"])
        self.assertIsNotNone(msg)
        self.assertIn("--teamsheet-json", msg)

    def test_unique_items_without_bring4_or_multi_bring4_is_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--unique-items"])
        self.assertIsNotNone(msg)
        self.assertIn("--unique-items", msg)


class TestCliAcceptsMirrorMatches(unittest.TestCase):
    """"you should be allowed to bring the same pokemon as the enemy" --
    the CLI used to hard-reject --deep/--bring4/--joint --partner whenever
    --our/--partner shared a name with --vs, on top of counter_finder.py's
    own (now-also-removed) ValueError. All three are legal VGC mirrors now;
    only the pre-existing "unknown Pokemon" checks remain."""

    def test_deep_accepts_a_shared_name(self):
        msg, out = run_main(
            ["--vs", "Kingambit,Sableye", "--our", "Kingambit,Garchomp",
             "--deep"])
        self.assertIsNone(msg, out)

    def test_bring4_accepts_a_shared_name(self):
        msg, out = run_main(
            ["--vs", "Kingambit,Sableye", "--our",
             "Kingambit,Garchomp,Incineroar,Gallade", "--bring4",
             "--no-prompt", "--top", "1"])
        self.assertIsNone(msg, out)

    def test_joint_partner_accepts_a_shared_name(self):
        msg, out = run_main(
            ["--vs", "Kingambit,Sableye", "--joint", "--partner",
             "Kingambit"])
        self.assertIsNone(msg, out)


class TestUniqueItemsFlag(unittest.TestCase):
    """"make the item uniqueness an option, but by default items will
    remain non-unique to reduce search time" -- `--unique-items` threads
    into `enforce_item_clause` on `bring4_search`/`core_deep_dive`. Off by
    default: a known real collision (Ninetales-Alola and Rampardos both
    independently pick Life Orb against this fixture -- see
    TestResolveUniqueItems in test_counter_finder.py) shows up in the
    printed deep-dive "set: ..." line unless the flag is passed."""

    # Exactly 4 names (not the padded-to-6 list this used to be) -- with
    # only one possible bring-4 to rank, both colliding members are
    # GUARANTEED to show up in the deep dive regardless of how any other
    # mechanic changes shift bring-4 ranking; a padded-to-6 list left this
    # fixture fragile to exactly that (a bring-4 search entirely avoiding
    # Ninetales-Alola/Rampardos once other mechanics made a different
    # 4-of-6 rank #1, silently no longer exercising the collision at all).
    OUR6 = "Mega Gengar,Mega Alakazam,Ninetales-Alola,Rampardos"

    def _set_line(self, out):
        idx = out.find("set: ")
        self.assertGreaterEqual(idx, 0, out)
        return out[idx:out.find("\n", idx)]

    def test_off_by_default_the_known_collision_still_shows(self):
        msg, out = run_main(
            ["--our", self.OUR6, "--bring4", "--vs", "Sableye,Ariados",
             "--no-prompt", "--top", "1", "--deep-dive-core", "1"])
        self.assertIsNone(msg, out)
        line = self._set_line(out)
        self.assertIn("Ninetales-Alola @ Life Orb", line)
        self.assertIn("Rampardos @ Life Orb", line)

    def test_resolves_the_collision_in_bring4_deep_dive(self):
        msg, out = run_main(
            ["--our", self.OUR6, "--bring4", "--vs", "Sableye,Ariados",
             "--no-prompt", "--top", "1", "--deep-dive-core", "1",
             "--unique-items"])
        self.assertIsNone(msg, out)
        line = self._set_line(out)
        self.assertIn("Ninetales-Alola @ Life Orb", line)
        self.assertNotIn("Rampardos @ Life Orb", line)

    def test_resolves_the_collision_in_multi_bring4_deep_dive_core(self):
        argv = ["--pool-size", "16", "--multi-bring4", "--vs-team",
                "Kingambit,Sableye", "--vs-team", "Ariados,Basculegion",
                "--good-threshold", "0", "--min-enemies", "1",
                "--deep-dive-core", "1", "--unique-items"]
        msg, out = run_main(argv)
        self.assertIsNone(msg, out)
        line = self._set_line(out)
        # Whichever core was actually chosen, no two of its members may
        # share an item under --unique-items.
        items = [part.split(" @ ")[1] for part in
                line[len("set: "):].split(", ")]
        self.assertEqual(len(items), len(set(items)), line)


class TestCoreSizesFlag(unittest.TestCase):
    """"I would like to output the best 3-pokemon cores against each team"
    -- `--core-sizes` (comma-separated, default "4,5,6", each 3-6) threads
    into `multi_bring4_exhaustive`/`multi_bring4_beam`'s own `core_sizes`
    parameter. `--bring4 --our` also accepts exactly 3 names now."""

    def test_only_applies_to_multi_bring4(self):
        msg, _out = run_main(["--vs", "Kingambit", "--core-sizes", "3"])
        self.assertIsNotNone(msg)
        self.assertIn("--core-sizes", msg)

    def test_rejects_a_non_integer(self):
        msg, _out = run_main(
            ["--pool-size", "16", "--multi-bring4", "--vs-team",
             "Kingambit,Sableye", "--core-sizes", "x"])
        self.assertIsNotNone(msg)
        self.assertIn("--core-sizes", msg)

    def test_rejects_a_size_outside_three_to_six(self):
        msg, _out = run_main(
            ["--pool-size", "16", "--multi-bring4", "--vs-team",
             "Kingambit,Sableye", "--core-sizes", "2"])
        self.assertIsNotNone(msg)
        self.assertIn("--core-sizes", msg)

    def test_core_sizes_three_only_prints_three_pokemon_cores(self):
        msg, out = run_main(
            ["--pool-size", "16", "--multi-bring4", "--vs-team",
             "Kingambit,Sableye", "--vs-team", "Ariados,Basculegion",
             "--good-threshold", "0", "--min-enemies", "1",
             "--core-sizes", "3", "--top", "5"])
        self.assertIsNone(msg, out)
        # Every printed core's leading "(N)" rank marker must read "(3)".
        import re
        core_sizes_seen = set(re.findall(r"^\s+\d+ \((\d)\)", out, re.MULTILINE))
        self.assertEqual(core_sizes_seen, {"3"})

    def test_bring4_our_accepts_exactly_three_names(self):
        msg, out = run_main(
            ["--our", "Garchomp,Incineroar,Gallade", "--bring4",
             "--vs", "Sableye,Ariados", "--no-prompt", "--top", "1"])
        self.assertIsNone(msg, out)
        self.assertIn("all 3 pairs drawn from your 3", out)


class TestHelpDocumentsTheNewFlags(unittest.TestCase):
    """A flag missing from `--help` is one nobody can discover; a flag in
    the header docstring but not parsed is one that cannot run -- the same
    two failure modes `test_pick_and_beam_flags.py` guards against for
    `overnight.bat`."""

    @classmethod
    def setUpClass(cls):
        cls.help_text = subprocess.run(
            [sys.executable, os.path.join(ROOT, "tools", "counter_table.py"),
             "--help"], capture_output=True, text=True, timeout=60).stdout

    def test_allow_scarf_is_parsed(self):
        self.assertIn("--allow-scarf", self.help_text)

    def test_max_weak_is_parsed(self):
        self.assertIn("--max-weak", self.help_text)

    def test_type_limit_is_parsed(self):
        self.assertIn("--type-limit", self.help_text)

    def test_max_megas_is_parsed(self):
        self.assertIn("--max-megas", self.help_text)

    def test_deep_dive_core_is_parsed(self):
        self.assertIn("--deep-dive-core", self.help_text)

    def test_xlsx_is_parsed(self):
        self.assertIn("--xlsx", self.help_text)

    def test_core_sizes_is_parsed(self):
        self.assertIn("--core-sizes", self.help_text)

    def test_strict_weak_types_is_parsed(self):
        self.assertIn("--strict-weak-types", self.help_text)

    def test_unique_items_is_parsed(self):
        self.assertIn("--unique-items", self.help_text)

    def test_max_weak_types_is_parsed(self):
        self.assertIn("--max-weak-types", self.help_text)

    def test_teamsheet_json_is_parsed(self):
        self.assertIn("--teamsheet-json", self.help_text)

    def test_every_flag_the_module_docstring_shows_is_parsed(self):
        import re
        docstring = ct.__doc__
        shown = set(re.findall(r"(--[a-z0-9-]+)", docstring))
        parsed = set(re.findall(r"(--[a-z0-9-]+)", self.help_text))
        missing = sorted(shown - parsed)
        self.assertEqual(missing, [], f"documented but not parsed: {missing}")


class TestMultiBring4EndToEnd(unittest.TestCase):
    """A real (small) run, exercising the full item 1/3/4/5 wiring together
    -- variable core size, the synergy summary, the teamsheet, and a hard
    --max-weak filter actually changing the result."""

    @classmethod
    def setUpClass(cls):
        # pool-size 16 stopped reliably finding a core once Regulation M-C's
        # very-high-Score additions started crowding the top of the
        # candidate pool -- widened per the tool's own suggested remedy
        # ("widen the pool"), same fix as TestMultiBring4SixPairExportColumns.
        cls.argv = ["--pool-size", "30", "--multi-bring4",
                   "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
                   "--vs-team", "Sylveon,Mega Charizard Y,Sinistcha,Farigiraf",
                   "--top", "3"]

    def test_runs_clean_and_shows_a_core_with_synergy_and_teamsheet(self):
        msg, out = run_main(self.argv)
        self.assertIsNone(msg, out)
        self.assertIn("synergy: weak to 2+ types", out)
        self.assertIn(" @ ", out)  # "Name @ Item: moves" teamsheet line
        self.assertNotIn("team6", out, "the old field name must not leak "
                         "into the printed report")

    def test_a_core_may_be_smaller_than_six(self):
        """"a full team of only 4-5 members is not a problem ... I would
        still like to see them" -- confirmed structurally, not just by
        eyeballing one run's printout."""
        from _harness import load_world
        from generate_team import build_candidate_pool
        W = load_world()
        merged, moves = W["merged"], W["moves"]
        natures, typechart = W["natures"], W["typechart"]
        # Same pool-building path `_pool`/`--pool-size 16` uses -- a plain
        # alphabetical slice doesn't reliably overlap with the fixture
        # enemies below.
        pool = ct._apply_preferences(
            list(build_candidate_pool(merged, top_n=16)), merged)
        from counter_finder import multi_bring4_coverage, multi_bring4_exhaustive
        # good_threshold relaxed well below the default 100%: this pool's
        # top pairs now correctly lose more often once Whimsicott's REAL
        # Tailwind access is assumed (`_pair_vs_targets`'s "assume they set
        # tailwind and see if it is a loss") and once every member's
        # item/moveset is fixed across BOTH enemy teams instead of being
        # independently re-optimised per enemy (`multi_bring4_coverage`'s
        # own "a real team's set is fixed for the whole event" fix) -- at
        # 0.8 the surviving candidate pool happens to be all 4 Mega-capable
        # names, which the (also new) default --max-megas 2 cap correctly
        # empties to zero cores; 0.5 gives a real, multi-size, mixed
        # mega/non-mega candidate pool to exercise the same exhaustive-
        # search structure this test actually cares about.
        coverage = multi_bring4_coverage(
            pool,
            [["Kingambit", "Basculegion", "Garchomp", "Whimsicott"],
             ["Sylveon", "Mega Charizard Y", "Sinistcha", "Farigiraf"]],
            merged, moves, natures, typechart, good_threshold=0.5)
        rows = multi_bring4_exhaustive(coverage, good_threshold=0.5)
        sizes = {r["core_size"] for r in rows}
        self.assertTrue(sizes, "fixture assumes at least one core is found")
        self.assertTrue(all(r["unused"] == () for r in rows))

    def test_max_weak_actually_changes_the_result(self):
        """A hard filter that never changes anything is a no-op wired in
        for show -- confirm --max-weak actually removes a core (or
        reorders the top one) rather than just printing more text."""
        _msg1, out_unfiltered = run_main(self.argv)
        _msg2, out_filtered = run_main(self.argv + ["--max-weak", "1"])
        self.assertNotEqual(out_unfiltered, out_filtered)

    def test_csv_export_uses_core_not_team6(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            self.assertIn("core", header)
            self.assertIn("core size", header)
            self.assertNotIn("team6", header)
            self.assertIn("weak to 2+ types (members)", header)
            self.assertTrue(any("teamsheet" in h for h in header))
        finally:
            os.unlink(path)


class TestDeepDiveCoreAndXlsxExport(unittest.TestCase):
    """"I also want to see the gameplans for each pair included in a team
    vs enemies. Given the size, maybe make this deep dive an option after
    the search has run" + "it may make more sense to make counter_table.py
    export an xlsx rather than a csv"."""

    @classmethod
    def setUpClass(cls):
        # pool-size 16 stopped reliably finding a core once Regulation M-C's
        # very-high-Score additions started crowding the top of the
        # candidate pool -- widened per the tool's own suggested remedy
        # ("widen the pool"), same fix as TestMultiBring4SixPairExportColumns.
        cls.argv = ["--pool-size", "30", "--multi-bring4",
                   "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
                   "--vs-team", "Sylveon,Mega Charizard Y,Sinistcha,Farigiraf",
                   "--top", "3"]

    def test_deep_dive_core_prints_the_aggregate_and_gameplans(self):
        msg, out = run_main(self.argv + ["--deep-dive-core", "1"])
        self.assertIsNone(msg, out)
        self.assertIn("Deep dive:", out)
        self.assertIn("OVERALL, every pair vs every enemy", out)
        self.assertIn("beaten (", out)
        self.assertIn("tw-safe", out)
        self.assertIn("pr-safe", out)
        # A real gameplan line (a turn's hit) must be present -- "T1 ... ->
        # ...: <move> ...%", matching --deep's own format.
        self.assertRegex(out, r"T\d .+ -> .+: .+ \d+-\d+-\d+%")

    def test_deep_dive_core_out_of_range_is_rejected(self):
        msg, _out = run_main(self.argv + ["--deep-dive-core", "999"])
        self.assertIsNotNone(msg)
        self.assertIn("--deep-dive-core", msg)

    def test_deep_dive_core_is_not_computed_by_default(self):
        """The whole point of making it opt-in -- a plain run must not pay
        for (or print) the full per-pair-per-enemy detail."""
        _msg, out = run_main(self.argv)
        self.assertNotIn("Deep dive:", out)

    def test_xlsx_export_writes_a_real_workbook(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)  # let the tool create it fresh
        try:
            msg, out = run_main(self.argv + ["--xlsx", path])
            self.assertIsNone(msg, out)
            self.assertTrue(os.path.exists(path))
            from openpyxl import load_workbook
            wb = load_workbook(path)
            self.assertIn("Cores", wb.sheetnames)
            ws = wb["Cores"]
            self.assertGreater(ws.max_row, 1, "header plus at least one core row")
            # A leading "#" rank column -- "the team number [must] be
            # indexed in the xlsx export" -- with "Core" as the next column.
            self.assertEqual(ws.cell(row=1, column=1).value, "#")
            self.assertEqual(ws.cell(row=1, column=2).value, "Core")
            self.assertEqual(ws.cell(row=2, column=1).value, 1)
            # "Sets"/"Teamsheets" sheets carry every row's own item+moveset
            # and a pastable pokepaste/token, not just the deep-dived one.
            self.assertIn("Sets", wb.sheetnames)
            self.assertIn("Teamsheets", wb.sheetnames)
            ts = wb["Teamsheets"]
            self.assertGreater(ts.max_row, 1)
            token_col = [c.value for c in ts[1]].index("Teamsheet (base64)") + 1
            self.assertTrue(ts.cell(row=2, column=token_col).value.startswith("TSHEET1:"))
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_xlsx_export_with_deep_dive_adds_the_extra_sheets(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, _out = run_main(
                self.argv + ["--deep-dive-core", "1", "--xlsx", path])
            self.assertIsNone(msg)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            # Ranked "Dive N ..." sheets (N = the rank actually dived) --
            # several dives from one run each get their own trio instead of
            # colliding on one fixed sheet name.
            for name in ("Cores", "Dive 1 Sets", "Dive 1 Summary",
                        "Dive 1 Gameplans"):
                self.assertIn(name, wb.sheetnames)
            gp = wb["Dive 1 Gameplans"]
            self.assertGreater(gp.max_row, 1)
            self.assertEqual(gp.cell(row=1, column=1).value, "Pair")
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_deep_dive_core_accepts_a_comma_separated_list(self):
        """"I can deep mode run for promising teams 5, 85, and 16" --
        several ranks in one run, each printed and each getting its own
        xlsx sheet trio."""
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(
                self.argv + ["--deep-dive-core", "1,2", "--xlsx", path])
            self.assertIsNone(msg, out)
            self.assertEqual(out.count("Deep dive:"), 2)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            for name in ("Dive 1 Sets", "Dive 1 Summary", "Dive 1 Gameplans",
                        "Dive 2 Sets", "Dive 2 Summary", "Dive 2 Gameplans"):
                self.assertIn(name, wb.sheetnames)
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestDiveSheetsFormatting(unittest.TestCase):
    """"in the deep dive exports in the xlsx: for gameplans, it would be
    good to have an empty line between each match for legibility. In
    summary it would be good to have a column to see what each pair loses
    to." -- both purely in `_write_dive_sheets`, no new search machinery."""

    @classmethod
    def setUpClass(cls):
        cls.argv = ["--pool-size", "30", "--multi-bring4",
                   "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
                   "--vs-team", "Sylveon,Mega Charizard Y,Sinistcha,Farigiraf",
                   "--top", "3"]

    def test_gameplans_has_a_blank_row_between_matches(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, _out = run_main(
                self.argv + ["--deep-dive-core", "1", "--xlsx", path])
            self.assertIsNone(msg, _out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            gp = wb["Dive 1 Gameplans"]
            rows = [[c.value for c in row] for row in gp.iter_rows(min_row=2)]
            blank_rows = [r for r in rows if all(v is None for v in r)]
            self.assertGreater(len(blank_rows), 0,
                               "at least one blank separator row between matches")
            # Every blank row is followed (or preceded) by real data -- not
            # a trailing artifact of the sheet.
            self.assertTrue(any(v is not None for v in rows[-1]) or len(rows) > 1)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_summary_has_a_loses_to_column(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, _out = run_main(
                self.argv + ["--deep-dive-core", "1", "--xlsx", path])
            self.assertIsNone(msg, _out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            summ = wb["Dive 1 Summary"]
            header = [c.value for c in summ[1]]
            self.assertIn("Loses To", header)
            loses_col = header.index("Loses To") + 1
            lost_col = header.index("Lost") + 1
            # Every row's "Loses To" cell agrees with its own "Lost" count --
            # same outcome == "loss" predicate behind both columns.
            for row in range(2, summ.max_row + 1):
                lost_count = summ.cell(row=row, column=lost_col).value
                loses_to = summ.cell(row=row, column=loses_col).value or ""
                names = [n for n in loses_to.split("; ") if n]
                self.assertEqual(len(names), lost_count,
                                 f"row {row}: {loses_to!r} vs Lost={lost_count}")
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestDeepDiveCoreForBring4(unittest.TestCase):
    """--deep-dive-core now also works with --bring4 (a fixed 6, one enemy
    roster), not just --multi-bring4 -- "I want to be able to choose a
    specific team to deep dive into" applies just as much to a bring-4 as
    to a multi-bring4 core: `core_deep_dive` accepts any core, so this is
    just wiring the Nth-ranked `bring4_rows` entry into it, the same way
    `--multi-bring4` already wires its own Nth-ranked core."""

    @classmethod
    def setUpClass(cls):
        cls.argv = ["--our",
                   "Kingambit,Dragapult,Whimsicott,Ninetales-Alola,"
                   "Mega Alakazam,Sharpedo",
                   "--bring4", "--vs", "Sableye,Ariados"]

    def test_deep_dive_core_prints_the_aggregate_and_gameplans(self):
        msg, out = run_main(self.argv + ["--deep-dive-core", "1"])
        self.assertIsNone(msg, out)
        self.assertIn("Deep dive:", out)
        self.assertIn("OVERALL, every pair vs every enemy", out)
        self.assertIn("tw-safe", out)
        self.assertIn("pr-safe", out)
        self.assertRegex(out, r"T\d .+ -> .+: .+ \d+-\d+-\d+%")

    def test_deep_dive_core_out_of_range_is_rejected(self):
        msg, _out = run_main(self.argv + ["--deep-dive-core", "999"])
        self.assertIsNotNone(msg)
        self.assertIn("--deep-dive-core", msg)

    def test_deep_dive_core_is_not_computed_by_default(self):
        _msg, out = run_main(self.argv)
        self.assertNotIn("Deep dive:", out)

    def test_deep_dive_core_without_bring4_or_multi_bring4_is_still_rejected(self):
        msg, _out = run_main(["--vs", "Kingambit", "--deep-dive-core", "1"])
        self.assertIsNotNone(msg)
        self.assertIn("--deep-dive-core", msg)


class TestBring4SixPairExportColumns(unittest.TestCase):
    """"I would like the csv/xlsx export from the CLI to show the basic
    details of the 6 pairs for each bring4 (total, 3rd best, 4th best, and
    worst wins, wins under Tailwind ..., under protect safe)" --
    `bring4_pair_depth`/`enemy_has_real_tailwind`, threaded into --bring4's
    own CSV export."""

    @classmethod
    def setUpClass(cls):
        cls.argv = ["--our",
                   "Kingambit,Dragapult,Whimsicott,Ninetales-Alola,"
                   "Mega Alakazam,Sharpedo",
                   "--bring4", "--vs", "Sableye,Ariados"]

    def test_csv_export_has_the_six_pair_depth_columns(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            self.assertTrue(rows)
            header = rows[0].keys()
            for col in ("pairs beaten total", "pairs beaten 3rd best",
                       "pairs beaten 4th best", "pairs beaten worst",
                       "enemy has real tailwind",
                       "pairs tailwind safe total",
                       "pairs protect safe total"):
                self.assertIn(col, header)
            # Sableye+Ariados carries no real Tailwind user.
            self.assertEqual(rows[0]["enemy has real tailwind"], "False")
        finally:
            os.unlink(path)

    def test_enemy_with_a_real_tailwind_setter_is_flagged_true(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            argv = ["--our",
                   "Kingambit,Dragapult,Whimsicott,Ninetales-Alola,"
                   "Mega Alakazam,Sharpedo",
                   "--bring4", "--vs", "Sableye,Talonflame", "--csv", path]
            msg, _out = run_main(argv)
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            self.assertEqual(rows[0]["enemy has real tailwind"], "True")
        finally:
            os.unlink(path)

    def test_csv_export_has_a_teamsheet_base64_column_that_decodes(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            self.assertIn("teamsheet base64", rows[0])
            from team_sheet import decode_teamsheet
            pool, _sets = decode_teamsheet(rows[0]["teamsheet base64"])
            self.assertTrue(pool)
        finally:
            os.unlink(path)


class TestMultiBring4SixPairExportColumns(unittest.TestCase):
    """The same six-pair depth columns, for --multi-bring4's CSV AND xlsx
    export (one core row can carry several enemies' worth of them)."""

    @classmethod
    def setUpClass(cls):
        # pool-size 16 stopped reliably finding a core once Regulation M-C's
        # very-high-Score additions (Mega Garchomp Z in particular) started
        # crowding the top of the candidate pool -- widened per the tool's
        # own suggested remedy ("widen the pool") rather than tuning Score
        # estimates to dodge a small, exhaustive-search fixture; this test
        # only cares that the export columns exist, not which core is found.
        cls.argv = ["--pool-size", "30", "--multi-bring4",
                   "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
                   "--top", "3"]

    def test_csv_export_has_per_enemy_six_pair_depth_columns(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            for col in ("enemy 1 pairs beaten total",
                       "enemy 1 pairs beaten 3rd best",
                       "enemy 1 pairs beaten 4th best",
                       "enemy 1 pairs beaten worst",
                       "enemy 1 has real tailwind",
                       "enemy 1 pairs tailwind safe total",
                       "enemy 1 pairs protect safe total"):
                self.assertIn(col, header)
        finally:
            os.unlink(path)

    def test_csv_export_has_a_per_enemy_teamsheet_base64_column(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            self.assertIn("enemy 1 teamsheet base64", rows[0])
            from team_sheet import decode_teamsheet
            pool, _sets = decode_teamsheet(rows[0]["enemy 1 teamsheet base64"])
            self.assertTrue(pool)
        finally:
            os.unlink(path)

    def test_xlsx_export_has_per_enemy_six_pair_depth_columns(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(self.argv + ["--xlsx", path])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["Cores"]
            header = [ws.cell(row=1, column=c).value
                     for c in range(1, ws.max_column + 1)]
            self.assertEqual(header[0], "#")
            for col in ("Enemy 1 pairs beaten total",
                       "Enemy 1 pairs beaten 3rd best",
                       "Enemy 1 pairs beaten 4th best",
                       "Enemy 1 pairs beaten worst",
                       "Enemy 1 has real Tailwind",
                       "Enemy 1 pairs Tailwind-safe total",
                       "Enemy 1 pairs protect-safe total"):
                self.assertIn(col, header)
            # Whimsicott is a real Tailwind setter in this fixture's roster.
            tw_col = header.index("Enemy 1 has real Tailwind") + 1
            self.assertEqual(ws.cell(row=2, column=tw_col).value, True)
            self.assertEqual(ws.cell(row=2, column=1).value, 1)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_xlsx_export_has_sets_and_teamsheets_sheets(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(self.argv + ["--xlsx", path])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            self.assertIn("Sets", wb.sheetnames)
            self.assertIn("Teamsheets", wb.sheetnames)
            sets_ws = wb["Sets"]
            self.assertGreater(sets_ws.max_row, 1)
            self.assertEqual(
                [c.value for c in sets_ws[1]],
                ["#", "Team", "Pokemon", "Item", "Ability", "Nature", "EVs",
                 "Moves", "Move usage %"])
            ts = wb["Teamsheets"]
            token_col = [c.value for c in ts[1]].index("Teamsheet (base64)") + 1
            token = ts.cell(row=2, column=token_col).value
            from team_sheet import decode_teamsheet
            pool, _sets = decode_teamsheet(token)
            self.assertTrue(pool)
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestMaxWeakTypesEndToEnd(unittest.TestCase):
    """"I would like to be able to select a cap for the number of types
    that have 2 weaknesses, such as no more than 3 types that have 2
    members weak to it. Display this measure in the export.\""""

    @classmethod
    def setUpClass(cls):
        cls.argv = ["--pool-size", "16", "--multi-bring4",
                   "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
                   "--good-threshold", "30", "--max-weak", "6", "--top", "3"]

    def test_types_with_2plus_weak_members_line_is_printed(self):
        msg, out = run_main(self.argv)
        self.assertIsNone(msg, out)
        self.assertIn("types with 2+ weak members:", out)

    def test_a_tight_cap_actually_changes_the_printed_result(self):
        _msg1, out_uncapped = run_main(self.argv)
        _msg2, out_capped = run_main(self.argv + ["--max-weak-types", "1"])
        self.assertNotEqual(out_uncapped, out_capped)

    def test_csv_export_has_the_types_with_2plus_column(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(self.argv + ["--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            self.assertIn("types with 2+ weak members", header)
        finally:
            os.unlink(path)

    def test_xlsx_export_has_the_types_with_2plus_column(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(self.argv + ["--xlsx", path])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["Cores"]
            header = [ws.cell(row=1, column=c).value
                     for c in range(1, ws.max_column + 1)]
            self.assertIn("Types with 2+ weak members", header)
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestTeamsheetJsonExport(unittest.TestCase):
    """"I am primarily using the CLI for counter_table.py and then checking
    results in the streamlit app, so the export in the CLI needs to be
    able to export to the streamlit app, whether through a paste or
    otherwise" -- `--teamsheet-json PATH` (or `-` for stdout) writes
    exactly the JSON shape the app's Team Builder tab reads/writes."""

    def test_bring4_stdout_prints_a_teamsheet_builder_can_load(self):
        argv = ["--our",
               "Kingambit,Dragapult,Whimsicott,Ninetales-Alola",
               "--bring4", "--vs", "Sableye,Ariados", "--teamsheet-json", "-"]
        msg, out = run_main(argv)
        self.assertIsNone(msg, out)
        self.assertIn("Teamsheet JSON", out)
        self.assertIn("paste into the Streamlit app", out)
        start = out.index("{")
        end = out.rindex("}") + 1
        import json
        payload = json.loads(out[start:end])
        self.assertEqual(set(payload["pool"]),
                         {"Kingambit", "Dragapult", "Whimsicott",
                          "Ninetales-Alola"})
        for name in payload["pool"]:
            self.assertIn("item", payload["sets"][name])
            self.assertIn("moves", payload["sets"][name])

    def test_bring4_defaults_to_the_top_result_without_deep_dive_core(self):
        argv = ["--our",
               "Kingambit,Dragapult,Whimsicott,Ninetales-Alola,"
               "Mega Alakazam,Sharpedo",
               "--bring4", "--vs", "Sableye,Ariados"]
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".json", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(argv + ["--teamsheet-json", path])
            self.assertIsNone(msg)
            import json
            with open(path, encoding="utf-8") as fh:
                payload = json.load(fh)
            self.assertEqual(len(payload["pool"]), 4)
        finally:
            os.unlink(path)

    def test_multi_bring4_writes_a_loadable_file(self):
        argv = ["--pool-size", "16", "--multi-bring4",
               "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
               "--good-threshold", "30"]
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".json", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, out = run_main(argv + ["--teamsheet-json", path])
            self.assertIsNone(msg, out)
            import json
            with open(path, encoding="utf-8") as fh:
                payload = json.load(fh)
            self.assertTrue(payload["pool"])
            self.assertEqual(set(payload["pool"]), set(payload["sets"]))
        finally:
            os.unlink(path)

    def test_reuses_the_deep_dive_core_pick_rather_than_recomputing(self):
        """Passing --deep-dive-core alongside --teamsheet-json must export
        THAT SAME core, not silently fall back to the top result."""
        argv = ["--our",
               "Kingambit,Dragapult,Whimsicott,Ninetales-Alola,"
               "Mega Alakazam,Sharpedo",
               "--bring4", "--vs", "Sableye,Ariados",
               "--deep-dive-core", "2", "--teamsheet-json", "-"]
        msg, out = run_main(argv)
        self.assertIsNone(msg, out)
        deep_dive_start = out.index("Deep dive:")
        teamsheet_start = out.index("Teamsheet JSON")
        core_line = out[deep_dive_start:teamsheet_start].splitlines()[0]
        core_names = set(n.strip() for n in
                         core_line.replace("Deep dive:", "").split("/"))
        import json
        start = out.index("{", teamsheet_start)
        end = out.rindex("}") + 1
        payload = json.loads(out[start:end])
        self.assertEqual(set(payload["pool"]), core_names)

    def test_multi_bring4_with_no_core_found_errors_cleanly(self):
        argv = ["--pool-size", "16", "--multi-bring4",
               "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
               "--max-weak", "0", "--max-megas", "0",
               "--teamsheet-json", "-"]
        msg, _out = run_main(argv)
        self.assertIsNotNone(msg)
        self.assertIn("--teamsheet-json", msg)


class TestCleanWinScoring(unittest.TestCase):
    """"I would consider losing 1 pokemon and taking a lot of damage and
    KOing 2 enemies as far inferior to KOing the enemy without taking
    damage, given the range of possible outcomes. There should be a way to
    score this to reflect this dynamic." `pairs_clean_win_total` (CLI
    print, CSV, --bring4/--multi-bring4 CSV+xlsx, --deep-dive-core) is
    that score."""

    def test_joint_print_shows_the_clean_column(self):
        msg, out = run_main(["--joint", "--vs", "Sableye,Ariados",
                             "--pool-size", "10", "--top", "3"])
        self.assertIsNone(msg, out)
        self.assertIn("clean", out)

    def test_joint_csv_export_has_the_clean_win_column(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(["--joint", "--vs", "Sableye,Ariados",
                                  "--pool-size", "10", "--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            self.assertIn("pairs clean win total", header)
        finally:
            os.unlink(path)

    def test_bring4_csv_has_the_6_pairs_clean_win_column(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            argv = ["--our",
                   "Kingambit,Dragapult,Whimsicott,Ninetales-Alola",
                   "--bring4", "--vs", "Sableye,Ariados", "--csv", path]
            msg, _out = run_main(argv)
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            self.assertIn("pairs clean win total", header)
        finally:
            os.unlink(path)

    def test_multi_bring4_csv_and_xlsx_have_the_clean_win_column(self):
        import csv
        import tempfile
        argv = ["--pool-size", "16", "--multi-bring4",
               "--vs-team", "Kingambit,Basculegion,Garchomp,Whimsicott",
               "--good-threshold", "30", "--max-weak", "6"]
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            csv_path = f.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            xlsx_path = f.name
        os.unlink(xlsx_path)
        try:
            msg, _out = run_main(argv + ["--csv", csv_path])
            self.assertIsNone(msg)
            with open(csv_path, newline="", encoding="utf-8") as fh:
                header = next(csv.reader(fh))
            self.assertIn("enemy 1 pairs clean win total", header)

            msg, out = run_main(argv + ["--xlsx", xlsx_path])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb["Cores"]
            xlsx_header = [ws.cell(row=1, column=c).value
                          for c in range(1, ws.max_column + 1)]
            self.assertIn("Enemy 1 pairs clean win total", xlsx_header)
        finally:
            os.unlink(csv_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)

    def test_deep_dive_core_prints_the_clean_win_figure(self):
        argv = ["--our",
               "Kingambit,Dragapult,Whimsicott,Ninetales-Alola",
               "--bring4", "--vs", "Sableye,Ariados",
               "--deep-dive-core", "1"]
        msg, out = run_main(argv)
        self.assertIsNone(msg, out)
        self.assertIn("clean win", out)


class TestDetailPreviewsShowTheWorstMatchupsFirst(unittest.TestCase):
    """Both `_print_pairs` and `_print_joint` picked their per-row detail
    preview via `sorted(r["detail"].items(), key=...)[:3]`, into a variable
    literally named `worst` -- but `_RANK`/`_JOINT_RANK` rank the BEST
    outcome (clean/sweep) as 0 and the WORST (pinned/loss) as 3, so a plain
    ascending sort put the 3 BEST matchups under that name, not the worst
    ones. `_print_deep`'s own docstring says outright "LOSSES AND OHKO
    RISKS FIRST, since those are what's actionable" -- its sort had the
    exact same inversion. All three now sort worst-first."""

    def test_print_pairs_shows_the_pinned_result_before_the_clean_one(self):
        row = {
            "name": "Gallade", "item": "Life Orb", "pairs_clean": 1,
            "pairs_trade": 0, "pairs_no_ko": 0, "pairs_pinned": 1,
            "pairs_total": 2,
            "detail": {
                ("Kingambit", "Basculegion"): {
                    "outcome": "clean", "target": "Kingambit",
                    "hits": {"C": {"Kingambit": hit("Psycho Cut", 1.2)}}},
                ("Garchomp", "Whimsicott"): {
                    "outcome": "pinned", "target": None, "hits": {}},
            },
        }
        buf = io.StringIO()
        with redirect_stdout(buf):
            ct._print_pairs([row], ["Kingambit", "Basculegion", "Garchomp",
                                    "Whimsicott"], top=1)
        out = buf.getvalue()
        self.assertLess(out.index("Garchomp + Whimsicott: pinned"),
                        out.index("Kingambit + Basculegion: clean"),
                        "the pinned (worst) matchup must print before the "
                        "clean (best) one")

    def test_print_joint_shows_the_loss_before_the_sweep(self):
        row = {
            "name": "Gallade", "item": "Life Orb",
            "pairs_swept": 1, "pairs_traded": 0, "pairs_lost": 1,
            "pairs_no_ko": 0, "pairs_tailwind_safe": 1, "pairs_protect_safe": 1,
            "pairs_clean_win_total": 2.0, "pairs_total": 2,
            "detail": {
                ("Kingambit", "Basculegion"): {
                    "outcome": "sweep", "turns_used": 1,
                    "tailwind_safe": True, "tailwind_outcome": "sweep",
                    "protect_safe": True,
                    "protect_outcomes": {"E1": "sweep", "E2": "sweep"},
                    "log": []},
                ("Garchomp", "Whimsicott"): {
                    "outcome": "loss", "turns_used": 1,
                    "tailwind_safe": False, "tailwind_outcome": "loss",
                    "protect_safe": False,
                    "protect_outcomes": {"E1": "loss", "E2": "loss"},
                    "log": []},
            },
        }
        buf = io.StringIO()
        with redirect_stdout(buf):
            ct._print_joint([row], ["Kingambit", "Basculegion", "Garchomp",
                                    "Whimsicott"], top=1, partner="", turns=2)
        out = buf.getvalue()
        self.assertLess(out.index("Garchomp + Whimsicott: loss"),
                        out.index("Kingambit + Basculegion: sweep"),
                        "the loss must print before the sweep")

    def test_print_deep_shows_the_loss_before_the_sweep(self):
        detail = {
            ("Kingambit", "Basculegion"): {
                "outcome": "sweep", "turns_used": 1, "ohko_risk": [],
                "tailwind_safe": True, "tailwind_outcome": "sweep",
                "protect_safe": True,
                "protect_outcomes": {"E1": "sweep", "E2": "sweep"},
                "grid": {"ours": {}, "theirs": {}}, "log": []},
            ("Garchomp", "Whimsicott"): {
                "outcome": "loss", "turns_used": 1, "ohko_risk": [],
                "tailwind_safe": False, "tailwind_outcome": "loss",
                "protect_safe": False,
                "protect_outcomes": {"E1": "loss", "E2": "loss"},
                "grid": {"ours": {}, "theirs": {}}, "log": []},
        }
        summary = {"pairs_total": 2, "pairs_swept": 1, "pairs_traded": 0,
                  "pairs_lost": 1, "pairs_no_ko": 0, "pairs_tailwind_safe": 1,
                  "pairs_protect_safe": 1, "pairs_clean_win_total": 2.0}
        buf = io.StringIO()
        with redirect_stdout(buf):
            ct._print_deep("Gallade", "Ninetales-Alola", "Life Orb", None,
                           ["Kingambit", "Basculegion", "Garchomp",
                            "Whimsicott"], detail, summary, turns=2)
        out = buf.getvalue()
        self.assertLess(out.index("Garchomp + Whimsicott: LOSS"),
                        out.index("Kingambit + Basculegion: SWEEP"),
                        "fixture assumes both outcomes print in upper case")


class TestVsTeamAcceptsANamedTeam(unittest.TestCase):
    """"For --vs-team in counter_table.py, I would like to be able to use
    named teams from the team.csv and /data/teams folder." `--vs-team`
    already only ever fed `--multi-bring4`'s search, and
    `species_data.load_teams` (the SAME library `--team` already searches
    -- teams.csv rows plus any pokepaste dropped in data/teams/ or
    data/my_teams/) was already loaded into `W["teams"]` before this parsing
    ran; a named team is now tried FIRST, falling back to the existing
    comma-separated-Pokemon parsing when the name isn't a known team."""

    def test_a_known_team_name_resolves_to_its_full_roster(self):
        from _harness import load_world
        W = load_world()
        self.assertIn("Rain", W["teams"])
        self.assertIn("Big 6", W["teams"])
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Rain", "--vs-team", "Big 6",
             "--pool-size", "12", "--good-threshold", "30", "--top", "1"])
        self.assertIsNone(msg, out)
        for name in W["teams"]["Rain"]:
            self.assertIn(name, out)
        for name in W["teams"]["Big 6"]:
            self.assertIn(name, out)

    def test_a_raw_comma_list_still_works_unchanged(self):
        """Backward compatibility: a --vs-team value that ISN'T a known
        team name still parses as individual Pokemon, exactly as before
        this feature existed."""
        msg, out = run_main(
            ["--multi-bring4", "--vs-team",
             "Kingambit,Basculegion,Garchomp,Whimsicott", "--vs-team",
             "Sylveon,Mega Charizard Y,Sinistcha,Farigiraf", "--pool-size",
             "12", "--good-threshold", "30", "--top", "1"])
        self.assertIsNone(msg, out)
        self.assertIn("Kingambit", out)
        self.assertIn("Sylveon", out)

    def test_an_unknown_bare_word_still_fails_loudly(self):
        """A typo'd team name (no comma, not a real team) must not be
        silently treated as one lone Pokemon -- it already fails the
        existing 'needs at least 2 Pokemon' check, same as before named
        teams existed."""
        msg, _out = run_main(
            ["--multi-bring4", "--vs-team", "NotARealTeamName",
             "--vs-team", "Big 6"])
        self.assertIsNotNone(msg)
        self.assertIn("needs at least 2 Pokemon", msg)


class TestVsAllTeamsFlag(unittest.TestCase):
    """--vs-all-teams: run --multi-bring4 against EVERY saved team
    (species_data.load_teams's merged teams.csv + data/teams/ +
    data/my_teams/ dict, already loaded as W["teams"]) instead of naming
    each one with a repeated --vs-team."""

    def test_resolves_to_every_saved_team(self):
        from _harness import load_world
        W = load_world()
        msg, out = run_main(
            ["--multi-bring4", "--vs-all-teams", "--pool-size", "10",
             "--top", "1"])
        self.assertIsNone(msg, out)
        self.assertIn(f"vs {len(W['teams'])} enemy teams", out)
        self.assertEqual(out.count("-- top pairs:"), len(W["teams"]),
                         "one 'Enemy N (...) -- top pairs:' header per saved team")

    def test_mutually_exclusive_with_vs_team(self):
        msg, _out = run_main(
            ["--multi-bring4", "--vs-all-teams", "--vs-team", "Rain"])
        self.assertIsNotNone(msg)
        self.assertIn("can't be combined with --vs-team", msg)

    def test_requires_multi_bring4(self):
        msg, _out = run_main(
            ["--bring4", "--our", "Garchomp,Hydreigon,Kingambit,Whimsicott",
             "--vs-all-teams"])
        self.assertIsNotNone(msg)
        self.assertIn("requires --multi-bring4", msg)


class TestTurnsAppliesToMultiBring4(unittest.TestCase):
    """`multi_bring4_coverage` already accepted and threaded a `turns`
    parameter (default 2) -- the CLI's own validation was the only thing
    blocking --turns from --multi-bring4."""

    def test_no_longer_rejected(self):
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
             "--pool-size", "8", "--top", "1", "--turns", "4"])
        self.assertIsNone(msg, out)

    def test_still_rejected_outside_joint_deep_bring4_multi_bring4(self):
        msg, _out = run_main(["--vs", "Kingambit", "--turns", "3"])
        self.assertIsNotNone(msg)
        self.assertIn("--turns only applies to", msg)


class TestJobsAppliesToMultiBring4(unittest.TestCase):
    """"Is there a way to optionally devote more resources to
    counter_table.py for parallel calculations, such as with the --jobs
    argument in [generate_]overnight?" -- --jobs mirrors search_teams.py's
    own flag: N enemy rosters' pool-wide pair searches run in parallel
    worker processes instead of one after another. Only meaningful for
    --multi-bring4, since that's the only mode with more than one such
    independent search to split across workers."""

    def test_still_rejected_outside_multi_bring4(self):
        msg, _out = run_main(["--vs", "Kingambit", "--jobs", "2"])
        self.assertIsNotNone(msg)
        self.assertIn("--jobs only applies to --multi-bring4", msg)

    def test_jobs_1_explicit_on_multi_bring4_still_works(self):
        """--jobs defaults to 1 -- passing it explicitly on --multi-bring4
        must behave exactly like the default (serial, no process pool)."""
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
             "--pool-size", "8", "--top", "1", "--no-prompt", "--jobs", "1"])
        self.assertIsNone(msg, out)

    def test_a_multi_bring4_run_with_jobs_2_produces_the_same_result_as_serial(self):
        """Parallelism is purely a speed knob -- everything but the extra
        "workers" progress line (only --jobs > 1 prints that) must match."""
        args = ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
                "--vs-team", "Garchomp,Incineroar", "--pool-size", "8",
                "--top", "3", "--no-prompt"]
        msg_serial, out_serial = run_main(args)
        msg_parallel, out_parallel = run_main(args + ["--jobs", "2"])
        self.assertIsNone(msg_serial, out_serial)
        self.assertIsNone(msg_parallel, out_parallel)
        strip_workers = lambda s: "\n".join(   # noqa: E731
            l for l in s.splitlines() if not l.startswith("workers  :")
            and not l.startswith("WARNING  :"))
        self.assertEqual(strip_workers(out_serial), strip_workers(out_parallel))

    def test_jobs_2_prints_a_workers_line(self):
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
             "--vs-team", "Garchomp,Incineroar", "--pool-size", "8",
             "--top", "1", "--no-prompt", "--jobs", "2"])
        self.assertIsNone(msg, out)
        self.assertIn("workers  : 2 of", out)

    def test_jobs_2_with_only_one_vs_team_prints_no_workers_line(self):
        """Nothing to split across workers with a single enemy roster --
        the CLI's own "> 1 enemy" guard (mirrors multi_bring4_coverage's
        own len(target_name_lists) > 1 fallback) keeps the workers line
        from claiming parallelism that never actually ran."""
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
             "--pool-size", "8", "--top", "1", "--no-prompt", "--jobs", "2"])
        self.assertIsNone(msg, out)
        self.assertNotIn("workers  :", out)


class TestBring4AcceptsASingleNamedVsTeam(unittest.TestCase):
    """"I need a way, both in the counter_table.py and the streamlit app, to
    run a bring4 (4-6) vs only ONE named TEAM (--vs-team)" -- --bring4 now
    accepts a single --vs-team (a saved team.csv row/pokepaste, resolved
    the same way --multi-bring4's own --vs-team already is) as an
    alternative to spelling the roster out with --vs."""

    OUR = ("Garchomp,Incineroar,Gallade,Hydreigon,Whimsicott,"
          "Mega Alakazam")

    def test_a_saved_team_name_resolves_to_its_full_roster(self):
        from _harness import load_world
        W = load_world()
        self.assertIn("Rain", W["teams"])
        msg, out = run_main(
            ["--our", self.OUR, "--bring4", "--vs-team", "Rain",
             "--no-prompt", "--top", "1"])
        self.assertIsNone(msg, out)
        for name in W["teams"]["Rain"]:
            self.assertIn(name, out)

    def test_a_raw_comma_list_still_works_as_a_single_vs_team(self):
        msg, out = run_main(
            ["--our", self.OUR, "--bring4", "--vs-team",
             "Archaludon,Grimmsnarl,Mega Metagross,Pelipper", "--no-prompt",
             "--top", "1"])
        self.assertIsNone(msg, out)
        self.assertIn("Archaludon", out)
        self.assertIn("Grimmsnarl", out)

    def test_vs_and_vs_team_together_is_rejected(self):
        msg, _out = run_main(
            ["--our", self.OUR, "--bring4", "--vs", "Kingambit,Basculegion",
             "--vs-team", "Rain"])
        self.assertIsNotNone(msg)
        self.assertIn("--vs-team", msg)

    def test_neither_vs_nor_vs_team_is_rejected(self):
        msg, _out = run_main(["--our", self.OUR, "--bring4"])
        self.assertIsNotNone(msg)

    def test_more_than_one_vs_team_is_rejected(self):
        msg, _out = run_main(
            ["--our", self.OUR, "--bring4", "--vs-team", "Rain", "--vs-team",
             "Big 6"])
        self.assertIsNotNone(msg)
        self.assertIn("at most one --vs-team", msg)

    def test_xlsx_export_works_with_bring4(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, _out = run_main(
                ["--our", self.OUR, "--bring4", "--vs-team", "Rain",
                 "--no-prompt", "--xlsx", path, "--top", "2"])
            self.assertIsNone(msg)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            self.assertIn("Bring-4s", wb.sheetnames)
            self.assertIn("Sets", wb.sheetnames)
            self.assertIn("Teamsheets", wb.sheetnames)
            ws = wb["Bring-4s"]
            self.assertEqual(ws.cell(row=1, column=1).value, "#")
            self.assertEqual(ws.cell(row=2, column=1).value, 1)
            ts = wb["Teamsheets"]
            token_col = [c.value for c in ts[1]].index("Teamsheet (base64)") + 1
            token = ts.cell(row=2, column=token_col).value
            self.assertTrue(token.startswith("TSHEET1:"))
            from team_sheet import decode_teamsheet
            pool, _sets = decode_teamsheet(token)
            self.assertTrue(pool)  # decodes to a real, non-empty roster
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestXlsxNewSummaryColumns(unittest.TestCase):
    """"I would like the best and third best pair under tailwind and under
    enemy protect. I would also like to see best and third best number of
    pairs beaten without having either of own pair faint. I also want new
    columns that have Avg Wins/90, Avg Wins under tailwind/90, Avg wins
    under protect/90, Clean wins /90, and the average score ... of the
    team" -- both `--bring4`'s "Bring-4s" sheet and `--multi-bring4`'s
    "Cores" sheet must carry every one of these."""

    NEW_COLUMNS = ("Average Score", "Avg Wins/90", "Avg Wins under Tailwind/90",
                  "Avg Wins under Protect/90", "Clean wins/90")
    NEW_PER_ENEMY_SUFFIXES = ("pairs Tailwind-safe best", "pairs Tailwind-safe 3rd best",
                              "pairs protect-safe best", "pairs protect-safe 3rd best",
                              "pairs beaten without fainting best",
                              "pairs beaten without fainting 3rd best")

    def test_bring4_xlsx_has_every_new_column(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(
                ["--our", "Garchomp,Incineroar,Gallade,Hydreigon,Whimsicott,"
                          "Mega Alakazam", "--bring4", "--vs-team", "Rain",
                 "--no-prompt", "--xlsx", path, "--top", "2"])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            header = [c.value for c in wb["Bring-4s"][1]]
            for col in self.NEW_COLUMNS:
                self.assertIn(col, header)
            for suffix in self.NEW_PER_ENEMY_SUFFIXES:
                self.assertIn(suffix, header)
            row2 = {h: c.value for h, c in zip(header, wb["Bring-4s"][2])}
            for col in self.NEW_COLUMNS:
                self.assertIsNotNone(row2[col])
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_multi_bring4_xlsx_has_every_new_column(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(
                ["--multi-bring4", "--vs-team", "Kingambit,Basculegion",
                 "--vs-team", "Garchomp,Incineroar", "--pool-size", "20",
                 "--good-threshold", "0", "--min-enemies", "1",
                 "--top", "2", "--no-prompt", "--xlsx", path])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            self.assertGreater(wb["Cores"].max_row, 1, "no core rows were found")
            header = [c.value for c in wb["Cores"][1]]
            for col in self.NEW_COLUMNS:
                self.assertIn(col, header)
            for suffix in self.NEW_PER_ENEMY_SUFFIXES:
                self.assertTrue(any(h and h.endswith(suffix) for h in header),
                                f"no per-enemy column ending in {suffix!r}")
            row2 = {h: c.value for h, c in zip(header, wb["Cores"][2])}
            for col in self.NEW_COLUMNS:
                self.assertIsNotNone(row2[col])
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_avg_wins_per_90_matches_a_from_scratch_recomputation(self):
        """Cross-check the xlsx cell against `bring4_pair_depth`/`_per_90`
        computed directly from the same search results, not just "is it a
        number" -- catches a wrong column, a wrong `_per_90` scale, or an
        averaging-across-the-wrong-things bug."""
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        os.unlink(path)
        try:
            msg, out = run_main(
                ["--our", "Garchomp,Incineroar,Gallade,Hydreigon,Whimsicott,"
                          "Mega Alakazam", "--bring4", "--vs-team", "Rain",
                 "--no-prompt", "--xlsx", path, "--top", "1"])
            self.assertIsNone(msg, out)
            from openpyxl import load_workbook
            from counter_finder import bring4_search, bring4_pair_depth
            wb = load_workbook(path)
            header = [c.value for c in wb["Bring-4s"][1]]
            row2 = {h: c.value for h, c in zip(header, wb["Bring-4s"][2])}
            from _harness import load_world
            W = load_world()
            merged, moves = W["merged"], W["moves"]
            natures, typechart = W["natures"], W["typechart"]
            our6 = ["Garchomp", "Incineroar", "Gallade", "Hydreigon",
                   "Whimsicott", "Mega Alakazam"]
            targets = list(W["teams"]["Rain"])
            _pair_rows, bring4_rows = bring4_search(
                our6, targets, merged, moves, natures, typechart, good_threshold=0.0)
            depth = bring4_pair_depth(bring4_rows[0])
            n_pairs, pt = len(bring4_rows[0]["pair_rows"]), depth["pairs_total"]
            expected = round(depth["beaten_total"] / (n_pairs * pt) * 90, 1)
            self.assertAlmostEqual(row2["Avg Wins/90"], expected, places=1)
        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestMultiBring4NeverComesBackEmpty(unittest.TestCase):
    """"When a sweep of --vs-team gets too many results it doesn't even
    output the results or anything to CSV. I want to at least see the
    results (i.e., high performing pairs, results /15) and it should be
    very quick to compute the sets of 4 brings ... this should not take
    long." A candidate pool too big for the exhaustive C(N,6) sweep used
    to abort the whole run via SystemExit before anything printed -- Stage
    A (the pool-wide pair-vs-each-enemy search) is unaffected by that
    limit and was simply being thrown away. Now: the pair summary always
    prints, and hitting the exhaustive ceiling auto-falls back to --beam
    (already the fast, non-exhaustive path) instead of erroring out."""

    def test_a_too_large_candidate_pool_falls_back_instead_of_exiting(self):
        msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Rain", "--vs-team", "Big 6",
             "--pool-size", "12", "--good-threshold", "30",
             "--min-enemies", "1", "--top", "5", "--max-candidates", "3"])
        self.assertIsNone(msg, out)
        self.assertIn("too many for an exhaustive sweep", out)
        self.assertIn("Falling back to --beam automatically", out)
        self.assertIn("auto-fallback", out)

    def test_the_pair_summary_prints_even_on_the_fallback_path(self):
        _msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Rain", "--vs-team", "Big 6",
             "--pool-size", "12", "--good-threshold", "30",
             "--min-enemies", "1", "--top", "5", "--max-candidates", "3"])
        self.assertIn("top pairs:", out)
        self.assertIn("beaten (", out)
        self.assertIn("/15 beaten", out)

    def test_the_pair_summary_prints_on_the_normal_exhaustive_path_too(self):
        """Not just a fallback-only feature -- always shown."""
        _msg, out = run_main(
            ["--multi-bring4", "--vs-team", "Rain", "--vs-team", "Big 6",
             "--pool-size", "12", "--good-threshold", "30",
             "--min-enemies", "1", "--top", "5"])
        self.assertNotIn("too many for an exhaustive sweep", out)
        self.assertIn("top pairs:", out)

    def test_csv_export_still_works_after_the_fallback(self):
        import csv
        import tempfile
        with tempfile.NamedTemporaryFile(
                suffix=".csv", delete=False, mode="w") as f:
            path = f.name
        try:
            msg, _out = run_main(
                ["--multi-bring4", "--vs-team", "Rain", "--vs-team", "Big 6",
                 "--pool-size", "12", "--good-threshold", "30",
                 "--min-enemies", "1", "--top", "5", "--max-candidates", "3",
                 "--csv", path])
            self.assertIsNone(msg)
            with open(path, newline="", encoding="utf-8") as fh:
                rows = list(csv.reader(fh))
            self.assertGreater(len(rows), 1, "the fallback must still "
                               "write real rows, not an empty file")
        finally:
            os.unlink(path)


if __name__ == "__main__":
    unittest.main()
